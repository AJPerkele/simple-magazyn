import sqlite3
import csv
import os
from datetime import datetime, timedelta
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

class DB:
    def __init__(self, path="data.db"):
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init()

    def _init(self):
        c = self.conn.cursor()
        c.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT UNIQUE,
            title TEXT,
            stock INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS purchase_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            total_pln REAL,
            date TEXT
        );

        CREATE TABLE IF NOT EXISTS purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            product_id INTEGER,
            qty INTEGER,
            unit_cost REAL,
            available_qty INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS purchase_stock_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_item_id INTEGER,
            product_id INTEGER,
            qty INTEGER,
            date TEXT,
            sale_order_id INTEGER DEFAULT NULL
        );

        CREATE TABLE IF NOT EXISTS sales_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT,
            total_pln REAL,
            total_eur REAL,
            purchase_cost REAL DEFAULT 0,
            date TEXT
        );

        CREATE TABLE IF NOT EXISTS sales_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            product_id INTEGER,
            qty INTEGER
        );

        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT UNIQUE,
            sale_order_id INTEGER,
            file_path TEXT,
            customer_name TEXT,
            customer_address TEXT,
            issue_date TEXT,
            total_amount REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sale_order_id) REFERENCES sales_orders(id) ON DELETE SET NULL
        );
        """)
        
        try:
            c.execute("SELECT purchase_cost FROM sales_orders LIMIT 1")
        except sqlite3.OperationalError:
            c.execute("ALTER TABLE sales_orders ADD COLUMN purchase_cost REAL DEFAULT 0")
            
        try:
            c.execute("SELECT available_qty FROM purchase_items LIMIT 1")
        except sqlite3.OperationalError:
            c.execute("ALTER TABLE purchase_items ADD COLUMN available_qty INTEGER DEFAULT 0")
            
        try:
            c.execute("SELECT unit_cost FROM purchase_items LIMIT 1")
        except sqlite3.OperationalError:
            c.execute("ALTER TABLE purchase_items ADD COLUMN unit_cost REAL DEFAULT 0")
            
        c.execute("""
            CREATE TABLE IF NOT EXISTS purchase_stock_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                purchase_item_id INTEGER,
                product_id INTEGER,
                qty INTEGER,
                date TEXT,
                sale_order_id INTEGER DEFAULT NULL
            )
        """)
        
        self.conn.commit()

    # ---------- PRODUCTS ----------
    def add_product(self, sku, title):
        self.conn.execute(
            "INSERT INTO products(sku,title,stock) VALUES(?,?,0)",
            (sku, title)
        )
        self.conn.commit()

    def check_sku_exists(self, sku):
        r = self.conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()
        return r is not None

    def get_product_id_by_sku(self, sku):
        r = self.conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()
        return r["id"] if r else None

    def delete_product(self, pid):
        c = self.conn.cursor()
        
        product = self.get_product_info(pid)
        if product and product["stock"] > 0:
            return False
            
        c.execute("DELETE FROM purchase_items WHERE product_id=?", (pid,))
        c.execute("DELETE FROM sales_items WHERE product_id=?", (pid,))
        c.execute("DELETE FROM purchase_stock_history WHERE product_id=?", (pid,))
        c.execute("DELETE FROM invoices WHERE sale_order_id IN (SELECT id FROM sales_orders WHERE id IN (SELECT order_id FROM sales_items WHERE product_id=?))", (pid,))
        c.execute("DELETE FROM products WHERE id=?", (pid,))
        
        self.conn.commit()
        return True

    def list_products(self):
        return self.conn.execute("SELECT * FROM products").fetchall()

    def update_stock(self, pid, delta):
        self.conn.execute(
            "UPDATE products SET stock = stock + ? WHERE id=?",
            (delta, pid)
        )
        self.conn.commit()

    def check_stock(self, pid: int, required_qty: int) -> bool:
        r = self.conn.execute("SELECT stock FROM products WHERE id=?", (pid,)).fetchone()
        if not r:
            return False
        return r["stock"] >= required_qty

    def get_product_info(self, pid: int):
        return self.conn.execute("SELECT sku, title, stock FROM products WHERE id=?", (pid,)).fetchone()

    # ---------- PURCHASES ----------
    def add_purchase_order(self, total_pln, date, items):
        c = self.conn.cursor()
        
        max_id_result = c.execute("SELECT MAX(id) as max_id FROM purchase_orders").fetchone()
        max_id = max_id_result["max_id"] or 0
        
        all_ids = [row[0] for row in c.execute("SELECT id FROM purchase_orders ORDER BY id").fetchall()]
        new_id = 1
        for existing_id in all_ids:
            if existing_id == new_id:
                new_id += 1
            else:
                break
        
        if new_id > max_id:
            new_id = max_id + 1
        
        c.execute(
            "INSERT INTO purchase_orders(id,total_pln,date) VALUES(?,?,?)",
            (new_id, total_pln, date)
        )
        oid = new_id

        total_qty = sum(qty for _, qty in items)
        
        for pid, qty in items:
            unit_cost = (total_pln / total_qty) if total_qty > 0 else 0
            
            max_pi_id = c.execute("SELECT MAX(id) as max_id FROM purchase_items").fetchone()
            max_pi_id_val = max_pi_id["max_id"] or 0
            
            pi_ids = [row[0] for row in c.execute("SELECT id FROM purchase_items ORDER BY id").fetchall()]
            new_pi_id = 1
            for existing_id in pi_ids:
                if existing_id == new_pi_id:
                    new_pi_id += 1
                else:
                    break
            
            if new_pi_id > max_pi_id_val:
                new_pi_id = max_pi_id_val + 1
            
            c.execute("""
                INSERT INTO purchase_items(id,order_id,product_id,qty,unit_cost,available_qty)
                VALUES(?,?,?,?,?,?)
            """, (new_pi_id, oid, pid, qty, unit_cost, qty))
            
            purchase_item_id = new_pi_id
            
            max_psh_id = c.execute("SELECT MAX(id) as max_id FROM purchase_stock_history").fetchone()
            max_psh_id_val = max_psh_id["max_id"] or 0
            
            for i in range(qty):
                psh_ids = [row[0] for row in c.execute("SELECT id FROM purchase_stock_history ORDER BY id").fetchall()]
                new_psh_id = 1
                for existing_id in psh_ids:
                    if existing_id == new_psh_id:
                        new_psh_id += 1
                    else:
                        break
                
                if new_psh_id > max_psh_id_val:
                    new_psh_id = max_psh_id_val + 1
                    max_psh_id_val = new_psh_id
                
                c.execute("""
                    INSERT INTO purchase_stock_history(id,purchase_item_id,product_id,qty,date)
                    VALUES(?,?,?,?,?)
                """, (new_psh_id, purchase_item_id, pid, 1, date))
            
            product = self.conn.execute("SELECT id FROM products WHERE id=?", (pid,)).fetchone()
            if product:
                self.update_stock(pid, qty)

        self.conn.commit()
        return oid

    def list_purchases(self):
        try:
            rows = self.conn.execute("""
                SELECT
                    o.id,
                    pr.sku,
                    pr.title,
                    i.qty,
                    ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM purchase_items 
                        WHERE order_id = o.id
                    ) * i.qty, 2) as item_cost,
                    o.date
                FROM purchase_orders o
                JOIN purchase_items i ON i.order_id = o.id
                JOIN products pr ON pr.id = i.product_id
                ORDER BY o.date DESC, o.id DESC
            """).fetchall()
            
            result = []
            for r in rows:
                result.append((
                    r["id"],
                    r["sku"],
                    r["title"],
                    r["qty"],
                    r["item_cost"],
                    r["date"]
                ))
            return result
        except Exception as e:
            print(f"Błąd w list_purchases: {e}")
            return []

    def get_detailed_purchases(self, date_from, date_to):
        try:
            rows = self.conn.execute("""
                SELECT
                    o.id as order_id,
                    o.date,
                    pr.sku,
                    pr.title,
                    i.qty,
                    ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM purchase_items 
                        WHERE order_id = o.id
                    ) * i.qty, 2) as item_cost,
                    i.unit_cost,
                    o.total_pln as order_total
                FROM purchase_orders o
                JOIN purchase_items i ON i.order_id = o.id
                JOIN products pr ON pr.id = i.product_id
                WHERE o.date BETWEEN ? AND ?
                ORDER BY o.date, o.id
            """, (date_from, date_to)).fetchall()
            
            result = []
            for r in rows:
                result.append({
                    'order_id': r['order_id'],
                    'date': r['date'],
                    'sku': r['sku'],
                    'title': r['title'],
                    'qty': r['qty'],
                    'item_cost': r['item_cost'],
                    'unit_cost': r['unit_cost'],
                    'order_total': r['order_total']
                })
            return result
        except Exception as e:
            print(f"Błąd w get_detailed_purchases: {e}")
            return []

    def delete_purchase(self, oid):
        try:
            c = self.conn.cursor()
            
            rows = c.execute("""
                SELECT product_id, qty 
                FROM purchase_items 
                WHERE order_id=?
            """, (oid,)).fetchall()

            for r in rows:
                self.update_stock(r["product_id"], -r["qty"])

            c.execute("""
                DELETE FROM purchase_stock_history 
                WHERE purchase_item_id IN (
                    SELECT id FROM purchase_items WHERE order_id=?
                )
            """, (oid,))
            
            c.execute("DELETE FROM purchase_items WHERE order_id=?", (oid,))
            c.execute("DELETE FROM purchase_orders WHERE id=?", (oid,))
            
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Błąd w delete_purchase: {e}")
            return False

    # ---------- FIFO FUNCTIONS ----------
    def get_fifo_batches(self, pid: int, required_qty: int):
        try:
            c = self.conn.cursor()
            c.execute("SELECT 1 FROM purchase_stock_history LIMIT 1")
        except sqlite3.OperationalError:
            return []
        
        batches = self.conn.execute("""
            SELECT 
                pi.id as purchase_item_id,
                pi.unit_cost,
                pi.available_qty,
                po.date,
                COUNT(psh.id) as available_units
            FROM purchase_items pi
            JOIN purchase_orders po ON po.id = pi.order_id
            LEFT JOIN purchase_stock_history psh ON psh.purchase_item_id = pi.id 
                AND psh.product_id = pi.product_id 
                AND psh.sale_order_id IS NULL
            WHERE pi.product_id = ? AND pi.available_qty > 0
            GROUP BY pi.id
            HAVING available_units > 0
            ORDER BY po.date ASC, pi.id ASC
        """, (pid,)).fetchall()
        
        result = []
        remaining_qty = required_qty
        
        for batch in batches:
            if remaining_qty <= 0:
                break
                
            available = batch["available_units"]
            take_qty = min(available, remaining_qty)
            
            result.append({
                "purchase_item_id": batch["purchase_item_id"],
                "unit_cost": batch["unit_cost"],
                "available_qty": take_qty,
                "date": batch["date"]
            })
            
            remaining_qty -= take_qty
        
        return result

    def allocate_fifo_stock(self, pid: int, qty: int, sale_order_id: int):
        batches = self.get_fifo_batches(pid, qty)
        total_cost = 0.0
        
        for batch in batches:
            batch_qty = batch["available_qty"]
            unit_cost = batch["unit_cost"]
            
            self.conn.execute("""
                UPDATE purchase_stock_history 
                SET sale_order_id = ?
                WHERE purchase_item_id = ? 
                AND product_id = ?
                AND sale_order_id IS NULL
                LIMIT ?
            """, (sale_order_id, batch["purchase_item_id"], pid, batch_qty))
            
            self.conn.execute("""
                UPDATE purchase_items 
                SET available_qty = available_qty - ?
                WHERE id = ?
            """, (batch_qty, batch["purchase_item_id"]))
            
            total_cost += unit_cost * batch_qty
        
        self.conn.commit()
        return total_cost

    # ---------- SALES ----------
    def add_sale_order_with_reset(self, platform, total_pln, total_eur, date, items, purchase_cost=0):
        c = self.conn.cursor()
        
        max_id_result = c.execute("SELECT MAX(id) as max_id FROM sales_orders").fetchone()
        max_id = max_id_result["max_id"] or 0
        
        all_ids = [row[0] for row in c.execute("SELECT id FROM sales_orders ORDER BY id").fetchall()]
        
        new_id = 1
        for existing_id in all_ids:
            if existing_id == new_id:
                new_id += 1
            else:
                break
        
        if new_id > max_id:
            new_id = max_id + 1
        
        c.execute("""
            INSERT INTO sales_orders(id,platform,total_pln,total_eur,purchase_cost,date)
            VALUES(?,?,?,?,?,?)
        """, (new_id, platform, total_pln, total_eur, purchase_cost, date))
        sale_order_id = new_id

        total_purchase_cost = 0
        
        for pid, qty in items:
            max_si_id = c.execute("SELECT MAX(id) as max_id FROM sales_items").fetchone()
            max_si_id_val = max_si_id["max_id"] or 0
            
            si_ids = [row[0] for row in c.execute("SELECT id FROM sales_items ORDER BY id").fetchall()]
            new_si_id = 1
            for existing_id in si_ids:
                if existing_id == new_si_id:
                    new_si_id += 1
                else:
                    break
            
            if new_si_id > max_si_id_val:
                new_si_id = max_si_id_val + 1
            
            c.execute("""
                INSERT INTO sales_items(id,order_id,product_id,qty)
                VALUES(?,?,?,?)
            """, (new_si_id, sale_order_id, pid, qty))
            
            try:
                item_cost = self.allocate_fifo_stock(pid, qty, sale_order_id)
                total_purchase_cost += item_cost
            except Exception as e:
                print(f"Błąd alokacji FIFO: {e}")
                total_purchase_cost += 0
            
            self.update_stock(pid, -qty)

        if total_purchase_cost > 0:
            c.execute("""
                UPDATE sales_orders 
                SET purchase_cost = ?
                WHERE id = ?
            """, (total_purchase_cost, sale_order_id))

        self.conn.commit()
        return sale_order_id

    def add_sale_order(self, platform, total_pln, total_eur, date, items, purchase_cost=0):
        return self.add_sale_order_with_reset(platform, total_pln, total_eur, date, items, purchase_cost)

    def list_sales(self):
        try:
            rows = self.conn.execute("""
                SELECT
                    o.id,
                    o.platform,
                    o.total_pln,
                    o.total_eur,
                    COALESCE(o.purchase_cost, 0) as purchase_cost,
                    o.date,
                    GROUP_CONCAT(pr.sku || ' x' || si.qty, ', ') as items
                FROM sales_orders o
                LEFT JOIN sales_items si ON si.order_id = o.id
                LEFT JOIN products pr ON pr.id = si.product_id
                GROUP BY o.id, o.platform, o.total_pln, o.total_eur, o.purchase_cost, o.date
                ORDER BY o.date DESC, o.id DESC
            """).fetchall()
        except sqlite3.OperationalError as e:
            rows = self.conn.execute("""
                SELECT
                    o.id,
                    o.platform,
                    o.total_pln,
                    o.total_eur,
                    COALESCE(o.purchase_cost, 0) as purchase_cost,
                    o.date,
                    GROUP_CONCAT(pr.sku || ' x' || si.qty, ', ') as items
                FROM sales_orders o
                LEFT JOIN sales_items si ON si.order_id = o.id
                LEFT JOIN products pr ON pr.id = si.product_id
                GROUP BY o.id, o.platform, o.total_pln, o.total_eur, o.date
                ORDER BY o.date DESC, o.id DESC
            """).fetchall()
        
        result = []
        for r in rows:
            purchase_cost = r["purchase_cost"] if r["purchase_cost"] is not None else 0
            profit = r["total_pln"] - purchase_cost
            result.append((
                r["id"],
                r["platform"],
                round(r["total_pln"], 2),
                round(r["total_eur"], 2),
                round(purchase_cost, 2),
                round(profit, 2),
                r["date"],
                r["items"] or ""
            ))
        return result

    def get_detailed_sales(self, date_from, date_to):
        try:
            rows = self.conn.execute("""
                SELECT
                    o.id as order_id,
                    o.platform,
                    o.date,
                    pr.sku,
                    pr.title,
                    si.qty,
                    ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as item_revenue_pln,
                    ROUND(o.total_eur / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as item_revenue_eur,
                    ROUND(o.purchase_cost / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as item_cost,
                    o.total_pln as order_total_pln,
                    o.total_eur as order_total_eur,
                    o.purchase_cost as order_total_cost
                FROM sales_orders o
                JOIN sales_items si ON si.order_id = o.id
                JOIN products pr ON pr.id = si.product_id
                WHERE o.date BETWEEN ? AND ?
                ORDER BY o.date, o.id
            """, (date_from, date_to)).fetchall()
            
            result = []
            for r in rows:
                item_profit = r['item_revenue_pln'] - r['item_cost']
                result.append({
                    'order_id': r['order_id'],
                    'platform': r['platform'],
                    'date': r['date'],
                    'sku': r['sku'],
                    'title': r['title'],
                    'qty': r['qty'],
                    'item_revenue_pln': r['item_revenue_pln'],
                    'item_revenue_eur': r['item_revenue_eur'],
                    'item_cost': r['item_cost'],
                    'item_profit': item_profit,
                    'order_total_pln': r['order_total_pln'],
                    'order_total_eur': r['order_total_eur'],
                    'order_total_cost': r['order_total_cost']
                })
            return result
        except Exception as e:
            print(f"Błąd w get_detailed_sales: {e}")
            return []

    def delete_sale(self, oid):
        try:
            c = self.conn.cursor()
            
            # Sprawdź czy dla tej sprzedaży istnieje faktura
            invoice = c.execute("SELECT invoice_number FROM invoices WHERE sale_order_id=?", (oid,)).fetchone()
            
            try:
                purchase_items = c.execute("""
                    SELECT psh.purchase_item_id, psh.product_id
                    FROM purchase_stock_history psh
                    WHERE psh.sale_order_id = ?
                """, (oid,)).fetchall()
                
                for item in purchase_items:
                    c.execute("""
                        UPDATE purchase_items 
                        SET available_qty = available_qty + 1
                        WHERE id = ?
                    """, (item["purchase_item_id"],))
                
                c.execute("""
                    UPDATE purchase_stock_history 
                    SET sale_order_id = NULL 
                    WHERE sale_order_id = ?
                """, (oid,))
            except:
                pass
            
            rows = c.execute("""
                SELECT product_id, qty FROM sales_items WHERE order_id=?
            """, (oid,)).fetchall()

            for r in rows:
                self.update_stock(r["product_id"], r["qty"])

            c.execute("DELETE FROM sales_items WHERE order_id=?", (oid,))
            c.execute("DELETE FROM sales_orders WHERE id=?", (oid,))
            
            # Nie usuwamy faktury z bazy, tylko odłączamy od sprzedaży
            if invoice:
                c.execute("UPDATE invoices SET sale_order_id = NULL WHERE sale_order_id = ?", (oid,))
            
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Błąd w delete_sale: {e}")
            return False

    # ---------- INVOICES ----------
    def add_invoice(self, invoice_number, sale_order_id, file_path, 
                    customer_name="", customer_address="", issue_date=None, total_amount=0):
        if issue_date is None:
            issue_date = datetime.now().strftime("%Y-%m-%d")
        
        try:
            self.conn.execute("""
                INSERT INTO invoices (invoice_number, sale_order_id, file_path, 
                                     customer_name, customer_address, issue_date, total_amount)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (invoice_number, sale_order_id, file_path, customer_name, 
                  customer_address, issue_date, total_amount))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            # Jeśli faktura już istnieje, zaktualizuj
            self.conn.execute("""
                UPDATE invoices 
                SET sale_order_id = ?, file_path = ?, customer_name = ?, 
                    customer_address = ?, issue_date = ?, total_amount = ?
                WHERE invoice_number = ?
            """, (sale_order_id, file_path, customer_name, customer_address, 
                  issue_date, total_amount, invoice_number))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Błąd dodawania faktury: {e}")
            return False

    def list_invoices(self, date_from=None, date_to=None):
        try:
            query = """
                SELECT 
                    i.id,
                    i.invoice_number,
                    i.sale_order_id,
                    i.file_path,
                    i.customer_name,
                    i.customer_address,
                    i.issue_date,
                    i.total_amount,
                    i.created_at,
                    s.platform,
                    s.date as sale_date
                FROM invoices i
                LEFT JOIN sales_orders s ON s.id = i.sale_order_id
            """
            
            params = []
            if date_from and date_to:
                query += " WHERE i.issue_date BETWEEN ? AND ?"
                params = [date_from, date_to]
            
            query += " ORDER BY i.issue_date DESC, i.created_at DESC"
            
            rows = self.conn.execute(query, params).fetchall()
            
            result = []
            for r in rows:
                result.append({
                    'id': r['id'],
                    'invoice_number': r['invoice_number'],
                    'sale_order_id': r['sale_order_id'],
                    'file_path': r['file_path'],
                    'customer_name': r['customer_name'],
                    'customer_address': r['customer_address'],
                    'issue_date': r['issue_date'],
                    'total_amount': r['total_amount'],
                    'created_at': r['created_at'],
                    'platform': r['platform'],
                    'sale_date': r['sale_date']
                })
            return result
        except Exception as e:
            print(f"Błąd w list_invoices: {e}")
            return []

    def delete_invoice(self, invoice_id):
        try:
            c = self.conn.cursor()
            
            # Pobierz ścieżkę do pliku przed usunięciem
            invoice = c.execute("SELECT file_path FROM invoices WHERE id=?", (invoice_id,)).fetchone()
            
            if invoice and invoice['file_path'] and os.path.exists(invoice['file_path']):
                try:
                    os.remove(invoice['file_path'])
                except:
                    pass  # Nie przerywaj jeśli nie można usunąć pliku
            
            c.execute("DELETE FROM invoices WHERE id=?", (invoice_id,))
            self.conn.commit()
            return True
        except Exception as e:
            print(f"Błąd usuwania faktury: {e}")
            return False

    def get_invoice_by_number(self, invoice_number):
        return self.conn.execute(
            "SELECT * FROM invoices WHERE invoice_number=?",
            (invoice_number,)
        ).fetchone()

    # ---------- SPRZEDAŻ NARASTAJĄCA ----------
    def get_simple_sales_register_with_cumulative(self, date_from, date_to, personal_data):
        try:
            rows = self.conn.execute("""
                SELECT
                    o.id as order_id,
                    o.date as data_sprzedazy,
                    o.platform as platforma,
                    pr.sku as kod_produktu,
                    pr.title as nazwa_produktu,
                    si.qty as ilosc,
                    ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as cena_sprzedazy,
                    ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as wartosc_sprzedazy_pln,
                    ROUND(o.purchase_cost / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) as koszt_zakupu,
                    (ROUND(o.total_pln / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2) - 
                     ROUND(o.purchase_cost / (
                        SELECT SUM(qty) 
                        FROM sales_items 
                        WHERE order_id = o.id
                    ) * si.qty, 2)) as zysk_brutto
                FROM sales_orders o
                JOIN sales_items si ON si.order_id = o.id
                JOIN products pr ON pr.id = si.product_id
                WHERE o.date BETWEEN ? AND ?
                ORDER BY o.date ASC, o.id ASC, pr.sku
            """, (date_from, date_to)).fetchall()
            
            result = {
                "sprzedawca": {
                    "imie_nazwisko": personal_data.get("name", ""),
                    "adres": personal_data.get("address", ""),
                    "kod_pocztowy": personal_data.get("postal_code", ""),
                    "miejscowosc": personal_data.get("city", ""),
                    "pesel": personal_data.get("pesel", ""),
                    "nip": personal_data.get("nip", ""),
                    "regon": personal_data.get("regon", "")
                },
                "okres": f"{date_from} - {date_to}",
                "transakcje": [],
                "podsumowanie_miesieczne_narastajaco": [],
                "podsumowanie_roczne_narastajaco": []
            }
            
            cumulative_revenue = 0
            cumulative_cost = 0
            cumulative_profit = 0
            
            monthly_summary = {}
            unique_order_ids = set()
            
            for r in rows:
                transaction = {
                    'order_id': r['order_id'],
                    'data_sprzedazy': r['data_sprzedazy'],
                    'platforma': r['platforma'],
                    'kod_produktu': r['kod_produktu'],
                    'nazwa_produktu': r['nazwa_produktu'],
                    'ilosc': r['ilosc'],
                    'cena_sprzedazy': r['cena_sprzedazy'],
                    'wartosc_sprzedazy_pln': r['wartosc_sprzedazy_pln'],
                    'koszt_zakupu': r['koszt_zakupu'],
                    'zysk_brutto': r['zysk_brutto'],
                    'sprzedaz_narastajaca_pln': 0,
                    'sprzedaz_narastajaca_koszt': 0,
                    'sprzedaz_narastajaca_zysk': 0
                }
                
                cumulative_revenue += r['wartosc_sprzedazy_pln']
                cumulative_cost += r['koszt_zakupu']
                cumulative_profit += r['zysk_brutto']
                
                transaction['sprzedaz_narastajaca_pln'] = round(cumulative_revenue, 2)
                transaction['sprzedaz_narastajaca_koszt'] = round(cumulative_cost, 2)
                transaction['sprzedaz_narastajaca_zysk'] = round(cumulative_profit, 2)
                
                result["transakcje"].append(transaction)
                
                year_month = r['data_sprzedazy'][:7]
                if year_month not in monthly_summary:
                    monthly_summary[year_month] = {
                        'przychod': 0,
                        'koszt': 0,
                        'zysk': 0,
                        'unikalne_zamowienia': set(),
                        'liczba_pozycji': 0
                    }
                
                monthly_summary[year_month]['przychod'] += r['wartosc_sprzedazy_pln']
                monthly_summary[year_month]['koszt'] += r['koszt_zakupu']
                monthly_summary[year_month]['zysk'] += r['zysk_brutto']
                monthly_summary[year_month]['unikalne_zamowienia'].add(r['order_id'])
                monthly_summary[year_month]['liczba_pozycji'] += 1
                
                unique_order_ids.add(r['order_id'])
            
            # Przelicz liczbę transakcji na podstawie unikalnych zamówień
            for month_data in monthly_summary.values():
                month_data['liczba_transakcji'] = len(month_data['unikalne_zamowienia'])
                del month_data['unikalne_zamowienia']
            
            cumulative_monthly = 0
            for year_month in sorted(monthly_summary.keys()):
                month_data = monthly_summary[year_month]
                cumulative_monthly += month_data['przychod']
                
                result["podsumowanie_miesieczne_narastajaco"].append({
                    'miesiac': year_month,
                    'przychod_miesiac': round(month_data['przychod'], 2),
                    'koszt_miesiac': round(month_data['koszt'], 2),
                    'zysk_miesiac': round(month_data['zysk'], 2),
                    'liczba_transakcji': month_data['liczba_transakcji'],
                    'liczba_pozycji': month_data['liczba_pozycji'],
                    'przychod_narastajaco': round(cumulative_monthly, 2),
                    'zysk_narastajaco': round(sum(m['zysk'] for k, m in monthly_summary.items() 
                                                 if k <= year_month), 2)
                })
            
            result["podsumowanie_roczne_narastajaco"] = []
            cumulative_yearly = 0
            
            yearly_summary = {}
            for year_month, data in monthly_summary.items():
                year = year_month[:4]
                if year not in yearly_summary:
                    yearly_summary[year] = {'przychod': 0, 'zysk': 0, 'liczba_transakcji': 0}
                yearly_summary[year]['przychod'] += data['przychod']
                yearly_summary[year]['zysk'] += data['zysk']
                yearly_summary[year]['liczba_transakcji'] += data['liczba_transakcji']
            
            for year in sorted(yearly_summary.keys()):
                cumulative_yearly += yearly_summary[year]['przychod']
                result["podsumowanie_roczne_narastajaco"].append({
                    'rok': year,
                    'przychod_rok': round(yearly_summary[year]['przychod'], 2),
                    'zysk_rok': round(yearly_summary[year]['zysk'], 2),
                    'liczba_transakcji': yearly_summary[year]['liczba_transakcji'],
                    'przychod_narastajaco': round(cumulative_yearly, 2)
                })
            
            result["podsumowanie_ogolne"] = {
                'przychod_calkowity': round(cumulative_revenue, 2),
                'koszt_calkowity': round(cumulative_cost, 2),
                'zysk_calkowity': round(cumulative_profit, 2),
                'liczba_transakcji': len(unique_order_ids),
                'liczba_pozycji': len(rows)
            }
            
            return result
        except Exception as e:
            print(f"Błąd w get_simple_sales_register_with_cumulative: {e}")
            return None

    # ---------- REPORT ----------
    def report(self, d1, d2):
        try:
            return self.conn.execute("""
                SELECT
                    SUM(o.total_pln) AS sales_pln,
                    SUM(o.total_eur) AS sales_eur,
                    SUM(o.purchase_cost) AS purchase_costs,
                    (
                        SELECT SUM(po.total_pln)
                        FROM purchase_orders po
                        WHERE po.date BETWEEN ? AND ?
                    ) AS all_costs
                FROM sales_orders o
                WHERE o.date BETWEEN ? AND ?
            """, (d1, d2, d1, d2)).fetchone()
        except sqlite3.OperationalError:
            return self.conn.execute("""
                SELECT
                    SUM(o.total_pln) AS sales_pln,
                    SUM(o.total_eur) AS sales_eur,
                    0 AS purchase_costs,
                    (
                        SELECT SUM(po.total_pln)
                        FROM purchase_orders po
                        WHERE po.date BETWEEN ? AND ?
                    ) AS all_costs
                FROM sales_orders o
                WHERE o.date BETWEEN ? AND ?
            """, (d1, d2, d1, d2)).fetchone()

    # ---------- CSV ----------
    def export_sales_csv(self, path):
        rows = self.list_sales()
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["ID", "Platforma", "PLN", "EUR", "Koszt zakupu", "Zysk", "Data", "Pozycje"])
            for r in rows:
                w.writerow(r)

    # ---------- SZCZEGÓŁOWY RAPORT CSV ----------
    def export_detailed_report_csv(self, path, date_from, date_to, 
                                 include_purchases=True, include_sales=True,
                                 include_summary=True, include_products=False,
                                 personal_data=None, report_type=None, config=None):
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                
                # Użyj report_type w tytule jeśli dostępny
                title = "EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
                if report_type == "quarterly":
                    title = "KWARTALNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
                elif report_type == "monthly":
                    title = "MIESIĘCZNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
                elif report_type == "yearly":
                    title = "ROCZNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
                
                w.writerow([title])
                w.writerow([f"Okres: {date_from} - {date_to}"])
                w.writerow([f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                w.writerow([])
                
                if personal_data:
                    w.writerow(["DANE SPRZEDAWCY"])
                    w.writerow([f"Imię i nazwisko: {personal_data.get('name', '')}"])
                    w.writerow([f"Adres: {personal_data.get('address', '')}"])
                    w.writerow([f"Kod pocztowy: {personal_data.get('postal_code', '')}"])
                    w.writerow([f"Miejscowość: {personal_data.get('city', '')}"])
                    w.writerow([f"PESEL: {personal_data.get('pesel', '')}"])
                    if personal_data.get('nip'):
                        w.writerow([f"NIP: {personal_data.get('nip', '')}"])
                    if personal_data.get('regon'):
                        w.writerow([f"REGON: {personal_data.get('regon', '')}"])
                    w.writerow([])
                
                register_data = self.get_simple_sales_register_with_cumulative(date_from, date_to, personal_data or {})
                
                if register_data and register_data["transakcje"]:
                    w.writerow(["PODSUMOWANIE OGÓLNE OKRESU"])
                    w.writerow(["Przychód całkowity:", f"{register_data['podsumowanie_ogolne']['przychod_calkowity']:.2f} PLN"])
                    w.writerow(["Koszt całkowity:", f"{register_data['podsumowanie_ogolne']['koszt_calkowity']:.2f} PLN"])
                    w.writerow(["Zysk całkowity:", f"{register_data['podsumowanie_ogolne']['zysk_calkowity']:.2f} PLN"])
                    w.writerow(["Liczba transakcji (zamówień):", register_data['podsumowanie_ogolne']['liczba_transakcji']])
                    w.writerow(["Liczba pozycji sprzedaży:", register_data['podsumowanie_ogolne']['liczba_pozycji']])
                    w.writerow([])
                    
                    w.writerow(["PODSUMOWANIE MIESIĘCZNE (NARASTAJĄCO)"])
                    w.writerow(["Miesiąc", "Przychód miesiąc", "Zysk miesiąc", "Liczba transakcji", "Liczba pozycji",
                               "PRZYCHÓD NARASTAJĄCO", "ZYSK NARASTAJĄCO"])
                    
                    for month_data in register_data.get("podsumowanie_miesieczne_narastajaco", []):
                        w.writerow([
                            month_data['miesiac'],
                            f"{month_data['przychod_miesiac']:.2f}",
                            f"{month_data['zysk_miesiac']:.2f}",
                            month_data['liczba_transakcji'],
                            month_data.get('liczba_pozycji', 0),
                            f"{month_data['przychod_narastajaco']:.2f}",
                            f"{month_data['zysk_narastajaco']:.2f}"
                        ])
                    w.writerow([])
                    
                    w.writerow(["PODSUMOWANIE ROCZNE (NARASTAJĄCO)"])
                    w.writerow(["Rok", "Przychód roczny", "Zysk roczny", "Liczba transakcji", "PRZYCHÓD NARASTAJĄCO"])
                    
                    for year_data in register_data.get("podsumowanie_roczne_narastajaco", []):
                        w.writerow([
                            year_data['rok'],
                            f"{year_data['przychod_rok']:.2f}",
                            f"{year_data['zysk_rok']:.2f}",
                            year_data.get('liczba_transakcji', 0),
                            f"{year_data['przychod_narastajaco']:.2f}"
                        ])
                    w.writerow([])
                    
                    w.writerow(["SZCZEGÓŁOWA EWIDENCJA TRANSAKCJI"])
                    w.writerow(["Data", "Platforma", "Produkt", "Ilość", "Cena sprzedaży", 
                               "Wartość sprzedaży", "Koszt zakupu", "Zysk",
                               "PRZYCHÓD NARASTAJĄCO", "KOSZT NARASTAJĄCO", "ZYSK NARASTAJĄCO"])
                    
                    for transaction in register_data["transakcje"]:
                        product_display = f"{transaction['nazwa_produktu']} ({transaction['kod_produktu']})"
                        w.writerow([
                            transaction['data_sprzedazy'],
                            transaction['platforma'],
                            product_display,
                            transaction['ilosc'],
                            f"{transaction['cena_sprzedazy']:.2f}",
                            f"{transaction['wartosc_sprzedazy_pln']:.2f}",
                            f"{transaction['koszt_zakupu']:.2f}",
                            f"{transaction['zysk_brutto']:.2f}",
                            f"{transaction['sprzedaz_narastajaca_pln']:.2f}",
                            f"{transaction['sprzedaz_narastajaca_koszt']:.2f}",
                            f"{transaction['sprzedaz_narastajaca_zysk']:.2f}"
                        ])
                    w.writerow([])
                    
                    # Użyj config jeśli dostępny do pobrania limitów
                    minimal_wage = 4242
                    limit_multiplier = 0.75  # domyślnie miesięczny
                    
                    if config:
                        try:
                            year = int(date_from[:4])
                            minimal_wage = config.get_minimal_wage(year)
                            
                            # Sprawdź czy to raport kwartalny i czy używamy limitów kwartalnych
                            if report_type == "quarterly" and config.use_quarterly_limits():
                                limit_multiplier = config.get_limits_config().get("quarterly_limit_multiplier", 2.25)
                                limit = minimal_wage * limit_multiplier
                                limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia (limit kwartalny)"
                            else:
                                limit = minimal_wage * 0.75
                                limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                        except:
                            limit = minimal_wage * 0.75
                            limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                    else:
                        limit = minimal_wage * 0.75
                        limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                    
                    total_revenue = register_data['podsumowanie_ogolne']['przychod_calkowity']
                    year = date_from[:4]
                    
                    w.writerow([f"ANALIZA PROGU LIMITU ({year} r.)"])
                    w.writerow([f"Minimalne wynagrodzenie: {minimal_wage} PLN"])
                    w.writerow([f"{limit_text}: {limit:.2f} PLN"])
                    w.writerow([f"Przychód narastająco w roku {year}: {total_revenue:.2f} PLN"])
                    
                    if total_revenue > limit:
                        w.writerow(["UWAGA: Przekroczono limit działalności nierejestrowanej!"])
                        w.writerow(["Konieczna rejestracja działalności gospodarczej"])
                    else:
                        w.writerow(["OK: Przychód mieści się w limicie działalności nierejestrowanej"])
                    
                    w.writerow([])
                else:
                    w.writerow(["Brak danych sprzedaży w wybranym okresie"])
                    w.writerow([])
                    
            return True
        except Exception as e:
            print(f"Błąd w export_detailed_report_csv: {e}")
            return False

    # ---------- SZCZEGÓŁOWY RAPORT EXCEL ----------
    def export_detailed_report_excel(self, path, date_from, date_to,
                                   include_purchases=True, include_sales=True,
                                   include_summary=True, include_products=False,
                                   personal_data=None, report_type=None, config=None):
        if not HAS_EXCEL:
            raise ImportError("Biblioteka openpyxl nie jest zainstalowana. Użyj: pip install openpyxl")
        
        try:
            wb = openpyxl.Workbook()
            
            header_font = Font(bold=True, color="FFFFFF")
            warning_font = Font(bold=True, color="FF0000")
            cumulative_font = Font(bold=True, color="2E7D32")
            
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cumulative_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            warning_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
            
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            money_format = '#,##0.00'
            
            register_data = self.get_simple_sales_register_with_cumulative(date_from, date_to, personal_data or {})
            
            ws_summary = wb.active
            ws_summary.title = "Podsumowanie"
            
            row = 1
            
            # Tytuł zależny od typu raportu
            title = "EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
            if report_type == "quarterly":
                title = "KWARTALNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
            elif report_type == "monthly":
                title = "MIESIĘCZNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
            elif report_type == "yearly":
                title = "ROCZNA EWIDENCJA SPRZEDAŻY DLA DZIAŁALNOŚCI NIEREJESTROWANEJ"
            
            ws_summary.cell(row=row, column=1, value=title)
            ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)
            row += 2
            
            ws_summary.cell(row=row, column=1, value=f"Okres: {date_from} - {date_to}")
            ws_summary.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            ws_summary.cell(row=row, column=1, value=f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            row += 2
            
            if personal_data:
                ws_summary.cell(row=row, column=1, value="DANE SPRZEDAWCY").font = Font(bold=True)
                row += 1
                
                ws_summary.cell(row=row, column=1, value=f"Imię i nazwisko: {personal_data.get('name', '')}")
                row += 1
                ws_summary.cell(row=row, column=1, value=f"Adres: {personal_data.get('address', '')}")
                row += 1
                ws_summary.cell(row=row, column=1, value=f"Kod pocztowy: {personal_data.get('postal_code', '')}")
                row += 1
                ws_summary.cell(row=row, column=1, value=f"Miejscowość: {personal_data.get('city', '')}")
                row += 1
                ws_summary.cell(row=row, column=1, value=f"PESEL: {personal_data.get('pesel', '')}")
                row += 1
                
                if personal_data.get('nip'):
                    ws_summary.cell(row=row, column=1, value=f"NIP: {personal_data.get('nip', '')}")
                    row += 1
                if personal_data.get('regon'):
                    ws_summary.cell(row=row, column=1, value=f"REGON: {personal_data.get('regon', '')}")
                    row += 1
                row += 1
            
            if register_data and register_data.get("transakcje"):
                summary = register_data["podsumowanie_ogolne"]
                
                ws_summary.cell(row=row, column=1, value="PODSUMOWANIE OGÓLNE OKRESU").font = Font(bold=True)
                row += 1
                
                data = [
                    ["Przychód całkowity:", f"{summary['przychod_calkowity']:.2f}"],
                    ["Koszt całkowity:", f"{summary['koszt_calkowity']:.2f}"],
                    ["Zysk całkowity:", f"{summary['zysk_calkowity']:.2f}"],
                    ["Liczba transakcji (zamówień):", summary['liczba_transakcji']],
                    ["Liczba pozycji sprzedaży:", summary['liczba_pozycji']]
                ]
                
                for i, (label, value) in enumerate(data):
                    ws_summary.cell(row=row, column=1, value=label)
                    if i < 3:
                        ws_summary.cell(row=row, column=2, value=float(value.split()[0]))
                        ws_summary.cell(row=row, column=2).number_format = money_format
                        ws_summary.cell(row=row, column=2).font = Font(bold=True)
                    else:
                        ws_summary.cell(row=row, column=2, value=int(value))
                    row += 1
                
                row += 1
                
                ws_monthly = wb.create_sheet("Podsumowanie miesięczne")
                ws_monthly.cell(row=1, column=1, value="PODSUMOWANIE MIESIĘCZNE (NARASTAJĄCO)").font = Font(bold=True, size=12)
                
                headers = ["Miesiąc", "Przychód miesiąc", "Zysk miesiąc", "Liczba transakcji", "Liczba pozycji",
                          "PRZYCHÓD NARASTAJĄCO", "ZYSK NARASTAJĄCO"]
                
                for col, header in enumerate(headers, start=1):
                    cell = ws_monthly.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                row_monthly = 4
                for month_data in register_data.get("podsumowanie_miesieczne_narastajaco", []):
                    ws_monthly.cell(row=row_monthly, column=1, value=month_data['miesiac'])
                    ws_monthly.cell(row=row_monthly, column=2, value=month_data['przychod_miesiac']).number_format = money_format
                    ws_monthly.cell(row=row_monthly, column=3, value=month_data['zysk_miesiac']).number_format = money_format
                    ws_monthly.cell(row=row_monthly, column=4, value=month_data['liczba_transakcji'])
                    ws_monthly.cell(row=row_monthly, column=5, value=month_data.get('liczba_pozycji', 0))
                    ws_monthly.cell(row=row_monthly, column=6, value=month_data['przychod_narastajaco']).number_format = money_format
                    ws_monthly.cell(row=row_monthly, column=6).font = cumulative_font
                    ws_monthly.cell(row=row_monthly, column=6).fill = cumulative_fill
                    ws_monthly.cell(row=row_monthly, column=7, value=month_data['zysk_narastajaco']).number_format = money_format
                    ws_monthly.cell(row=row_monthly, column=7).font = cumulative_font
                    row_monthly += 1
                
                ws_details = wb.create_sheet("Transakcje")
                ws_details.cell(row=1, column=1, value="SZCZEGÓŁOWA EWIDENCJA TRANSAKCJI Z NARASTANIEM").font = Font(bold=True, size=12)
                
                detail_headers = ["Data", "Platforma", "Produkt", "Ilość", "Cena sprzedaży", 
                                "Wartość sprzedaży", "Koszt zakupu", "Zysk",
                                "PRZYCHÓD NARASTAJĄCO", "KOSZT NARASTAJĄCO", "ZYSK NARASTAJĄCO"]
                
                for col, header in enumerate(detail_headers, start=1):
                    cell = ws_details.cell(row=3, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                row_detail = 4
                for transaction in register_data["transakcje"]:
                    ws_details.cell(row=row_detail, column=1, value=transaction['data_sprzedazy'])
                    ws_details.cell(row=row_detail, column=2, value=transaction['platforma'])
                    product_display = f"{transaction['nazwa_produktu']} ({transaction['kod_produktu']})"
                    ws_details.cell(row=row_detail, column=3, value=product_display)
                    ws_details.cell(row=row_detail, column=4, value=transaction['ilosc'])
                    ws_details.cell(row=row_detail, column=5, value=transaction['cena_sprzedazy']).number_format = money_format
                    ws_details.cell(row=row_detail, column=6, value=transaction['wartosc_sprzedazy_pln']).number_format = money_format
                    ws_details.cell(row=row_detail, column=7, value=transaction['koszt_zakupu']).number_format = money_format
                    ws_details.cell(row=row_detail, column=8, value=transaction['zysk_brutto']).number_format = money_format
                    
                    for col_offset, key in enumerate(['sprzedaz_narastajaca_pln', 'sprzedaz_narastajaca_koszt', 'sprzedaz_narastajaca_zysk'], start=9):
                        ws_details.cell(row=row_detail, column=col_offset, value=transaction[key]).number_format = money_format
                        ws_details.cell(row=row_detail, column=col_offset).font = Font(bold=True)
                        ws_details.cell(row=row_detail, column=col_offset).fill = cumulative_fill
                    
                    row_detail += 1
                
                ws_analysis = wb.create_sheet("Analiza progu US")
                ws_analysis.cell(row=1, column=1, value="ANALIZA PROGU LIMITU DLA US").font = Font(bold=True, size=14, color="FF0000")
                
                # Użyj config jeśli dostępny do pobrania limitów
                minimal_wage = 4242
                limit_multiplier = 0.75  # domyślnie miesięczny
                limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                
                if config:
                    try:
                        year = int(date_from[:4])
                        minimal_wage = config.get_minimal_wage(year)
                        
                        # Sprawdź czy to raport kwartalny i czy używamy limitów kwartalnych
                        if report_type == "quarterly" and config.use_quarterly_limits():
                            limit_multiplier = config.get_limits_config().get("quarterly_limit_multiplier", 2.25)
                            limit = minimal_wage * limit_multiplier
                            limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia (limit kwartalny)"
                        else:
                            limit = minimal_wage * 0.75
                            limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                    except:
                        limit = minimal_wage * 0.75
                        limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                else:
                    limit = minimal_wage * 0.75
                    limit_text = "75% minimalnego wynagrodzenia (limit miesięczny)"
                
                total_revenue = register_data['podsumowanie_ogolne']['przychod_calkowity']
                year = date_from[:4]
                
                analysis_data = [
                    ["Parametr", "Wartość"],
                    ["Minimalne wynagrodzenie", f"{minimal_wage} PLN"],
                    [limit_text, f"{limit:.2f} PLN"],
                    ["Przychód narastająco", f"{total_revenue:.2f} PLN"],
                    ["", ""],
                    ["ANALIZA:", ""]
                ]
                
                row_analysis = 3
                for label, value in analysis_data:
                    ws_analysis.cell(row=row_analysis, column=1, value=label)
                    ws_analysis.cell(row=row_analysis, column=2, value=value)
                    row_analysis += 1
                
                if total_revenue > limit:
                    ws_analysis.cell(row=row_analysis, column=1, value="UWAGA: PRZEKROCZONO LIMIT!")
                    ws_analysis.cell(row=row_analysis, column=1).font = warning_font
                    ws_analysis.cell(row=row_analysis, column=1).fill = warning_fill
                    ws_analysis.cell(row=row_analysis, column=2, value="Konieczna rejestracja działalności gospodarczej")
                    ws_analysis.cell(row=row_analysis, column=2).font = warning_font
                else:
                    ws_analysis.cell(row=row_analysis, column=1, value="OK: W LIMICIE")
                    ws_analysis.cell(row=row_analysis, column=1).font = cumulative_font
                    ws_analysis.cell(row=row_analysis, column=2, value="Działalność nierejestrowana może być kontynuowana")
                
                # Dodatkowe informacje dla raportów kwartalnych
                if report_type == "quarterly" and config and config.use_quarterly_limits():
                    row_analysis += 2
                    ws_analysis.cell(row=row_analysis, column=1, value="UWAGA OD 2026 ROKU:")
                    ws_analysis.cell(row=row_analysis, column=1).font = Font(bold=True, color="2E7D32")
                    row_analysis += 1
                    ws_analysis.cell(row=row_analysis, column=1, value="Od 2026 roku obowiązują limity kwartalne dla działalności nierejestrowanej.")
                    row_analysis += 1
                    ws_analysis.cell(row=row_analysis, column=1, value=f"Limit kwartalny wynosi {limit_multiplier*100:.0f}% minimalnego wynagrodzenia.")
                    row_analysis += 1
                    ws_analysis.cell(row=row_analysis, column=1, value=f"Przychód w tym kwartale: {total_revenue:.2f} PLN")
                    ws_analysis.cell(row=row_analysis, column=1).font = Font(bold=True)
                
                for ws in [ws_summary, ws_monthly, ws_details, ws_analysis]:
                    for col in range(1, 20):
                        ws.column_dimensions[get_column_letter(col)].width = 15
            
            wb.save(path)
            return True
            
        except Exception as e:
            print(f"Błąd w export_detailed_report_excel: {e}")
            return False

    # ---------- DEBUG ----------
    def debug_stock_changes(self, pid: int):
        current = self.conn.execute("SELECT stock FROM products WHERE id=?", (pid,)).fetchone()
        
        purchases = self.conn.execute("""
            SELECT pi.qty, pi.unit_cost, po.date, pi.available_qty
            FROM purchase_items pi
            JOIN purchase_orders po ON po.id = pi.order_id
            WHERE pi.product_id=?
            ORDER BY po.date
        """, (pid,)).fetchall()
        
        return {
            "current_stock": current["stock"] if current else 0,
            "purchases": purchases,
        }
