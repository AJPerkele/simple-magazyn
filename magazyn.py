"""
╔══════════════════════════════════════════════════════════╗
║    SYSTEM MAGAZYNOWO-SPRZEDAŻOWY  v3.1                  ║
║    Działalność nierejestrowana | PySide6 | SQLite        ║
╚══════════════════════════════════════════════════════════╝
Autor: @AJPerkele  |  Licencja: GNU GPL v3.0
"""

import sys, os, csv, json, shutil, sqlite3, requests
from datetime import datetime, timedelta

from PySide6.QtWidgets import *
from PySide6.QtCore import QDate, Qt, QTimer, QSize
from PySide6.QtGui import QFont, QAction, QColor, QPainter, QPen, QBrush, QLinearGradient

APP_VERSION = "3.1.0"
APP_NAME    = "System Magazynowo-Sprzedażowy"
APP_AUTHOR  = "@AJPerkele"
APP_LICENSE = "GNU GPL v3.0"
BUILD_DATE  = datetime.now().strftime("%Y-%m-%d")
PLATFORMS   = ["Vinted", "OLX", "Allegro Lokalnie", "FB Marketplace", "Inne"]
PLATFORM_LIMIT = 29  # limit sprzedaży na platformę (działalność nierejestrowana)

# ─────────────────────────────────────────────────────────
#  MOTYWY – DZIENNY I NOCNY
# ─────────────────────────────────────────────────────────

THEME_DAY = {
    "name": "Dzienny",
    "bg":           "#FFFFFF",
    "bg2":          "#F5F5F5",
    "bg3":          "#FAFAFA",
    "border":       "#E0E0E0",
    "text":         "#1A1A1A",
    "text2":        "#555555",
    "text3":        "#888888",
    "accent":       "#C62828",
    "accent_hover": "#B71C1C",
    "accent_press": "#8B0000",
    "success":      "#2E7D32",
    "warning":      "#E65100",
    "danger":       "#C62828",
    "menubar_bg":   "#C62828",
    "menubar_text": "#FFFFFF",
    "header_bg":    "#F0F0F0",
    "header_text":  "#333333",
    "sel_bg":       "#C62828",
    "sel_text":     "#FFFFFF",
    "toolbar_bg":   "#F5F5F5",
    "card_bg":      "#FFFFFF",
    "chart_bar":    "#C62828",
    "chart_bar2":   "#EF9A9A",
    "chart_label":  "#555555",
    "chart_grid":   "#E0E0E0",
    "progress_bg":  "#E0E0E0",
    "input_bg":     "#FFFFFF",
    "status_bg":    "#EEEEEE",
    "kpi_rev":      "#C62828",
    "kpi_profit":   "#2E7D32",
    "kpi_sales":    "#1565C0",
    "kpi_prod":     "#6A1B9A",
    "kpi_stock":    "#E65100",
}

THEME_NIGHT = {
    "name": "Nocny",
    "bg":           "#1A1D23",
    "bg2":          "#1E2128",
    "bg3":          "#22262E",
    "border":       "#2C2F38",
    "text":         "#E8EAF0",
    "text2":        "#C0C4CE",
    "text3":        "#8B919E",
    "accent":       "#E36A35",
    "accent_hover": "#F07840",
    "accent_press": "#C85E2A",
    "success":      "#43A047",
    "warning":      "#FB8C00",
    "danger":       "#E53935",
    "menubar_bg":   "#13151A",
    "menubar_text": "#C9CDD4",
    "header_bg":    "#13151A",
    "header_text":  "#8B919E",
    "sel_bg":       "#E36A35",
    "sel_text":     "#FFFFFF",
    "toolbar_bg":   "#1A1D23",
    "card_bg":      "#1E2128",
    "chart_bar":    "#E36A35",
    "chart_bar2":   "#7B3F1A",
    "chart_label":  "#8B919E",
    "chart_grid":   "#2C2F38",
    "progress_bg":  "#22262E",
    "input_bg":     "#22262E",
    "status_bg":    "#13151A",
    "kpi_rev":      "#E36A35",
    "kpi_profit":   "#43A047",
    "kpi_sales":    "#42A5F5",
    "kpi_prod":     "#AB47BC",
    "kpi_stock":    "#FFA726",
}

CURRENT_THEME = THEME_DAY  # domyślnie dzienny


def build_qss(t):
    return f"""
QWidget {{
    background-color: {t['bg']};
    color: {t['text']};
    font-family: 'Segoe UI', 'Helvetica Neue', Arial, sans-serif;
    font-size: 13px;
}}
QMainWindow {{ background-color: {t['bg']}; }}
QDialog {{ background-color: {t['bg']}; }}

/* ── MENUBAR ── */
QMenuBar {{
    background-color: {t['menubar_bg']};
    color: {t['menubar_text']};
    font-size: 13px;
    font-weight: 600;
    padding: 2px 0;
    border-bottom: 2px solid {t['accent']};
}}
QMenuBar::item {{
    padding: 7px 16px;
    border-radius: 3px;
    background: transparent;
    color: {t['menubar_text']};
}}
QMenuBar::item:selected, QMenuBar::item:pressed {{
    background-color: {t['accent']};
    color: #FFFFFF;
}}
QMenu {{
    background-color: {t['bg2']};
    border: 1px solid {t['border']};
    border-radius: 6px;
    padding: 4px;
}}
QMenu::item {{
    padding: 8px 28px 8px 14px;
    border-radius: 4px;
    color: {t['text']};
}}
QMenu::item:selected {{ background-color: {t['accent']}; color: #FFFFFF; }}
QMenu::item:disabled {{ color: {t['text3']}; }}
QMenu::separator {{ height: 1px; background: {t['border']}; margin: 4px 8px; }}

/* ── TOOLBAR ── */
QToolBar {{
    background-color: {t['toolbar_bg']};
    border-bottom: 1px solid {t['border']};
    spacing: 3px;
    padding: 5px 10px;
}}
QToolButton {{
    padding: 6px 12px;
    border-radius: 5px;
    color: {t['text']};
    background: transparent;
    font-size: 12px;
    font-weight: 500;
}}
QToolButton:hover {{ background-color: {t['bg2']}; color: {t['accent']}; }}
QToolButton:pressed {{ background-color: {t['bg3']}; }}

/* ── TABELE ── */
QTableWidget {{
    background-color: {t['bg2']};
    alternate-background-color: {t['bg3']};
    border: 1px solid {t['border']};
    border-radius: 6px;
    gridline-color: {t['border']};
    color: {t['text']};
    selection-background-color: {t['sel_bg']};
    selection-color: {t['sel_text']};
}}
QTableWidget::item {{ padding: 5px 8px; color: {t['text']}; }}
QTableWidget::item:selected {{ background-color: {t['sel_bg']}; color: {t['sel_text']}; }}
QHeaderView::section {{
    background-color: {t['header_bg']};
    color: {t['header_text']};
    padding: 7px 8px;
    border: none;
    border-right: 1px solid {t['border']};
    border-bottom: 2px solid {t['accent']};
    font-weight: 700;
    font-size: 11px;
}}

/* ── PRZYCISKI ── */
QPushButton {{
    background-color: {t['accent']};
    color: #FFFFFF;
    padding: 7px 16px;
    border-radius: 5px;
    font-weight: 700;
    border: none;
    font-size: 12px;
}}
QPushButton:hover {{ background-color: {t['accent_hover']}; }}
QPushButton:pressed {{ background-color: {t['accent_press']}; }}
QPushButton:disabled {{ background-color: {t['border']}; color: {t['text3']}; }}
QPushButton[secondary="true"] {{
    background-color: {t['bg2']};
    color: {t['text']};
    border: 1px solid {t['border']};
}}
QPushButton[secondary="true"]:hover {{ background-color: {t['bg3']}; color: {t['accent']}; }}
QPushButton[danger="true"] {{ background-color: {t['danger']}; }}
QPushButton[danger="true"]:hover {{ background-color: #FF5252; }}
QPushButton[success="true"] {{ background-color: {t['success']}; }}
QPushButton[success="true"]:hover {{ background-color: #66BB6A; }}

/* ── INPUTY ── */
QLineEdit, QTextEdit, QPlainTextEdit {{
    background-color: {t['input_bg']};
    border: 1.5px solid {t['border']};
    border-radius: 5px;
    padding: 6px 10px;
    color: {t['text']};
    selection-background-color: {t['accent']};
}}
QLineEdit:focus, QTextEdit:focus {{ border-color: {t['accent']}; }}
QLineEdit:disabled {{ color: {t['text3']}; }}
QSpinBox, QDoubleSpinBox {{
    background-color: {t['input_bg']};
    border: 1.5px solid {t['border']};
    border-radius: 5px;
    padding: 5px 8px;
    color: {t['text']};
}}
QSpinBox:focus, QDoubleSpinBox:focus {{ border-color: {t['accent']}; }}
QSpinBox::up-button, QDoubleSpinBox::up-button,
QSpinBox::down-button, QDoubleSpinBox::down-button {{
    background-color: {t['bg2']}; border: none; width: 18px;
}}
QComboBox {{
    background-color: {t['input_bg']};
    border: 1.5px solid {t['border']};
    border-radius: 5px;
    padding: 5px 10px;
    color: {t['text']};
    min-width: 120px;
}}
QComboBox:focus {{ border-color: {t['accent']}; }}
QComboBox::drop-down {{ border: none; background: transparent; width: 24px; }}
QComboBox QAbstractItemView {{
    background-color: {t['bg2']};
    border: 1px solid {t['border']};
    selection-background-color: {t['accent']};
    color: {t['text']};
}}
QDateEdit {{
    background-color: {t['input_bg']};
    border: 1.5px solid {t['border']};
    border-radius: 5px;
    padding: 5px 10px;
    color: {t['text']};
}}
QDateEdit:focus {{ border-color: {t['accent']}; }}
QDateEdit::drop-down {{ border: none; width: 20px; }}

/* ── GROUP BOX ── */
QGroupBox {{
    border: 1px solid {t['border']};
    border-radius: 6px;
    margin-top: 14px;
    padding: 12px 8px 8px 8px;
    color: {t['text3']};
    font-weight: 700;
    font-size: 11px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 12px;
    top: -1px;
    padding: 0 6px;
    background-color: {t['bg']};
    color: {t['accent']};
    font-weight: 700;
}}

/* ── CHECKBOX / RADIO ── */
QCheckBox, QRadioButton {{ color: {t['text2']}; spacing: 7px; }}
QCheckBox::indicator, QRadioButton::indicator {{
    width: 16px; height: 16px;
    border: 2px solid {t['border']};
    border-radius: 3px;
    background: {t['input_bg']};
}}
QCheckBox::indicator:checked {{ background-color: {t['accent']}; border-color: {t['accent']}; }}
QRadioButton::indicator {{ border-radius: 8px; }}
QRadioButton::indicator:checked {{ background-color: {t['accent']}; border-color: {t['accent']}; }}

/* ── TABS ── */
QTabWidget::pane {{
    border: 1px solid {t['border']};
    border-radius: 0 6px 6px 6px;
    background: {t['bg']};
}}
QTabBar::tab {{
    background: {t['bg2']};
    border: 1px solid {t['border']};
    border-bottom: none;
    padding: 8px 20px;
    color: {t['text3']};
    margin-right: 2px;
    border-radius: 6px 6px 0 0;
    font-weight: 600;
}}
QTabBar::tab:selected {{ background: {t['bg']}; color: {t['accent']}; border-bottom: 2px solid {t['accent']}; }}
QTabBar::tab:hover {{ color: {t['text']}; }}

/* ── STATUSBAR ── */
QStatusBar {{
    background-color: {t['status_bg']};
    color: {t['text3']};
    border-top: 1px solid {t['border']};
    font-size: 11px;
    padding: 2px 8px;
}}

/* ── SCROLLBAR ── */
QScrollBar:vertical {{
    background: {t['bg']}; width: 8px; border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {t['border']}; border-radius: 4px; min-height: 20px;
}}
QScrollBar::handle:vertical:hover {{ background: {t['accent']}; }}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
QScrollBar:horizontal {{
    background: {t['bg']}; height: 8px; border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {t['border']}; border-radius: 4px;
}}
QScrollBar::handle:horizontal:hover {{ background: {t['accent']}; }}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ width: 0; }}

/* ── PROGRESS ── */
QProgressBar {{
    background-color: {t['progress_bg']};
    border: 1px solid {t['border']};
    border-radius: 5px;
    height: 10px;
    text-align: center;
    color: transparent;
}}
QProgressBar::chunk {{ background-color: {t['accent']}; border-radius: 5px; }}

/* ── LIST WIDGET ── */
QListWidget {{
    background: {t['bg2']};
    border: 1px solid {t['border']};
    border-radius: 5px;
    color: {t['text']};
}}
QListWidget::item:selected {{ background-color: {t['accent']}; color: #FFFFFF; }}
QListWidget::item:hover {{ background-color: {t['bg3']}; }}

/* ── LABEL ── */
QLabel {{ color: {t['text']}; background: transparent; }}

/* ── MESSAGE BOX ── */
QMessageBox {{ background-color: {t['bg']}; }}
QMessageBox QLabel {{ color: {t['text']}; font-size: 13px; }}

/* ── FRAME ── */
QFrame {{ color: {t['text']}; }}
"""


def apply_theme(app, theme):
    global CURRENT_THEME
    CURRENT_THEME = theme
    app.setStyleSheet(build_qss(theme))


# ─────────────────────────────────────────────────────────
#  KURS EUR
# ─────────────────────────────────────────────────────────
def get_eur_rate(date_str=None):
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    try:
        url = f"http://api.nbp.pl/api/exchangerates/rates/a/eur/{date_str}/"
        r = requests.get(url, headers={"Accept": "application/json"}, timeout=4)
        if r.status_code == 200:
            return r.json()["rates"][0]["mid"]
    except Exception:
        pass
    return 4.25


# ─────────────────────────────────────────────────────────
#  KONFIGURACJA
# ─────────────────────────────────────────────────────────
class Config:
    DEFAULTS = {
        "database_path":  "data.db",
        "last_opened":    None,
        "invoice_counter": 1,
        "invoice_prefix": "R",
        "invoice_year":   None,
        "save_pdf":       True,
        "theme":          "day",
        "business_info":  {},
        "invoice_config": {"seller_info": "", "footer_text": "Dziękuję za zakup!"},
        "limits": {
            "minimal_wage": 4666.0,
            "quarterly_multiplier": 2.25,
            "use_quarterly": True,
            "year_limits": {
                "2025": {"minimal_wage": 4666.0},
                "2026": {"minimal_wage": 4666.0},
            }
        }
    }

    def __init__(self, path="config.json"):
        self.path = path
        if os.path.exists(path):
            with open(path, encoding="utf-8") as f:
                self._d = json.load(f)
        else:
            self._d = dict(self.DEFAULTS)
            self._save()

    def _save(self):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self._d, f, ensure_ascii=False, indent=2)

    def get(self, key, default=None):
        return self._d.get(key, default)

    def set(self, key, value):
        self._d[key] = value
        self._save()

    def get_db_path(self):      return self._d.get("database_path", "data.db")
    def set_db_path(self, p):
        self._d["database_path"] = p
        self._d["last_opened"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        self._save()

    def get_business_info(self):        return self._d.get("business_info", {})
    def update_business_info(self, d):  self._d["business_info"] = d; self._save()

    def get_invoice_config(self):       return self._d.get("invoice_config", self.DEFAULTS["invoice_config"])
    def update_invoice_config(self, d): self._d["invoice_config"] = d; self._save()

    def get_next_invoice_number(self):
        prefix = self._d.get("invoice_prefix", "R")
        year   = datetime.now().year
        if self._d.get("invoice_year") != year:
            self._d["invoice_counter"] = 1
            self._d["invoice_year"] = year
        n = self._d["invoice_counter"]
        self._d["invoice_counter"] += 1
        self._save()
        return f"{prefix}/{n:04d}/{year}"

    def reset_invoice_counter(self):
        self._d["invoice_counter"] = 1
        self._d["invoice_year"] = datetime.now().year
        self._save()

    def should_save_pdf(self): return self._d.get("save_pdf", True)

    def get_limits(self):         return self._d.get("limits", self.DEFAULTS["limits"])
    def update_limits(self, d):   self._d["limits"] = d; self._save()

    def get_minimal_wage(self, year=None):
        if year is None: year = datetime.now().year
        lim = self.get_limits()
        yl  = lim.get("year_limits", {})
        return yl.get(str(year), {}).get("minimal_wage", lim.get("minimal_wage", 4666))

    def use_quarterly_limits(self):   return self.get_limits().get("use_quarterly", True)
    def get_quarterly_multiplier(self): return self.get_limits().get("quarterly_multiplier", 2.25)


# ─────────────────────────────────────────────────────────
#  BAZA DANYCH
# ─────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font as XFont, PatternFill as XFill, Alignment as XAlign
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class DB:
    def __init__(self, path="data.db"):
        self.path = path
        self.conn = sqlite3.connect(path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._migrate()

    def _migrate(self):
        c = self.conn.cursor()
        c.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT UNIQUE, title TEXT, stock INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS purchase_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT, total_pln REAL, date TEXT
        );
        CREATE TABLE IF NOT EXISTS purchase_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER, product_id INTEGER,
            qty INTEGER, unit_cost REAL DEFAULT 0, available_qty INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS purchase_stock_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_item_id INTEGER, product_id INTEGER,
            qty INTEGER, date TEXT, sale_order_id INTEGER DEFAULT NULL
        );
        CREATE TABLE IF NOT EXISTS sales_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT, total_pln REAL, total_eur REAL,
            purchase_cost REAL DEFAULT 0, date TEXT
        );
        CREATE TABLE IF NOT EXISTS sales_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER, product_id INTEGER, qty INTEGER
        );
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_number TEXT UNIQUE, sale_order_id INTEGER,
            file_path TEXT, customer_name TEXT, customer_address TEXT,
            issue_date TEXT, total_amount REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sale_order_id) REFERENCES sales_orders(id) ON DELETE SET NULL
        );
        """)
        for col, tbl in [("purchase_cost","sales_orders"),("unit_cost","purchase_items"),
                         ("available_qty","purchase_items")]:
            try: c.execute(f"SELECT {col} FROM {tbl} LIMIT 1")
            except sqlite3.OperationalError:
                c.execute(f"ALTER TABLE {tbl} ADD COLUMN {col} REAL DEFAULT 0")
        self.conn.commit()

    # ── PRODUCTS ──
    def add_product(self, sku, title):
        self.conn.execute("INSERT INTO products(sku,title,stock) VALUES(?,?,0)", (sku,title))
        self.conn.commit()

    def check_sku_exists(self, sku):
        return self.conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone() is not None

    def get_product_id_by_sku(self, sku):
        r = self.conn.execute("SELECT id FROM products WHERE sku=?", (sku,)).fetchone()
        return r["id"] if r else None

    def list_products(self):
        return self.conn.execute("SELECT * FROM products ORDER BY title").fetchall()

    def get_product_info(self, pid):
        return self.conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()

    def update_stock(self, pid, delta):
        self.conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (delta,pid))
        self.conn.commit()

    def check_stock(self, pid, qty):
        r = self.conn.execute("SELECT stock FROM products WHERE id=?", (pid,)).fetchone()
        return r and r["stock"] >= qty

    def update_product(self, pid, sku, title):
        self.conn.execute("UPDATE products SET sku=?,title=? WHERE id=?", (sku,title,pid))
        self.conn.commit()

    def delete_product(self, pid):
        p = self.get_product_info(pid)
        if p and p["stock"] > 0: return False
        c = self.conn.cursor()
        c.execute("DELETE FROM purchase_items WHERE product_id=?", (pid,))
        c.execute("DELETE FROM sales_items WHERE product_id=?", (pid,))
        c.execute("DELETE FROM purchase_stock_history WHERE product_id=?", (pid,))
        c.execute("DELETE FROM products WHERE id=?", (pid,))
        self.conn.commit()
        return True

    # ── PURCHASES ──
    def add_purchase_order(self, total_pln, date, items):
        c = self.conn.cursor()
        c.execute("INSERT INTO purchase_orders(total_pln,date) VALUES(?,?)", (total_pln,date))
        oid = c.lastrowid
        total_qty = sum(q for _,q in items)
        for pid, qty in items:
            unit = (total_pln * qty / total_qty) / qty if total_qty > 0 and qty > 0 else 0
            c.execute(
                "INSERT INTO purchase_items(order_id,product_id,qty,unit_cost,available_qty) VALUES(?,?,?,?,?)",
                (oid,pid,qty,unit,qty))
            pi_id = c.lastrowid
            c.execute(
                "INSERT INTO purchase_stock_history(purchase_item_id,product_id,qty,date) VALUES(?,?,?,?)",
                (pi_id,pid,qty,date))
            self.conn.execute("UPDATE products SET stock=stock+? WHERE id=?", (qty,pid))
        self.conn.commit()

    def list_purchases(self):
        return self.conn.execute("""
            SELECT pi.id, p.sku, p.title, pi.qty, po.total_pln, po.date
            FROM purchase_items pi
            JOIN purchase_orders po ON po.id=pi.order_id
            JOIN products p ON p.id=pi.product_id
            ORDER BY po.date DESC
        """).fetchall()

    def delete_purchase(self, item_id):
        c = self.conn.cursor()
        item = c.execute("SELECT * FROM purchase_items WHERE id=?", (item_id,)).fetchone()
        if not item: return
        c.execute("UPDATE products SET stock=stock-? WHERE id=?", (item["qty"],item["product_id"]))
        c.execute("DELETE FROM purchase_stock_history WHERE purchase_item_id=?", (item_id,))
        c.execute("DELETE FROM purchase_items WHERE id=?", (item_id,))
        self.conn.commit()

    def get_fifo_batches(self, pid, qty):
        return self.conn.execute("""
            SELECT pi.id, pi.unit_cost, pi.available_qty
            FROM purchase_items pi
            JOIN purchase_orders po ON po.id=pi.order_id
            WHERE pi.product_id=? AND pi.available_qty>0
            ORDER BY po.date ASC, pi.id ASC LIMIT ?
        """, (pid, qty*10)).fetchall()

    # ── SALES ──
    def add_sale_order(self, platform, total_pln, total_eur, items, fifo_cost, date):
        c = self.conn.cursor()
        c.execute(
            "INSERT INTO sales_orders(platform,total_pln,total_eur,purchase_cost,date) VALUES(?,?,?,?,?)",
            (platform,total_pln,total_eur,fifo_cost,date))
        oid = c.lastrowid
        for pid, qty in items:
            c.execute("INSERT INTO sales_items(order_id,product_id,qty) VALUES(?,?,?)", (oid,pid,qty))
            remaining = qty
            batches = self.conn.execute("""
                SELECT pi.id, pi.available_qty FROM purchase_items pi
                JOIN purchase_orders po ON po.id=pi.order_id
                WHERE pi.product_id=? AND pi.available_qty>0
                ORDER BY po.date ASC, pi.id ASC
            """, (pid,)).fetchall()
            for b in batches:
                if remaining <= 0: break
                take = min(remaining, b["available_qty"])
                self.conn.execute(
                    "UPDATE purchase_items SET available_qty=available_qty-? WHERE id=?", (take,b["id"]))
                remaining -= take
            self.conn.execute("UPDATE products SET stock=stock-? WHERE id=?", (qty,pid))
        self.conn.commit()
        return oid

    def list_sales(self):
        return self.conn.execute("""
            SELECT so.id, so.platform, so.total_pln, so.total_eur,
                   so.purchase_cost,
                   (so.total_pln - so.purchase_cost) AS profit,
                   so.date,
                   GROUP_CONCAT(p.sku || ' x' || si.qty, ', ') AS items
            FROM sales_orders so
            LEFT JOIN sales_items si ON si.order_id=so.id
            LEFT JOIN products p ON p.id=si.product_id
            GROUP BY so.id ORDER BY so.date DESC
        """).fetchall()

    def delete_sale(self, order_id):
        c = self.conn.cursor()
        items = c.execute("SELECT product_id,qty FROM sales_items WHERE order_id=?", (order_id,)).fetchall()
        for item in items:
            c.execute("UPDATE products SET stock=stock+? WHERE id=?", (item["qty"],item["product_id"]))
        c.execute("DELETE FROM sales_items WHERE order_id=?", (order_id,))
        c.execute("DELETE FROM invoices WHERE sale_order_id=?", (order_id,))
        c.execute("DELETE FROM sales_orders WHERE id=?", (order_id,))
        self.conn.commit()

    def get_detailed_sales(self, date_from, date_to):
        return self.conn.execute("""
            SELECT so.id AS order_id, so.platform, so.date,
                   so.total_pln AS order_total_pln,
                   so.total_eur AS order_total_eur,
                   so.purchase_cost AS order_total_cost,
                   p.sku, p.title, si.qty,
                   so.total_pln AS item_revenue_pln,
                   so.purchase_cost AS item_cost,
                   (so.total_pln - so.purchase_cost) AS item_profit
            FROM sales_orders so
            JOIN sales_items si ON si.order_id=so.id
            JOIN products p ON p.id=si.product_id
            WHERE so.date BETWEEN ? AND ?
            ORDER BY so.date, so.id
        """, (date_from,date_to)).fetchall()

    # ── INVOICES ──
    def add_invoice(self, invoice_number, sale_id, file_path, customer_name, customer_address, amount):
        self.conn.execute("""
            INSERT INTO invoices(invoice_number,sale_order_id,file_path,customer_name,
                                 customer_address,issue_date,total_amount)
            VALUES(?,?,?,?,?,?,?)
        """, (invoice_number,sale_id,file_path,customer_name,customer_address,
              datetime.now().strftime("%Y-%m-%d"),amount))
        self.conn.commit()

    def list_invoices(self, date_from=None, date_to=None):
        if date_from and date_to:
            return self.conn.execute("""
                SELECT i.*, so.platform FROM invoices i
                LEFT JOIN sales_orders so ON so.id=i.sale_order_id
                WHERE i.issue_date BETWEEN ? AND ?
                ORDER BY i.created_at DESC
            """, (date_from,date_to)).fetchall()
        return self.conn.execute("""
            SELECT i.*, so.platform FROM invoices i
            LEFT JOIN sales_orders so ON so.id=i.sale_order_id
            ORDER BY i.created_at DESC
        """).fetchall()

    def delete_invoice(self, iid):
        self.conn.execute("DELETE FROM invoices WHERE id=?", (iid,))
        self.conn.commit()

    # ── STATS ──
    def get_stats(self, year=None):
        if year is None: year = datetime.now().year
        r = self.conn.execute("""
            SELECT COUNT(*) AS sc,
                   COALESCE(SUM(total_pln),0) AS rev,
                   COALESCE(SUM(total_pln-purchase_cost),0) AS profit,
                   COALESCE(SUM(purchase_cost),0) AS cost
            FROM sales_orders WHERE strftime('%Y',date)=?
        """, (str(year),)).fetchone()
        pc = self.conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        ts = self.conn.execute("SELECT COALESCE(SUM(stock),0) FROM products").fetchone()[0]
        return {"sale_count":r["sc"],"revenue":r["rev"],"profit":r["profit"],
                "cost":r["cost"],"prod_count":pc,"total_stock":ts}

    def get_monthly_revenue(self, year=None):
        if year is None: year = datetime.now().year
        rows = self.conn.execute("""
            SELECT strftime('%m',date) AS m,
                   COALESCE(SUM(total_pln),0) AS rev,
                   COALESCE(SUM(total_pln-purchase_cost),0) AS profit
            FROM sales_orders WHERE strftime('%Y',date)=?
            GROUP BY m ORDER BY m
        """, (str(year),)).fetchall()
        return {r["m"]:{"rev":r["rev"],"profit":r["profit"]} for r in rows}

    def get_platform_breakdown(self, year=None):
        if year is None: year = datetime.now().year
        return self.conn.execute("""
            SELECT platform,
                   COUNT(*) AS cnt,
                   COALESCE(SUM(total_pln),0) AS rev,
                   COALESCE(SUM(total_pln-purchase_cost),0) AS profit
            FROM sales_orders WHERE strftime('%Y',date)=?
            GROUP BY platform ORDER BY rev DESC
        """, (str(year),)).fetchall()

    def get_platform_sales_count(self, platform, year=None):
        """Liczba sprzedaży na danej platformie w bieżącym roku"""
        if year is None: year = datetime.now().year
        r = self.conn.execute("""
            SELECT COUNT(*) AS cnt FROM sales_orders
            WHERE platform=? AND strftime('%Y',date)=?
        """, (platform,str(year))).fetchone()
        return r["cnt"] if r else 0

    def backup(self, dest_path):
        self.conn.execute("PRAGMA wal_checkpoint(FULL)")
        shutil.copy2(self.path, dest_path)
        return dest_path

    def export_csv(self, path, date_from, date_to):
        sales = self.get_detailed_sales(date_from, date_to)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(["Data","Platforma","SKU","Nazwa","Ilość",
                        "Przychód PLN","Koszt PLN","Zysk PLN"])
            for s in sales:
                profit = s["order_total_pln"] - s["order_total_cost"]
                w.writerow([s["date"],s["platform"],s["sku"],s["title"],
                            s["qty"],f"{s['order_total_pln']:.2f}",
                            f"{s['order_total_cost']:.2f}",f"{profit:.2f}"])
        return True


# ─────────────────────────────────────────────────────────
#  POMOCNICZE
# ─────────────────────────────────────────────────────────
def product_combo(db):
    combo = QComboBox()
    for p in db.list_products():
        combo.addItem(f"{p['sku']} – {p['title']}  (stan: {p['stock']})", p["id"])
    return combo


def T():
    """Zwraca aktualny motyw"""
    return CURRENT_THEME


class SortableTable(QTableWidget):
    def __init__(self, rows=0, cols=0, parent=None):
        super().__init__(rows, cols, parent)
        self._sort_asc = {}
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().sectionClicked.connect(self._hdr)

    def _hdr(self, col):
        asc = not self._sort_asc.get(col, False)
        self._sort_asc[col] = asc
        self.sortItems(col, Qt.AscendingOrder if asc else Qt.DescendingOrder)


def Separator(parent=None):
    f = QFrame(parent)
    f.setFrameShape(QFrame.HLine)
    f.setStyleSheet(f"background:{T()['border']};max-height:1px;")
    return f


def btn(text, style="primary", parent=None):
    b = QPushButton(text, parent)
    if style == "secondary": b.setProperty("secondary","true")
    elif style == "danger":  b.setProperty("danger","true")
    elif style == "success": b.setProperty("success","true")
    return b


# ─────────────────────────────────────────────────────────
#  WYKRES SŁUPKOWY – miesięczne przychody / zysk
# ─────────────────────────────────────────────────────────
class MonthlyBarChart(QWidget):
    MONTHS = ["Sty","Lut","Mar","Kwi","Maj","Cze","Lip","Sie","Wrz","Paź","Lis","Gru"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._data   = {}
        self._show   = "rev"   # "rev" lub "profit"
        self.setMinimumHeight(160)

    def set_data(self, data):
        self._data = data
        self.update()

    def set_mode(self, mode):
        self._show = mode
        self.update()

    def paintEvent(self, event):
        t  = T()
        p  = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        W, H = self.width(), self.height()
        PL, PR, PT, PB = 48, 10, 10, 28

        chart_w = W - PL - PR
        chart_h = H - PT - PB

        rev_vals    = [self._data.get(f"{m:02d}",{}).get("rev",0) for m in range(1,13)]
        profit_vals = [self._data.get(f"{m:02d}",{}).get("profit",0) for m in range(1,13)]
        vals        = profit_vals if self._show == "profit" else rev_vals
        max_val     = max(max(vals), 1)

        bar_col   = QColor(t["chart_bar"])
        prof_col  = QColor(t["success"])
        label_col = QColor(t["chart_label"])
        grid_col  = QColor(t["chart_grid"])
        text_col  = QColor(t["text"])

        # siatka pozioma
        p.setPen(QPen(grid_col, 1, Qt.DotLine))
        for i in range(5):
            y = PT + chart_h - (i * chart_h / 4)
            p.drawLine(PL, int(y), W - PR, int(y))
            # etykieta osi Y
            val_label = f"{int(max_val * i / 4)}"
            p.setPen(QPen(label_col))
            p.setFont(QFont("Segoe UI", 7))
            p.drawText(0, int(y) - 8, PL - 4, 16, Qt.AlignRight | Qt.AlignVCenter, val_label)
            p.setPen(QPen(grid_col, 1, Qt.DotLine))

        # słupki
        bar_w = chart_w / 12
        gap   = bar_w * 0.18
        for i, (rv, pv) in enumerate(zip(rev_vals, profit_vals)):
            val = pv if self._show == "profit" else rv
            x   = PL + i * bar_w + gap / 2
            bw  = bar_w - gap

            # tło (max)
            p.setBrush(QBrush(QColor(t["chart_grid"])))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(int(x), PT, int(bw), chart_h, 3, 3)

            # wartość
            bh = max(int(val / max_val * chart_h), 0)
            by = PT + chart_h - bh
            col = prof_col if self._show == "profit" else bar_col
            p.setBrush(QBrush(col))
            p.drawRoundedRect(int(x), int(by), int(bw), bh, 3, 3)

            # wartość nad słupkiem
            if val > 0:
                p.setPen(QPen(text_col))
                p.setFont(QFont("Segoe UI", 7))
                label = f"{int(val/1000)}k" if val >= 1000 else str(int(val))
                p.drawText(int(x) - 4, int(by) - 14, int(bw) + 8, 13, Qt.AlignCenter, label)

            # miesiąc
            p.setPen(QPen(label_col))
            p.setFont(QFont("Segoe UI", 8))
            p.drawText(int(x), H - PB + 4, int(bw), 20, Qt.AlignCenter, self.MONTHS[i])

        p.end()


# ─────────────────────────────────────────────────────────
#  WYKRES – limit platform (kołowy / pasek)
# ─────────────────────────────────────────────────────────
class PlatformLimitChart(QWidget):
    """Pasek postępu z etykietami dla każdej platformy"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._platforms = []  # list of (name, used, limit, rev, profit)
        self.setMinimumHeight(30)

    def set_data(self, platforms):
        self._platforms = platforms
        h = max(len(platforms) * 48 + 20, 60)
        self.setMinimumHeight(h)
        self.update()

    def paintEvent(self, event):
        if not self._platforms:
            return
        t  = T()
        p  = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        W, H   = self.width(), self.height()
        row_h  = H / len(self._platforms)
        bar_h  = 14
        PL     = 130   # lewa kolumna – nazwa
        PR     = 10
        bar_w  = W - PL - PR - 110  # 110 px na tekst po prawej

        bar_ok     = QColor(T()["success"])
        bar_warn   = QColor(T()["warning"])
        bar_danger = QColor(T()["danger"])
        bar_bg     = QColor(T()["chart_grid"])
        text_main  = QColor(T()["text"])
        text_dim   = QColor(T()["text3"])

        for i, (name, used, limit, rev, profit) in enumerate(self._platforms):
            cy = i * row_h + row_h / 2

            # nazwa platformy
            p.setPen(QPen(text_main))
            p.setFont(QFont("Segoe UI", 10, QFont.Bold))
            p.drawText(0, int(cy - row_h/2), PL - 8, int(row_h), Qt.AlignRight | Qt.AlignVCenter, name)

            # tło paska
            bx = PL
            by = int(cy - bar_h / 2)
            p.setBrush(QBrush(bar_bg))
            p.setPen(Qt.NoPen)
            p.drawRoundedRect(bx, by, int(bar_w), bar_h, 5, 5)

            # wypełnienie
            pct = min(used / limit, 1.0) if limit > 0 else 0
            fill_w = int(bar_w * pct)
            color = bar_ok if pct < 0.70 else (bar_warn if pct < 0.90 else bar_danger)
            p.setBrush(QBrush(color))
            p.drawRoundedRect(bx, by, max(fill_w, 0), bar_h, 5, 5)

            # tekst: X/29 sprzedaży i pozostało
            tx = bx + bar_w + 8
            remaining = max(limit - used, 0)
            p.setPen(QPen(text_main))
            p.setFont(QFont("Segoe UI", 9, QFont.Bold))
            p.drawText(tx, by - 1, 110, bar_h, Qt.AlignLeft | Qt.AlignVCenter,
                       f"{used}/{limit}  (–{remaining})")

            # kwoty pod paskiem
            p.setPen(QPen(text_dim))
            p.setFont(QFont("Segoe UI", 8))
            p.drawText(bx, by + bar_h + 2, int(bar_w), 14, Qt.AlignLeft,
                       f"Przychód: {rev:,.2f} PLN   Zysk netto: {profit:,.2f} PLN")

        p.end()


# ─────────────────────────────────────────────────────────
#  KARTA KPI
# ─────────────────────────────────────────────────────────
class KpiCard(QFrame):
    def __init__(self, label, value, unit="", color=None, parent=None):
        super().__init__(parent)
        self.setFrameShape(QFrame.StyledPanel)
        self._color = color or T()["accent"]
        self.setMinimumWidth(150)

        v = QVBoxLayout(self)
        v.setContentsMargins(14, 12, 14, 12)
        v.setSpacing(3)

        self._lbl = QLabel(label)
        self._lbl.setStyleSheet(f"color:{T()['text3']};font-size:10px;font-weight:700;"
                                 "text-transform:uppercase;letter-spacing:0.5px;")
        v.addWidget(self._lbl)

        self._val = QLabel(str(value))
        self._val.setStyleSheet(f"color:{self._color};font-size:21px;font-weight:800;")
        v.addWidget(self._val)

        if unit:
            self._unit = QLabel(unit)
            self._unit.setStyleSheet(f"color:{T()['text3']};font-size:10px;")
            v.addWidget(self._unit)

        self._update_bg()

    def _update_bg(self):
        t = T()
        self.setStyleSheet(
            f"QFrame{{background:{t['card_bg']};border:1px solid {t['border']};"
            f"border-radius:8px;border-left:4px solid {self._color};}}"
        )

    def set_value(self, v):
        self._val.setText(str(v))

    def refresh_theme(self, color=None):
        if color: self._color = color
        self._lbl.setStyleSheet(f"color:{T()['text3']};font-size:10px;font-weight:700;")
        self._val.setStyleSheet(f"color:{self._color};font-size:21px;font-weight:800;")
        self._update_bg()


# ─────────────────────────────────────────────────────────
#  DASHBOARD
# ─────────────────────────────────────────────────────────
class DashboardWidget(QWidget):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db     = db
        self.config = config
        self._build()

    def _build(self):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)

        inner = QWidget()
        v = QVBoxLayout(inner)
        v.setSpacing(16)
        v.setContentsMargins(16,16,16,16)

        # ── nagłówek ──
        hdr = QHBoxLayout()
        yr_lbl = QLabel(f"Dashboard – rok {datetime.now().year}")
        yr_lbl.setStyleSheet(f"font-size:16px;font-weight:800;color:{T()['text']};")
        hdr.addWidget(yr_lbl)
        hdr.addStretch()

        # przełącznik motywu
        self.theme_btn = btn("🌙 Motyw nocny", "secondary")
        self.theme_btn.clicked.connect(self._toggle_theme)
        hdr.addWidget(self.theme_btn)
        ref_btn = btn("⟳ Odśwież", "secondary")
        ref_btn.clicked.connect(self.refresh)
        hdr.addWidget(ref_btn)
        v.addLayout(hdr)

        # ── KPI cards ──
        kpi_row = QHBoxLayout(); kpi_row.setSpacing(10)
        t = T()
        self.kpi_rev   = KpiCard("Przychód roczny",       "—", "PLN",       t["kpi_rev"])
        self.kpi_prof  = KpiCard("Zysk netto roczny",     "—", "PLN",       t["kpi_profit"])
        self.kpi_cost  = KpiCard("Koszt zakupów",         "—", "PLN",       t["kpi_stock"])
        self.kpi_cnt   = KpiCard("Liczba sprzedaży",      "—", "transakcji",t["kpi_sales"])
        self.kpi_prod  = KpiCard("Produkty w magazynie",  "—", "SKU",       t["kpi_prod"])
        for k in [self.kpi_rev, self.kpi_prof, self.kpi_cost, self.kpi_cnt, self.kpi_prod]:
            kpi_row.addWidget(k)
        v.addLayout(kpi_row)

        # ── limit US ──
        lim_grp = QGroupBox("Limit działalności nierejestrowanej (bieżący rok)")
        ll = QVBoxLayout()
        self.lim_bar  = QProgressBar()
        self.lim_bar.setFormat("")
        self.lim_bar.setFixedHeight(16)
        self.lim_info = QLabel()
        self.lim_info.setStyleSheet(f"color:{T()['text2']};font-size:11px;")
        ll.addWidget(self.lim_bar)
        ll.addWidget(self.lim_info)
        lim_grp.setLayout(ll)
        v.addWidget(lim_grp)

        # ── wykresy ──
        charts = QHBoxLayout(); charts.setSpacing(12)

        # miesięczny wykres
        chart_grp = QGroupBox("Miesięczne wyniki (bieżący rok)")
        cl = QVBoxLayout()
        mode_row = QHBoxLayout()
        mode_row.addWidget(QLabel("Pokaż:"))
        self.mode_rev  = QRadioButton("Przychód"); self.mode_rev.setChecked(True)
        self.mode_prof = QRadioButton("Zysk netto")
        self.mode_rev.toggled.connect(lambda c: self.chart.set_mode("rev") if c else None)
        self.mode_prof.toggled.connect(lambda c: self.chart.set_mode("profit") if c else None)
        mode_row.addWidget(self.mode_rev)
        mode_row.addWidget(self.mode_prof)
        mode_row.addStretch()
        cl.addLayout(mode_row)
        self.chart = MonthlyBarChart()
        self.chart.setMinimumHeight(170)
        cl.addWidget(self.chart)
        chart_grp.setLayout(cl)
        charts.addWidget(chart_grp, 3)

        v.addLayout(charts)

        # ── platformy z limitem 29 sprzedaży ──
        plat_grp = QGroupBox(f"Sprzedaż wg platform – limit {PLATFORM_LIMIT} szt./rok/platforma")
        pl = QVBoxLayout()

        # legenda
        leg = QHBoxLayout()
        for color, txt in [(T()["success"],"≤70% limitu"), (T()["warning"],"70–90%"), (T()["danger"],"≥90% – uwaga!")]:
            dot = QLabel("●")
            dot.setStyleSheet(f"color:{color};font-size:16px;")
            leg.addWidget(dot)
            leg.addWidget(QLabel(txt))
            leg.addSpacing(16)
        leg.addStretch()
        pl.addLayout(leg)

        self.plat_chart = PlatformLimitChart()
        pl.addWidget(self.plat_chart)

        # tabela podsumowania
        self.plat_tbl = SortableTable(0, 6)
        self.plat_tbl.setHorizontalHeaderLabels([
            "Platforma","Sprzedaży","Pozostało","Przychód PLN","Zysk netto PLN","% limitu"])
        for i, w in enumerate([QHeaderView.Stretch,
                                QHeaderView.ResizeToContents,
                                QHeaderView.ResizeToContents,
                                QHeaderView.ResizeToContents,
                                QHeaderView.ResizeToContents,
                                QHeaderView.ResizeToContents]):
            self.plat_tbl.horizontalHeader().setSectionResizeMode(i, w)
        self.plat_tbl.setFixedHeight(140)
        pl.addWidget(self.plat_tbl)
        plat_grp.setLayout(pl)
        v.addWidget(plat_grp)

        v.addStretch()
        scroll.setWidget(inner)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0,0,0,0)
        outer.addWidget(scroll)

        self._inner = inner
        self.refresh()

    def _toggle_theme(self):
        global CURRENT_THEME
        app = QApplication.instance()
        if CURRENT_THEME is THEME_DAY:
            apply_theme(app, THEME_NIGHT)
            self.theme_btn.setText("☀️ Motyw dzienny")
        else:
            apply_theme(app, THEME_DAY)
            self.theme_btn.setText("🌙 Motyw nocny")
        # odśwież wykres
        self.chart.update()
        self.plat_chart.update()

    def refresh(self):
        t     = T()
        stats = self.db.get_stats()
        self.kpi_rev.set_value(f"{stats['revenue']:,.2f}")
        self.kpi_prof.set_value(f"{stats['profit']:,.2f}")
        self.kpi_cost.set_value(f"{stats['cost']:,.2f}")
        self.kpi_cnt.set_value(str(stats['sale_count']))
        self.kpi_prod.set_value(str(stats['prod_count']))

        # limit bar
        wage  = self.config.get_minimal_wage()
        lim   = self.config.get_limits()
        if lim.get("use_quarterly", True):
            limit = wage * lim.get("quarterly_multiplier", 2.25)
            now   = datetime.now()
            q     = (now.month - 1) // 3 + 1
            label = f"Limit kwartalny Q{q}/{now.year}"
        else:
            limit = wage * 0.75
            label = f"Limit miesięczny {datetime.now().month}/{datetime.now().year}"

        rev = stats["revenue"]
        pct = min(int(rev / limit * 100), 100) if limit > 0 else 0
        self.lim_bar.setValue(pct)
        bar_color = t["success"] if pct < 70 else (t["warning"] if pct < 90 else t["danger"])
        self.lim_bar.setStyleSheet(
            f"QProgressBar::chunk{{background-color:{bar_color};border-radius:5px;}}"
            f"QProgressBar{{background:{t['progress_bg']};border:1px solid {t['border']};border-radius:5px;}}"
        )
        warn = "  ⚠️  Zbliżasz się do limitu!" if pct >= 80 else ""
        self.lim_info.setText(
            f"{label}: {rev:,.2f} PLN / {limit:,.2f} PLN  ({pct}%){warn}"
        )
        self.lim_info.setStyleSheet(f"color:{t['text2']};font-size:11px;")

        # wykres miesięczny
        monthly = self.db.get_monthly_revenue()
        self.chart.set_data(monthly)

        # platformy
        platforms = self.db.get_platform_breakdown()
        year = datetime.now().year

        plat_data = []
        self.plat_tbl.setRowCount(len(platforms))
        for i, pd in enumerate(platforms):
            name   = pd["platform"]
            used   = self.db.get_platform_sales_count(name, year)
            remain = max(PLATFORM_LIMIT - used, 0)
            rev_p  = pd["rev"]
            profit_p = pd["profit"]
            pct_p  = min(int(used / PLATFORM_LIMIT * 100), 100)

            plat_data.append((name, used, PLATFORM_LIMIT, rev_p, profit_p))

            vals = [name, str(used), str(remain),
                    f"{rev_p:,.2f}", f"{profit_p:,.2f}", f"{pct_p}%"]
            for j, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if j in [1,2,3,4,5]:
                    item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                # kolorowanie wiersza wg limitu
                if pct_p >= 90:
                    item.setForeground(QColor(t["danger"]))
                elif pct_p >= 70:
                    item.setForeground(QColor(t["warning"]))
                elif pct_p > 0:
                    item.setForeground(QColor(t["success"]))
                self.plat_tbl.setItem(i, j, item)

        self.plat_chart.set_data(plat_data)


# ─────────────────────────────────────────────────────────
#  PANEL PRODUKTÓW
# ─────────────────────────────────────────────────────────
class ProductsWidget(QWidget):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db     = db
        self.config = config
        self._build()
        self.refresh()

    def _build(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(12,12,12,12)
        v.setSpacing(8)

        tb = QHBoxLayout()
        add_b = btn("➕ Dodaj produkt"); add_b.clicked.connect(self._add)
        edi_b = btn("✏️ Edytuj","secondary"); edi_b.clicked.connect(self._edit)
        del_b = btn("🗑 Usuń","danger");     del_b.clicked.connect(self._delete)
        inv_b = btn("📋 Inwentaryzacja","secondary"); inv_b.clicked.connect(self._inventory)
        for b2 in [add_b, edi_b, del_b, inv_b]: tb.addWidget(b2)
        tb.addStretch()
        self.search = QLineEdit(); self.search.setPlaceholderText("🔍 Szukaj...")
        self.search.setFixedWidth(220); self.search.textChanged.connect(self._filter)
        tb.addWidget(self.search)
        v.addLayout(tb)

        self.tbl = SortableTable(0,5)
        self.tbl.setHorizontalHeaderLabels(["ID","SKU","Nazwa","Stan",""])
        hdr = self.tbl.horizontalHeader()
        for i,m in enumerate([QHeaderView.ResizeToContents, QHeaderView.ResizeToContents,
                               QHeaderView.Stretch, QHeaderView.ResizeToContents,
                               QHeaderView.ResizeToContents]):
            hdr.setSectionResizeMode(i,m)
        v.addWidget(self.tbl)

        self.status = QLabel()
        self.status.setStyleSheet(f"color:{T()['text3']};font-size:11px;")
        v.addWidget(self.status)

    def refresh(self):
        t = T()
        prods = self.db.list_products()
        self.tbl.setRowCount(len(prods))
        for i, p in enumerate(prods):
            for j, val in enumerate([p["id"],p["sku"],p["title"],p["stock"]]):
                item = QTableWidgetItem(str(val))
                if j in [0,3]: item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                if j == 3 and p["stock"] == 0:
                    item.setForeground(QColor(t["danger"]))
                self.tbl.setItem(i,j,item)
            xb = QPushButton("✕"); xb.setFixedSize(26,26)
            xb.setStyleSheet(f"background:{t['bg2']};color:{t['text3']};border-radius:4px;padding:0;font-size:11px;")
            xb.clicked.connect(lambda _, pid=p["id"]: self._quick_del(pid))
            self.tbl.setCellWidget(i,4,xb)
        self._filter(self.search.text())
        self.status.setText(f"Produktów: {len(prods)}")

    def _filter(self,text):
        text = text.lower()
        for r in range(self.tbl.rowCount()):
            m = any(text in (self.tbl.item(r,c).text().lower() if self.tbl.item(r,c) else "") for c in range(4))
            self.tbl.setRowHidden(r, not m)

    def _add(self):
        d = ProductDialog(self.db, parent=self)
        if d.exec(): self.refresh()

    def _edit(self):
        row = self.tbl.currentRow()
        if row < 0: QMessageBox.warning(self,"Brak wyboru","Kliknij wiersz produktu."); return
        pid = int(self.tbl.item(row,0).text())
        p   = self.db.get_product_info(pid)
        if p:
            if ProductDialog(self.db, dict(p), parent=self).exec(): self.refresh()

    def _delete(self):
        row = self.tbl.currentRow()
        if row < 0: QMessageBox.warning(self,"Brak wyboru","Kliknij wiersz."); return
        self._quick_del(int(self.tbl.item(row,0).text()))

    def _quick_del(self, pid):
        p = self.db.get_product_info(pid)
        if not p: return
        if QMessageBox.question(self,"Usuń",f"Usunąć: {p['sku']} – {p['title']}?",
                                QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            if not self.db.delete_product(pid):
                QMessageBox.warning(self,"Błąd","Nie można usunąć produktu z dodatnim stanem.")
            else: self.refresh()

    def _inventory(self):
        InventoryDialog(self.db,self).exec(); self.refresh()


# ─────────────────────────────────────────────────────────
#  DIALOGI
# ─────────────────────────────────────────────────────────
class ProductDialog(QDialog):
    def __init__(self, db, product=None, parent=None):
        super().__init__(parent)
        self.db = db; self.product = product
        self.setWindowTitle("Edytuj produkt" if product else "Nowy produkt")
        self.setFixedSize(420,260)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("📦  " + ("Edytuj produkt" if product else "Dodaj nowy produkt"),
                           styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        self.sku   = QLineEdit(product["sku"] if product else ""); self.sku.setPlaceholderText("np. PROD-001")
        self.title = QLineEdit(product["title"] if product else ""); self.title.setPlaceholderText("Nazwa produktu")
        self.stock = QSpinBox(); self.stock.setRange(0,999999)
        if product: self.stock.setValue(product["stock"]); self.stock.setEnabled(False)
        form.addRow("SKU:",self.sku); form.addRow("Nazwa:",self.title); form.addRow("Stan pocz.:",self.stock)
        v.addLayout(form); v.addStretch()
        btns = QHBoxLayout()
        ok = btn("Zapisz","success"); ok.clicked.connect(self._save)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)

    def _save(self):
        sku = self.sku.text().strip(); title = self.title.text().strip()
        if not sku or not title: QMessageBox.warning(self,"Błąd","SKU i Nazwa są wymagane."); return
        try:
            if self.product:
                self.db.update_product(self.product["id"],sku,title)
            else:
                if self.db.check_sku_exists(sku): QMessageBox.warning(self,"Błąd",f"SKU '{sku}' już istnieje."); return
                self.db.add_product(sku,title)
                if self.stock.value()>0:
                    pid = self.db.get_product_id_by_sku(sku)
                    self.db.add_purchase_order(0.0,datetime.now().strftime("%Y-%m-%d"),[(pid,self.stock.value())])
            self.accept()
        except Exception as e:
            QMessageBox.critical(self,"Błąd",str(e))


class PurchaseDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db; self.setWindowTitle("Nowy zakup"); self.resize(560,420)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("📦  Rejestruj zakup", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        self.cost = QDoubleSpinBox(); self.cost.setMaximum(1_000_000); self.cost.setDecimals(2); self.cost.setSuffix(" PLN")
        self.date_e = QDateEdit(QDate.currentDate()); self.date_e.setCalendarPopup(True)
        form.addRow("Koszt łączny:",self.cost); form.addRow("Data:",self.date_e); v.addLayout(form)
        v.addWidget(QLabel("Pozycje zamówienia", styleSheet=f"color:{T()['text3']};font-size:11px;font-weight:600;"))
        self.items_tbl = SortableTable(0,2)
        self.items_tbl.setHorizontalHeaderLabels(["Produkt","Ilość"])
        self.items_tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.items_tbl.horizontalHeader().setSectionResizeMode(1,QHeaderView.ResizeToContents)
        v.addWidget(self.items_tbl)
        add_btn = btn("➕ Dodaj pozycję","secondary"); add_btn.clicked.connect(self._add_row); v.addWidget(add_btn)
        v.addWidget(Separator(self))
        btns = QHBoxLayout()
        ok = btn("💾 Zapisz zakup","success"); ok.clicked.connect(self._save)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)
        self._add_row()
        self.result_items = []; self.result_cost = 0; self.result_date = ""

    def _add_row(self):
        r = self.items_tbl.rowCount(); self.items_tbl.insertRow(r)
        self.items_tbl.setCellWidget(r,0,product_combo(self.db))
        qty = QSpinBox(); qty.setRange(1,999999); self.items_tbl.setCellWidget(r,1,qty)

    def _save(self):
        items = [(self.items_tbl.cellWidget(r,0).currentData(), self.items_tbl.cellWidget(r,1).value())
                 for r in range(self.items_tbl.rowCount())
                 if self.items_tbl.cellWidget(r,0) and self.items_tbl.cellWidget(r,1)]
        if not items: QMessageBox.warning(self,"Błąd","Dodaj przynajmniej jedną pozycję."); return
        self.result_items = items; self.result_cost = self.cost.value()
        self.result_date  = self.date_e.date().toString("yyyy-MM-dd")
        self.accept()


class SaleDialog(QDialog):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db = db; self.config = config
        self.setWindowTitle("Nowa sprzedaż"); self.resize(620,560)
        self._fifo = 0.0
        self._build()

    def _build(self):
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("💰  Rejestruj sprzedaż", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        self.platform = QComboBox(); self.platform.addItems(PLATFORMS)
        self.platform.currentTextChanged.connect(self._on_platform)
        form.addRow("Platforma:",self.platform)
        self.custom_plat = QLineEdit(); self.custom_plat.setPlaceholderText("Wpisz nazwę"); self.custom_plat.setVisible(False)
        form.addRow("Inna platforma:",self.custom_plat)
        self.pln = QDoubleSpinBox(); self.pln.setMaximum(1_000_000); self.pln.setDecimals(2); self.pln.setSuffix(" PLN")
        form.addRow("Cena sprzedaży:",self.pln)
        self.date_e = QDateEdit(QDate.currentDate()); self.date_e.setCalendarPopup(True)
        form.addRow("Data:",self.date_e)
        self.fifo_lbl = QLabel("Koszt FIFO: 0.00 PLN"); self.fifo_lbl.setStyleSheet(f"color:{T()['accent']};font-weight:700;")
        form.addRow("",self.fifo_lbl)

        # ostrzeżenie limitu
        self.limit_warn = QLabel("")
        self.limit_warn.setStyleSheet(f"color:{T()['danger']};font-weight:700;font-size:11px;")
        form.addRow("",self.limit_warn)
        self.platform.currentTextChanged.connect(self._check_limit)
        self._check_limit(self.platform.currentText())

        v.addLayout(form)
        self.inv_cb = QCheckBox("Wygeneruj rachunek PDF"); v.addWidget(self.inv_cb)
        self.inv_cb.stateChanged.connect(self._toggle_inv)
        self.client_grp = QGroupBox("Dane klienta"); self.client_grp.setVisible(False)
        cl = QFormLayout()
        self.client_name = QLineEdit(); self.client_name.setPlaceholderText("Imię i Nazwisko")
        self.client_addr = QLineEdit(); self.client_addr.setPlaceholderText("Adres")
        cl.addRow("Nabywca:",self.client_name); cl.addRow("Adres:",self.client_addr)
        self.client_grp.setLayout(cl); v.addWidget(self.client_grp)
        v.addWidget(QLabel("Pozycje sprzedaży", styleSheet=f"color:{T()['text3']};font-size:11px;font-weight:600;"))
        self.items_tbl = SortableTable(0,2); self.items_tbl.setHorizontalHeaderLabels(["Produkt","Ilość"])
        self.items_tbl.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        self.items_tbl.horizontalHeader().setSectionResizeMode(1,QHeaderView.ResizeToContents)
        v.addWidget(self.items_tbl)
        add_btn = btn("➕ Dodaj pozycję","secondary"); add_btn.clicked.connect(self._add_row); v.addWidget(add_btn)
        v.addWidget(Separator(self))
        btns = QHBoxLayout()
        ok_inv = btn("💾 Zapisz i rachunek","success"); ok_inv.clicked.connect(lambda: self._do_save(True))
        ok     = btn("Zapisz bez rachunku","secondary"); ok.clicked.connect(lambda: self._do_save(False))
        ca     = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok_inv); btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)
        self._add_row()

    def _on_platform(self, txt):
        self.custom_plat.setVisible(txt == "Inne")
        self._check_limit(txt)

    def _check_limit(self, platform_txt):
        if platform_txt == "Inne": platform_txt = "Inne"
        used = self.db.get_platform_sales_count(platform_txt)
        remaining = PLATFORM_LIMIT - used
        if remaining <= 0:
            self.limit_warn.setText(f"⛔  LIMIT {PLATFORM_LIMIT} SPRZEDAŻY NA PLATFORMIE OSIĄGNIĘTY!")
        elif remaining <= 5:
            self.limit_warn.setText(f"⚠️  Pozostało tylko {remaining} sprzedaży z {PLATFORM_LIMIT} na tej platformie!")
            self.limit_warn.setStyleSheet(f"color:{T()['warning']};font-weight:700;font-size:11px;")
        else:
            self.limit_warn.setText(f"✓  Wykorzystano: {used}/{PLATFORM_LIMIT}  (pozostało: {remaining})")
            self.limit_warn.setStyleSheet(f"color:{T()['success']};font-size:11px;")

    def _toggle_inv(self, state): self.client_grp.setVisible(bool(state))

    def _plat_name(self):
        if self.platform.currentText() == "Inne":
            return self.custom_plat.text().strip() or "Inne"
        return self.platform.currentText()

    def _add_row(self):
        r = self.items_tbl.rowCount(); self.items_tbl.insertRow(r)
        combo = product_combo(self.db); combo.currentIndexChanged.connect(self._update_fifo)
        self.items_tbl.setCellWidget(r,0,combo)
        qty = QSpinBox(); qty.setRange(1,999999); qty.valueChanged.connect(self._update_fifo)
        self.items_tbl.setCellWidget(r,1,qty)
        QTimer.singleShot(100,self._update_fifo)

    def _get_items(self):
        return [(self.items_tbl.cellWidget(r,0).currentData(), self.items_tbl.cellWidget(r,1).value())
                for r in range(self.items_tbl.rowCount())
                if self.items_tbl.cellWidget(r,0) and self.items_tbl.cellWidget(r,1)]

    def _update_fifo(self):
        total = 0.0
        for pid, qty in self._get_items():
            batches = self.db.get_fifo_batches(pid,qty)
            rem = qty
            for b in batches:
                take = min(rem,b["available_qty"]); total += b["unit_cost"]*take; rem -= take
                if rem <= 0: break
        self._fifo = total
        self.fifo_lbl.setText(f"Koszt FIFO: {total:.2f} PLN  |  Zysk netto: {self.pln.value()-total:.2f} PLN")

    def _do_save(self, with_invoice):
        items = self._get_items()
        if not items: QMessageBox.warning(self,"Błąd","Dodaj przynajmniej jedną pozycję."); return
        for pid, qty in items:
            if not self.db.check_stock(pid,qty):
                QMessageBox.warning(self,"Brak towaru",f"Niewystarczający stan produktu ID {pid}."); return
        self._update_fifo()
        pln  = self.pln.value()
        date = self.date_e.date().toString("yyyy-MM-dd")
        eur  = pln / get_eur_rate(date)
        sale_id = self.db.add_sale_order(self._plat_name(),pln,eur,items,self._fifo,date)
        if with_invoice:
            try: self._gen_invoice(sale_id,items,pln,date)
            except Exception as e:
                QMessageBox.warning(self,"Rachunek",f"Sprzedaż zapisana, błąd rachunku:\n{e}")
        self.accept()

    def _gen_invoice(self, sale_id, items, total_pln, date):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            raise ImportError("Zainstaluj reportlab: pip install reportlab")

        font_name = "Helvetica"
        for fp in ["C:/Windows/Fonts/arial.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                   "/usr/share/fonts/truetype/msttcorefonts/Arial.ttf"]:
            if os.path.exists(fp):
                try: pdfmetrics.registerFont(TTFont("PolishFont",fp)); font_name = "PolishFont"; break
                except: pass

        biz    = self.config.get_business_info()
        cfg    = self.config.get_invoice_config()
        inv_no = self.config.get_next_invoice_number()
        pdf_dir = os.path.join(os.getcwd(),"rachunki")
        os.makedirs(pdf_dir,exist_ok=True)
        path = os.path.join(pdf_dir,f"rachunek_{inv_no.replace('/','_')}.pdf")

        doc   = SimpleDocTemplate(path, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm,
                                  topMargin=2*cm, bottomMargin=2*cm)
        sty   = getSampleStyleSheet()
        def ps(name,**kw): return ParagraphStyle(name,parent=sty["Normal"],fontName=font_name,**kw)
        story = [
            Paragraph(f"RACHUNEK UPROSZCZONY NR {inv_no}", ps("T",fontSize=14,spaceAfter=16,alignment=1)),
            Paragraph(f"Data wystawienia: {date}", ps("N",fontSize=10,spaceAfter=4)),
            Paragraph(f"Sprzedający: {biz.get('name','')}  {biz.get('address','')}  "
                      f"{biz.get('postal_code','')} {biz.get('city','')}", ps("N",fontSize=10,spaceAfter=4)),
            Paragraph(f"PESEL: {biz.get('pesel','')}", ps("N",fontSize=10,spaceAfter=10)),
        ]
        if self.client_name.text():
            story.append(Paragraph(f"Nabywca: {self.client_name.text()}  {self.client_addr.text()}",
                                   ps("N",fontSize=10,spaceAfter=10)))

        tbl_data = [["Produkt","Ilość","Cena jedn.","Wartość"]]
        unit_price = total_pln / sum(q for _,q in items) if items else 0
        for pid,qty in items:
            p = self.db.get_product_info(pid)
            tbl_data.append([p["title"] if p else f"ID{pid}", str(qty),
                             f"{unit_price:.2f} PLN", f"{unit_price*qty:.2f} PLN"])
        tbl_data.append(["","","RAZEM:",f"{total_pln:.2f} PLN"])

        tbl = Table(tbl_data,colWidths=[9*cm,2*cm,4*cm,4*cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#C62828")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,-1),font_name),
            ("FONTSIZE",(0,0),(-1,-1),9),
            ("GRID",(0,0),(-1,-1),0.5,colors.grey),
            ("ALIGN",(1,0),(-1,-1),"CENTER"),
            ("ALIGN",(3,0),(-1,-1),"RIGHT"),
        ]))
        story.append(tbl); story.append(Spacer(1,16))
        if cfg.get("footer_text"):
            story.append(Paragraph(cfg["footer_text"],ps("F",fontSize=9,alignment=1)))
        doc.build(story)
        self.db.add_invoice(inv_no,sale_id,path,self.client_name.text(),self.client_addr.text(),total_pln)

        if QMessageBox.question(self,"Rachunek",f"Rachunek {inv_no} zapisany.\nCzy otworzyć PDF?",
                                QMessageBox.Yes|QMessageBox.No) == QMessageBox.Yes:
            import subprocess, platform as plt
            if plt.system()=="Windows": os.startfile(path)
            elif plt.system()=="Darwin": subprocess.Popen(["open",path])
            else: subprocess.Popen(["xdg-open",path])


class HistoryDialog(QDialog):
    def __init__(self, title, headers, rows, delete_cb, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title); self.resize(1000,550)
        self.rows = rows; self.delete_cb = delete_cb
        v = QVBoxLayout(self); v.setSpacing(8)
        row = QHBoxLayout()
        self.search = QLineEdit(); self.search.setPlaceholderText("🔍  Szukaj...")
        self.search.textChanged.connect(self._filter)
        row.addWidget(self.search); row.addStretch(); v.addLayout(row)
        self.tbl = SortableTable(0,len(headers))
        self.tbl.setHorizontalHeaderLabels(headers)
        for i in range(len(headers)):
            self.tbl.horizontalHeader().setSectionResizeMode(i,QHeaderView.Stretch)
        v.addWidget(self.tbl)
        btns = QHBoxLayout()
        del_b = btn("🗑 Usuń zaznaczony","danger"); del_b.clicked.connect(self._delete)
        cls_b = btn("Zamknij","secondary"); cls_b.clicked.connect(self.accept)
        btns.addWidget(del_b); btns.addStretch(); btns.addWidget(cls_b); v.addLayout(btns)
        self._load(rows)

    def _load(self, rows):
        self.tbl.setRowCount(len(rows))
        for i, r in enumerate(rows):
            for j, val in enumerate(r):
                item = QTableWidgetItem(str(val) if val is not None else "")
                if isinstance(val,(int,float)): item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                self.tbl.setItem(i,j,item)

    def _filter(self,text):
        text = text.lower()
        for r in range(self.tbl.rowCount()):
            m = any(text in (self.tbl.item(r,c).text().lower() if self.tbl.item(r,c) else "") for c in range(self.tbl.columnCount()))
            self.tbl.setRowHidden(r,not m)

    def _delete(self):
        row = self.tbl.currentRow()
        if row < 0: QMessageBox.warning(self,"Brak wyboru","Kliknij wiersz."); return
        oid = int(self.tbl.item(row,0).text())
        if QMessageBox.question(self,"Usuń","Usunąć ten wpis?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.delete_cb(oid); self.accept()


class InventoryDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db; self.setWindowTitle("Inwentaryzacja"); self.resize(700,450)
        v = QVBoxLayout(self)
        info = QLabel("Zmień wartości w kolumnie 'Stan rzeczywisty' i kliknij Zapisz.")
        info.setStyleSheet(f"color:{T()['text3']};font-size:11px;"); v.addWidget(info)
        self.tbl = QTableWidget(0,5)
        self.tbl.setHorizontalHeaderLabels(["ID","SKU","Nazwa","Stan sys.","Stan rzecz."])
        self.tbl.horizontalHeader().setSectionResizeMode(2,QHeaderView.Stretch)
        self.tbl.setAlternatingRowColors(True); v.addWidget(self.tbl)
        btns = QHBoxLayout()
        ok = btn("💾 Zapisz korekty","success"); ok.clicked.connect(self._apply)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)
        self._load()

    def _load(self):
        prods = self.db.list_products(); self.tbl.setRowCount(len(prods))
        for i, p in enumerate(prods):
            for j, val in enumerate([p["id"],p["sku"],p["title"],p["stock"]]):
                item = QTableWidgetItem(str(val)); item.setFlags(item.flags()&~Qt.ItemIsEditable)
                self.tbl.setItem(i,j,item)
            self.tbl.setItem(i,4,QTableWidgetItem(str(p["stock"])))

    def _apply(self):
        ch = 0
        for r in range(self.tbl.rowCount()):
            pid = int(self.tbl.item(r,0).text()); sys_s = int(self.tbl.item(r,3).text())
            try: real_s = int(self.tbl.item(r,4).text())
            except: continue
            delta = real_s - sys_s
            if delta != 0: self.db.update_stock(pid,delta); ch += 1
        QMessageBox.information(self,"OK",f"Zapisano korekty dla {ch} produktów."); self.accept()


class InvoicesDialog(QDialog):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db = db; self.config = config
        self.setWindowTitle("Historia rachunków"); self.resize(1100,550)
        v = QVBoxLayout(self); v.setSpacing(8)
        row = QHBoxLayout()
        row.addWidget(QLabel("Od:"))
        self.df = QDateEdit(QDate.currentDate().addMonths(-1)); self.df.setCalendarPopup(True); row.addWidget(self.df)
        row.addWidget(QLabel("Do:"))
        self.dt = QDateEdit(QDate.currentDate()); self.dt.setCalendarPopup(True); row.addWidget(self.dt)
        fb = btn("Filtruj","secondary"); fb.clicked.connect(self._load); row.addWidget(fb)
        rb = btn("⟳ Reset licznika","secondary"); rb.clicked.connect(self._reset); row.addWidget(rb)
        row.addStretch()
        self.search = QLineEdit(); self.search.setPlaceholderText("🔍 Szukaj..."); self.search.textChanged.connect(self._filter)
        row.addWidget(self.search); v.addLayout(row)
        self.tbl = SortableTable(0,8)
        self.tbl.setHorizontalHeaderLabels(["ID","Numer","Platforma","Klient","Kwota","Data","Plik","Akcja"])
        hdr = self.tbl.horizontalHeader()
        hdr.setSectionResizeMode(1,QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3,QHeaderView.Stretch)
        hdr.setSectionResizeMode(6,QHeaderView.Stretch)
        v.addWidget(self.tbl)
        btns = QHBoxLayout()
        db2 = btn("🗑 Usuń zaznaczony","danger"); db2.clicked.connect(self._delete)
        cb2 = btn("Zamknij","secondary"); cb2.clicked.connect(self.accept)
        btns.addWidget(db2); btns.addStretch(); btns.addWidget(cb2); v.addLayout(btns)
        self._load()

    def _load(self):
        df = self.df.date().toString("yyyy-MM-dd"); dt = self.dt.date().toString("yyyy-MM-dd")
        invs = self.db.list_invoices(df,dt); self.tbl.setRowCount(len(invs))
        for i, inv in enumerate(invs):
            vals = [inv["id"],inv["invoice_number"],inv.get("platform","—"),
                    inv["customer_name"] or "—",f"{inv['total_amount']:.2f} PLN",
                    inv["issue_date"],os.path.basename(inv["file_path"] or ""),""]
            for j, v2 in enumerate(vals):
                item = QTableWidgetItem(str(v2))
                if "PLN" in str(v2): item.setTextAlignment(Qt.AlignRight|Qt.AlignVCenter)
                self.tbl.setItem(i,j,item)
            ob = QPushButton("📂 Otwórz"); ob.setFixedHeight(26)
            fp = inv["file_path"]
            ob.clicked.connect(lambda _,f=fp: self._open(f))
            self.tbl.setCellWidget(i,7,ob)

    def _filter(self,text):
        text = text.lower()
        for r in range(self.tbl.rowCount()):
            m = any(text in (self.tbl.item(r,c).text().lower() if self.tbl.item(r,c) else "") for c in range(self.tbl.columnCount()))
            self.tbl.setRowHidden(r,not m)

    def _open(self,path):
        if not path or not os.path.exists(path):
            QMessageBox.warning(self,"Brak pliku",f"Plik nie istnieje:\n{path}"); return
        import subprocess, platform as plt
        if plt.system()=="Windows": os.startfile(path)
        elif plt.system()=="Darwin": subprocess.Popen(["open",path])
        else: subprocess.Popen(["xdg-open",path])

    def _delete(self):
        row = self.tbl.currentRow()
        if row < 0: QMessageBox.warning(self,"Brak wyboru","Kliknij wiersz."); return
        iid = int(self.tbl.item(row,0).text())
        if QMessageBox.question(self,"Usuń","Usunąć rachunek?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.db.delete_invoice(iid); self._load()

    def _reset(self):
        if QMessageBox.question(self,"Reset","Resetować licznik numeracji?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            self.config.reset_invoice_counter(); QMessageBox.information(self,"OK","Licznik zresetowany.")


class ReportDialog(QDialog):
    """Generator raportów – miesięcznych, kwartalnych, rocznych i za dowolny okres.
    Obsługuje formaty CSV / XLSX / PDF z pełnymi danymi podatkowymi (US)."""

    def __init__(self, db, config, parent=None, report_type="monthly"):
        super().__init__(parent)
        self.db = db; self.config = config; self.report_type = report_type
        titles = {"monthly": "Raport miesięczny", "quarterly": "Raport kwartalny",
                  "yearly": "Raport roczny", "custom": "Raport za okres"}
        self.setWindowTitle(titles.get(report_type, "Raport"))
        self.resize(580, 560)
        self._build_ui()

    def _build_ui(self):
        v = QVBoxLayout(self); v.setSpacing(12)
        v.addWidget(QLabel(f"📊  {self.windowTitle()}",
                           styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))

        # ── zakres ──
        rg = QGroupBox("Zakres okresu"); rl = QHBoxLayout(); rl.setSpacing(8)
        now = datetime.now()
        if self.report_type == "monthly":
            self.month_cb = QComboBox()
            self.month_cb.addItems(["Styczen", "Luty", "Marzec", "Kwiecien", "Maj", "Czerwiec",
                                    "Lipiec", "Sierpien", "Wrzesien", "Pazdziernik", "Listopad", "Grudzien"])
            self.month_cb.setCurrentIndex(now.month - 1)
            self.year_sp = QSpinBox(); self.year_sp.setRange(2000, 2100); self.year_sp.setValue(now.year)
            rl.addWidget(QLabel("Miesiąc:")); rl.addWidget(self.month_cb)
            rl.addWidget(QLabel("Rok:"));     rl.addWidget(self.year_sp)
        elif self.report_type == "quarterly":
            self.qtr_cb = QComboBox()
            self.qtr_cb.addItems(["I kw. (sty-mar)", "II kw. (kwi-cze)",
                                  "III kw. (lip-wrz)", "IV kw. (paz-gru)"])
            self.qtr_cb.setCurrentIndex((now.month - 1) // 3)
            self.year_sp = QSpinBox(); self.year_sp.setRange(2000, 2100); self.year_sp.setValue(now.year)
            rl.addWidget(QLabel("Kwartal:")); rl.addWidget(self.qtr_cb)
            rl.addWidget(QLabel("Rok:"));     rl.addWidget(self.year_sp)
        elif self.report_type == "yearly":
            self.year_sp = QSpinBox(); self.year_sp.setRange(2000, 2100); self.year_sp.setValue(now.year)
            rl.addWidget(QLabel("Rok:")); rl.addWidget(self.year_sp)
        else:
            self.df_e = QDateEdit(QDate.currentDate().addMonths(-1)); self.df_e.setCalendarPopup(True)
            self.dt_e = QDateEdit(QDate.currentDate());                self.dt_e.setCalendarPopup(True)
            rl.addWidget(QLabel("Od:")); rl.addWidget(self.df_e)
            rl.addWidget(QLabel("Do:")); rl.addWidget(self.dt_e)
        rl.addStretch(); rg.setLayout(rl); v.addWidget(rg)

        # ── format ──
        fg = QGroupBox("Format eksportu"); fl = QHBoxLayout()
        self.rb_csv  = QRadioButton("CSV");  self.rb_csv.setChecked(True)
        self.rb_xlsx = QRadioButton("Excel (XLSX)")
        self.rb_pdf  = QRadioButton("PDF")
        for rb in [self.rb_csv, self.rb_xlsx, self.rb_pdf]: fl.addWidget(rb)
        fl.addStretch(); fg.setLayout(fl); v.addWidget(fg)

        # ── opcje ──
        og = QGroupBox("Zawartość raportu"); ol = QVBoxLayout(); ol.setSpacing(6)
        self.cb_sales     = QCheckBox("Sprzedaz (szczegolowa ewidencja)"); self.cb_sales.setChecked(True)
        self.cb_purchases = QCheckBox("Zakupy")
        self.cb_summary   = QCheckBox("Podsumowanie finansowe (przychod / zysk / koszty)"); self.cb_summary.setChecked(True)
        self.cb_us        = QCheckBox("Dane podatkowe US – imie, nazwisko, adres, PESEL, analiza limitu")
        self.cb_us.setChecked(True)   # domyślnie włączone
        self.cb_us.setStyleSheet(f"color:{T()['accent']};font-weight:600;")
        for cb in [self.cb_sales, self.cb_purchases, self.cb_summary, self.cb_us]:
            ol.addWidget(cb)
        og.setLayout(ol); v.addWidget(og)

        # ── stan danych osobowych ──
        biz = self.config.get_business_info()
        if biz.get("name") and biz.get("pesel"):
            biz_status = f"✅  Dane osobowe: {biz['name']}  |  PESEL: {biz['pesel']}"
            biz_color  = T()["success"]
        else:
            biz_status = "⚠️  Brak danych osobowych – uzupełnij w Konfiguracja → Dane osobiste"
            biz_color  = T()["warning"]
        biz_lbl = QLabel(biz_status)
        biz_lbl.setStyleSheet(f"color:{biz_color};font-size:11px;font-weight:600;")
        biz_lbl.setWordWrap(True)
        v.addWidget(biz_lbl)

        # ── info o limitach ──
        wage = self.config.get_minimal_wage()
        lim_lbl = QLabel(
            f"Limit miesieczny: {wage*0.75:.2f} PLN  |  "
            f"Limit kwartalny: {wage*self.config.get_quarterly_multiplier():.2f} PLN  "
            f"(min. wynagrodzenie: {wage:.2f} PLN)")
        lim_lbl.setStyleSheet(f"color:{T()['text3']};font-size:11px;")
        v.addWidget(lim_lbl)

        v.addStretch()
        v.addWidget(Separator(self))
        btns = QHBoxLayout()
        gb = btn("📊 Generuj raport", "success"); gb.clicked.connect(self._generate)
        ca = btn("Anuluj", "secondary");           ca.clicked.connect(self.reject)
        btns.addWidget(gb); btns.addStretch(); btns.addWidget(ca)
        v.addLayout(btns)

    # ── helpers ──────────────────────────────────────────
    def _get_range(self):
        now = datetime.now()
        if self.report_type == "monthly":
            m = self.month_cb.currentIndex() + 1; y = self.year_sp.value()
            last = (datetime(y, m % 12 + 1, 1) - timedelta(days=1)) if m < 12 else datetime(y, 12, 31)
            return f"{y}-{m:02d}-01", last.strftime("%Y-%m-%d")
        elif self.report_type == "quarterly":
            q = self.qtr_cb.currentIndex() + 1; y = self.year_sp.value()
            ends = {1: ("01-01","03-31"), 2: ("04-01","06-30"), 3: ("07-01","09-30"), 4: ("10-01","12-31")}
            s, e = ends[q]; return f"{y}-{s}", f"{y}-{e}"
        elif self.report_type == "yearly":
            y = self.year_sp.value(); return f"{y}-01-01", f"{y}-12-31"
        else:
            return self.df_e.date().toString("yyyy-MM-dd"), self.dt_e.date().toString("yyyy-MM-dd")

    def _report_title(self):
        rt = self.report_type
        if rt == "monthly":
            months_pl = ["Styczen","Luty","Marzec","Kwiecien","Maj","Czerwiec",
                         "Lipiec","Sierpien","Wrzesien","Pazdziernik","Listopad","Grudzien"]
            return f"Raport miesięczny – {months_pl[self.month_cb.currentIndex()]} {self.year_sp.value()}"
        elif rt == "quarterly":
            qtrs = ["I kwartał","II kwartał","III kwartał","IV kwartał"]
            return f"Raport kwartalny – {qtrs[self.qtr_cb.currentIndex()]} {self.year_sp.value()}"
        elif rt == "yearly":
            return f"Raport roczny – {self.year_sp.value()}"
        else:
            df, dt = self._get_range()
            return f"Raport za okres {df} – {dt}"

    def _include_us(self):
        return self.cb_us.isChecked()

    def _get_biz(self):
        return self.config.get_business_info() if self._include_us() else {}

    def _get_limit_info(self, df):
        year = int(df[:4])
        wage = self.config.get_minimal_wage(year)
        lim  = self.config.get_limits()
        use_q = lim.get("use_quarterly", True) and self.report_type == "quarterly"
        if use_q:
            limit = wage * lim.get("quarterly_multiplier", 2.25)
            label = "Limit kwartalny (225% min. wynagrodzenia)"
        else:
            limit = wage * 0.75
            label = "Limit miesieczny (75% min. wynagrodzenia)"
        return {"wage": wage, "limit": limit, "label": label, "year": year}

    # ── generowanie ──────────────────────────────────────
    def _generate(self):
        df, dt = self._get_range()
        if self.rb_csv.isChecked():
            ext, flt = ".csv",  "CSV Files (*.csv)"
        elif self.rb_xlsx.isChecked():
            ext, flt = ".xlsx", "Excel Files (*.xlsx)"
        else:
            ext, flt = ".pdf",  "PDF Files (*.pdf)"

        suggested = f"raport_{df.replace('-','')}_{dt.replace('-','')}{ext}"
        path, _ = QFileDialog.getSaveFileName(self, "Zapisz raport",
                                              os.path.join(os.getcwd(), suggested), flt)
        if not path: return
        try:
            if ext == ".csv":   ok = self._gen_csv(path, df, dt)
            elif ext == ".xlsx": ok = self._gen_xlsx(path, df, dt)
            else:               ok = self._gen_pdf(path, df, dt)
            if ok:
                QMessageBox.information(self, "Sukces", f"Raport zapisany:\n{path}")
                self.accept()
            else:
                QMessageBox.warning(self, "Blad", "Nie udalo sie wygenerowac raportu.")
        except Exception as e:
            QMessageBox.critical(self, "Blad", f"{type(e).__name__}:\n{str(e)}")

    # ──────────────────────────────────────────────────────
    #  CSV
    # ──────────────────────────────────────────────────────
    def _gen_csv(self, path, df, dt):
        biz   = self._get_biz()
        linfo = self._get_limit_info(df)
        sales = self.db.get_detailed_sales(df, dt) if self.cb_sales.isChecked() else []
        purch = self.db.list_purchases()            if self.cb_purchases.isChecked() else []

        # grupuj sprzedaż po zamówieniu
        orders = {}
        for s in sales:
            oid = s["order_id"]
            if oid not in orders:
                orders[oid] = {"date": s["date"], "platform": s["platform"],
                               "pln": s["order_total_pln"], "cost": s["order_total_cost"],
                               "products": []}
            orders[oid]["products"].append(f"{s['sku']} x{s['qty']}")
        order_list = list(orders.values())

        total_rev    = sum(o["pln"]  for o in order_list)
        total_cost   = sum(o["cost"] for o in order_list)
        total_profit = total_rev - total_cost

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")
            title = self._report_title()
            w.writerow([title])
            w.writerow([f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            w.writerow([f"Okres: {df} - {dt}"])
            w.writerow([f"Program: {APP_NAME} v{APP_VERSION}"])
            w.writerow([])

            # dane sprzedawcy
            if biz.get("name"):
                w.writerow(["=== DANE SPRZEDAWCY (DZIALALNOSC NIEREJESTROWANA) ==="])
                w.writerow(["Imie i nazwisko:", biz.get("name","")])
                w.writerow(["Adres:",           f"{biz.get('address','')} {biz.get('postal_code','')} {biz.get('city','')}"])
                w.writerow(["PESEL:",            biz.get("pesel","")])
                if biz.get("nip"):   w.writerow(["NIP:",   biz.get("nip","")])
                if biz.get("regon"): w.writerow(["REGON:", biz.get("regon","")])
                w.writerow([])

            # podsumowanie
            if self.cb_summary.isChecked():
                w.writerow(["=== PODSUMOWANIE FINANSOWE ==="])
                w.writerow(["Przychod calkowity PLN:", f"{total_rev:.2f}"])
                w.writerow(["Koszt zakupow PLN:",      f"{total_cost:.2f}"])
                w.writerow(["Zysk netto PLN:",         f"{total_profit:.2f}"])
                w.writerow(["Liczba transakcji:",       len(order_list)])
                w.writerow([])

            # analiza limitu US
            if self._include_us():
                w.writerow(["=== ANALIZA LIMITU US ==="])
                w.writerow(["Rok:", linfo["year"]])
                w.writerow(["Minimalne wynagrodzenie PLN:", f"{linfo['wage']:.2f}"])
                w.writerow([f"{linfo['label']}:", f"{linfo['limit']:.2f}"])
                w.writerow(["Przychod w okresie PLN:", f"{total_rev:.2f}"])
                pct = total_rev / linfo["limit"] * 100 if linfo["limit"] > 0 else 0
                w.writerow(["Uzycie limitu %:", f"{pct:.1f}%"])
                status = "PRZEKROCZONO LIMIT!" if total_rev > linfo["limit"] else "W granicach limitu"
                w.writerow(["Status:", status])
                w.writerow([])

            # szczegółowa sprzedaż
            if self.cb_sales.isChecked():
                w.writerow(["=== EWIDENCJA SPRZEDAZY ==="])
                w.writerow(["Data", "Platforma", "Produkty (SKU x ilosc)", "Przychod PLN",
                            "Koszt PLN", "Zysk netto PLN"])
                for o in order_list:
                    w.writerow([o["date"], o["platform"], ", ".join(o["products"]),
                                f"{o['pln']:.2f}", f"{o['cost']:.2f}",
                                f"{o['pln']-o['cost']:.2f}"])
                w.writerow([])

            # zakupy
            if self.cb_purchases.isChecked() and purch:
                w.writerow(["=== EWIDENCJA ZAKUPOW ==="])
                w.writerow(["ID", "SKU", "Nazwa", "Ilosc", "Koszt PLN", "Data"])
                for p in purch:
                    w.writerow([p["id"], p["sku"], p["title"], p["qty"],
                                f"{p['total_pln']:.2f}", p["date"]])
        return True

    # ──────────────────────────────────────────────────────
    #  XLSX
    # ──────────────────────────────────────────────────────
    def _gen_xlsx(self, path, df, dt):
        if not HAS_EXCEL:
            raise ImportError("Zainstaluj openpyxl:  pip install openpyxl")
        import openpyxl
        from openpyxl.styles import Font as F, PatternFill as PF, Alignment as AL, Border, Side
        from openpyxl.utils import get_column_letter

        biz   = self._get_biz()
        linfo = self._get_limit_info(df)
        sales = self.db.get_detailed_sales(df, dt) if self.cb_sales.isChecked() else []
        purch = self.db.list_purchases()            if self.cb_purchases.isChecked() else []

        # grupuj zamówienia
        orders = {}
        for s in sales:
            oid = s["order_id"]
            if oid not in orders:
                orders[oid] = {"date": s["date"], "platform": s["platform"],
                               "pln": s["order_total_pln"], "cost": s["order_total_cost"],
                               "products": []}
            orders[oid]["products"].append(f"{s['sku']} x{s['qty']}")
        order_list = list(orders.values())

        total_rev    = sum(o["pln"]  for o in order_list)
        total_cost   = sum(o["cost"] for o in order_list)
        total_profit = total_rev - total_cost

        wb = openpyxl.Workbook()

        # ── style ──
        RED   = "C62828"; WHT = "FFFFFF"; GRY = "F5F5F5"; DRK = "1A1A1A"
        LGRY  = "EEEEEE"; GRNG = "2E7D32"; ORNG = "E65100"

        def hdr_font(bold=True, color=WHT, sz=10):
            return F(name="Calibri", bold=bold, color=color, size=sz)
        def hdr_fill(color=RED):
            return PF(start_color=color, end_color=color, fill_type="solid")
        def centered():
            return AL(horizontal="center", vertical="center", wrap_text=True)
        def border():
            s = Side(style="thin", color="CCCCCC")
            return Border(left=s, right=s, top=s, bottom=s)
        def money_fmt(): return '#,##0.00 "PLN"'
        def pct_fmt():   return '0.0"%"'

        # ══ ARKUSZ 1: Ewidencja sprzedaży ══
        ws = wb.active; ws.title = "Ewidencja sprzedazy"
        ws.sheet_properties.tabColor = RED

        row = 1
        # nagłówek raportu
        ws.merge_cells(f"A{row}:H{row}")
        c = ws.cell(row=row, column=1, value=self._report_title())
        c.font = F(name="Calibri", bold=True, size=14, color=RED); row += 1

        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value=f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {APP_NAME} v{APP_VERSION}")
        ws.cell(row=row, column=1).font = F(name="Calibri", italic=True, size=9, color="777777"); row += 1

        ws.merge_cells(f"A{row}:H{row}")
        ws.cell(row=row, column=1, value=f"Okres: {df}  do  {dt}")
        ws.cell(row=row, column=1).font = F(name="Calibri", bold=True, size=10); row += 2

        # dane sprzedawcy US
        if biz.get("name"):
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1, value="DANE SPRZEDAWCY – DZIALALNOSC NIEREJESTROWANA")
            c.font = hdr_font(color=WHT, sz=10); c.fill = hdr_fill(RED); c.alignment = centered(); row += 1
            dane = [
                ("Imie i nazwisko:",  biz.get("name","")),
                ("Adres zamieszkania:", f"{biz.get('address','')}"),
                ("Kod pocztowy / Miasto:", f"{biz.get('postal_code','')}  {biz.get('city','')}"),
                ("PESEL:",            biz.get("pesel","")),
            ]
            if biz.get("nip"):   dane.append(("NIP:", biz.get("nip","")))
            if biz.get("regon"): dane.append(("REGON:", biz.get("regon","")))
            for label, value in dane:
                ws.cell(row=row, column=1, value=label).font = F(name="Calibri", bold=True, size=10)
                ws.cell(row=row, column=2, value=value).font  = F(name="Calibri", size=10)
                ws.cell(row=row, column=1).fill = hdr_fill("F5F5F5"); ws.cell(row=row, column=1).font = F(bold=True, color=DRK, size=10)
                row += 1
            row += 1

        # podsumowanie finansowe
        if self.cb_summary.isChecked():
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1, value="PODSUMOWANIE FINANSOWE")
            c.font = hdr_font(sz=10); c.fill = hdr_fill(RED); c.alignment = centered(); row += 1
            summary = [
                ("Przychod calkowity",  total_rev,    GRNG),
                ("Koszt zakupow",       total_cost,   ORNG),
                ("Zysk netto",          total_profit, GRNG if total_profit >= 0 else "C62828"),
                ("Liczba transakcji",   len(order_list), None),
            ]
            for label, value, color in summary:
                ws.cell(row=row, column=1, value=label).font = F(name="Calibri", bold=True, size=10)
                vc = ws.cell(row=row, column=2, value=value)
                vc.font = F(name="Calibri", bold=True, size=11, color=color or DRK)
                if isinstance(value, float): vc.number_format = money_fmt()
                row += 1
            row += 1

        # analiza limitu US
        if self._include_us():
            ws.merge_cells(f"A{row}:H{row}")
            c = ws.cell(row=row, column=1, value="ANALIZA LIMITU DZIALALNOSCI NIEREJESTROWANEJ")
            c.font = hdr_font(sz=10); c.fill = hdr_fill(RED); c.alignment = centered(); row += 1
            pct = total_rev / linfo["limit"] * 100 if linfo["limit"] > 0 else 0
            over = total_rev > linfo["limit"]
            lim_rows = [
                ("Rok podatkowy",               str(linfo["year"])),
                ("Minimalne wynagrodzenie PLN",  f"{linfo['wage']:.2f}"),
                (linfo["label"],                 f"{linfo['limit']:.2f}"),
                ("Przychod w okresie PLN",       f"{total_rev:.2f}"),
                ("Uzycie limitu",                f"{pct:.1f}%"),
                ("STATUS",  "PRZEKROCZONO LIMIT! Wymagana rejestracja DG!" if over else "W granicach limitu"),
            ]
            for label, value in lim_rows:
                ws.cell(row=row, column=1, value=label).font = F(name="Calibri", bold=True, size=10)
                vc = ws.cell(row=row, column=2, value=value)
                if label == "STATUS":
                    vc.font = F(name="Calibri", bold=True, size=10,
                                color=RED if over else GRNG)
                row += 1
            row += 1

        # tabela transakcji
        if self.cb_sales.isChecked():
            hdrs = ["Data", "Platforma", "Produkty (SKU x ilosc)", "Przychod PLN",
                    "Koszt PLN", "Zysk netto PLN", "Margin %"]
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = hdr_font(sz=9); c.fill = hdr_fill(RED); c.alignment = centered(); c.border = border()
            row += 1
            for oi, o in enumerate(order_list):
                margin = (o["pln"] - o["cost"]) / o["pln"] * 100 if o["pln"] > 0 else 0
                fill = hdr_fill("FFFFFF") if oi % 2 == 0 else hdr_fill(GRY)
                vals = [o["date"], o["platform"], ", ".join(o["products"]),
                        o["pln"], o["cost"], o["pln"] - o["cost"], margin / 100]
                for ci, val in enumerate(vals, 1):
                    c = ws.cell(row=row, column=ci, value=val)
                    c.fill = fill; c.border = border()
                    c.font = F(name="Calibri", size=9)
                    if ci == 4: c.number_format = money_fmt()
                    if ci == 5: c.number_format = money_fmt(); c.font = F(name="Calibri", size=9, color=ORNG)
                    if ci == 6:
                        c.number_format = money_fmt()
                        c.font = F(name="Calibri", bold=True, size=9,
                                   color=GRNG if o["pln"] >= o["cost"] else RED)
                    if ci == 7: c.number_format = '0.0"%"'
                    c.alignment = AL(horizontal="center" if ci in [1,2,4,5,6,7] else "left",
                                     vertical="center", wrap_text=ci==3)
                row += 1
            # wiersz sumy
            ws.cell(row=row, column=3, value="SUMA").font = F(name="Calibri", bold=True, size=10)
            for ci, val in [(4, total_rev), (5, total_cost), (6, total_profit)]:
                c = ws.cell(row=row, column=ci, value=val)
                c.number_format = money_fmt(); c.fill = hdr_fill("FFE0B2")
                c.font = F(name="Calibri", bold=True, size=10, color=RED)
            row += 2

        # arkusz zakupów
        if self.cb_purchases.isChecked() and purch:
            ws2 = wb.create_sheet("Zakupy"); ws2.sheet_properties.tabColor = "1565C0"
            ws2.merge_cells("A1:F1")
            c = ws2.cell(row=1, column=1, value=f"Ewidencja zakupow – {df} do {dt}")
            c.font = F(name="Calibri", bold=True, size=12, color="1565C0")
            hdrs2 = ["ID", "SKU", "Nazwa", "Ilosc", "Koszt PLN", "Data"]
            for ci, h in enumerate(hdrs2, 1):
                c = ws2.cell(row=2, column=ci, value=h)
                c.font = F(name="Calibri", bold=True, color=WHT)
                c.fill = hdr_fill("1565C0"); c.alignment = centered(); c.border = border()
            for ri, p in enumerate(purch, 3):
                vals = [p["id"], p["sku"], p["title"], p["qty"], p["total_pln"], p["date"]]
                for ci, val in enumerate(vals, 1):
                    c = ws2.cell(row=ri, column=ci, value=val)
                    c.font = F(name="Calibri", size=9); c.border = border()
                    if ci == 5: c.number_format = money_fmt()
                    c.fill = hdr_fill("FFFFFF") if ri % 2 == 0 else hdr_fill(GRY)
            for col in ws2.columns:
                ws2.column_dimensions[get_column_letter(col[0].column)].width = 16

        # szerokości kolumn arkusza głównego
        col_widths = [12, 14, 40, 16, 16, 16, 10]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"

        wb.save(path)
        return True

    # ──────────────────────────────────────────────────────
    #  PDF  – polskie znaki przez encoding lub DejaVu
    # ──────────────────────────────────────────────────────
    def _gen_pdf(self, path, df, dt):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer, HRFlowable, KeepTogether)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm, mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        except ImportError:
            raise ImportError("Zainstaluj reportlab:  pip install reportlab")

        # ── czcionka z polskimi znakami ──────────────────
        # Próbujemy wbudowaną DejaVuSans z reportlab (dostępna od v3.6)
        font_name  = "Helvetica"
        font_bold  = "Helvetica-Bold"

        _font_candidates = [
            # Windows
            ("C:/Windows/Fonts/arial.ttf",         "C:/Windows/Fonts/arialbd.ttf"),
            ("C:/Windows/Fonts/calibri.ttf",        "C:/Windows/Fonts/calibrib.ttf"),
            # Linux
            ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
             "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
            ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
             "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
            ("/usr/share/fonts/truetype/freefont/FreeSans.ttf",
             "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf"),
            # macOS
            ("/Library/Fonts/Arial.ttf", "/Library/Fonts/Arial Bold.ttf"),
            ("/System/Library/Fonts/Helvetica.ttc", None),
        ]

        # próba wbudowanej czcionki reportlab z obsługą Latin-2
        try:
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
        except Exception:
            pass

        for regular, bold in _font_candidates:
            if regular and os.path.exists(regular):
                try:
                    pdfmetrics.registerFont(TTFont("RaportFont", regular))
                    if bold and os.path.exists(bold):
                        pdfmetrics.registerFont(TTFont("RaportFont-Bold", bold))
                        font_bold = "RaportFont-Bold"
                    else:
                        font_bold = "RaportFont"
                    font_name = "RaportFont"
                    break
                except Exception:
                    continue

        # ── dane ────────────────────────────────────────
        biz   = self._get_biz()
        linfo = self._get_limit_info(df)
        sales = self.db.get_detailed_sales(df, dt) if self.cb_sales.isChecked() else []
        purch = self.db.list_purchases()            if self.cb_purchases.isChecked() else []

        orders = {}
        for s in sales:
            oid = s["order_id"]
            if oid not in orders:
                orders[oid] = {"date": s["date"], "platform": s["platform"],
                               "pln": s["order_total_pln"], "cost": s["order_total_cost"],
                               "products": []}
            orders[oid]["products"].append(f"{s['sku']} x{s['qty']}")
        order_list = list(orders.values())

        total_rev    = sum(o["pln"]  for o in order_list)
        total_cost   = sum(o["cost"] for o in order_list)
        total_profit = total_rev - total_cost

        # ── kolory ──────────────────────────────────────
        C_RED   = colors.HexColor("#C62828")
        C_RED2  = colors.HexColor("#FFCDD2")
        C_GRN   = colors.HexColor("#2E7D32")
        C_GRN2  = colors.HexColor("#E8F5E9")
        C_ORG   = colors.HexColor("#E65100")
        C_GRY   = colors.HexColor("#F5F5F5")
        C_GRY2  = colors.HexColor("#EEEEEE")
        C_DRK   = colors.HexColor("#1A1A1A")
        C_MED   = colors.HexColor("#555555")
        C_LIT   = colors.HexColor("#888888")
        C_BLBG  = colors.HexColor("#E3F2FD")

        # ── style paragrafów ────────────────────────────
        def ps(name, fn=None, fb=None, **kw):
            return ParagraphStyle(name,
                fontName=fn or font_name,
                **kw)

        sTitle  = ps("sTitle", fn=font_bold, fontSize=16, textColor=C_RED,
                     spaceAfter=4, alignment=TA_CENTER)
        sSub    = ps("sSub",   fn=font_name, fontSize=9,  textColor=C_MED,
                     spaceAfter=2, alignment=TA_CENTER)
        sSecHdr = ps("sHdr",   fn=font_bold, fontSize=10, textColor=C_RED,
                     spaceBefore=12, spaceAfter=4)
        sNorm   = ps("sNorm",  fn=font_name, fontSize=9,  textColor=C_DRK,
                     spaceAfter=2, leading=13)
        sBold   = ps("sBold",  fn=font_bold, fontSize=9,  textColor=C_DRK,
                     spaceAfter=2)
        sSmall  = ps("sSmall", fn=font_name, fontSize=7,  textColor=C_LIT,
                     spaceAfter=1, alignment=TA_CENTER)
        sWarn   = ps("sWarn",  fn=font_bold, fontSize=10, textColor=C_RED,
                     spaceAfter=4, alignment=TA_CENTER)
        sOK     = ps("sOK",    fn=font_bold, fontSize=10, textColor=C_GRN,
                     spaceAfter=4, alignment=TA_CENTER)

        # ── budowanie dokumentu ─────────────────────────
        doc = SimpleDocTemplate(
            path, pagesize=A4,
            rightMargin=1.8*cm, leftMargin=1.8*cm,
            topMargin=2*cm,     bottomMargin=2*cm,
            title=self._report_title(),
            author=biz.get("name", APP_NAME)
        )

        story = []
        W = A4[0] - 3.6*cm  # szerokość użytkowa

        # nagłówek dokumentu
        story.append(Paragraph(self._report_title(), sTitle))
        story.append(Paragraph(
            f"Okres: {df}  –  {dt}  |  Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sSub))
        story.append(Paragraph(f"{APP_NAME} v{APP_VERSION}", sSmall))
        story.append(HRFlowable(width="100%", thickness=2, color=C_RED, spaceAfter=8))

        # ── dane sprzedawcy US ──
        if biz.get("name"):
            story.append(Paragraph("DANE SPRZEDAWCY – DZIALALNOSC NIEREJESTROWANA", sSecHdr))
            biz_rows = [
                ["Imie i nazwisko:",        biz.get("name","")],
                ["Adres zamieszkania:",     biz.get("address","")],
                ["Kod pocztowy / Miasto:",  f"{biz.get('postal_code','')}  {biz.get('city','')}"],
                ["PESEL:",                  biz.get("pesel","")],
            ]
            if biz.get("nip"):   biz_rows.append(["NIP:",   biz.get("nip","")])
            if biz.get("regon"): biz_rows.append(["REGON:", biz.get("regon","")])
            biz_tbl = Table(biz_rows, colWidths=[5*cm, W-5*cm])
            biz_tbl.setStyle(TableStyle([
                ("FONTNAME",   (0,0), (-1,-1), font_name),
                ("FONTNAME",   (0,0), (0,-1),  font_bold),
                ("FONTSIZE",   (0,0), (-1,-1), 9),
                ("TEXTCOLOR",  (0,0), (0,-1),  C_MED),
                ("TEXTCOLOR",  (1,0), (1,-1),  C_DRK),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_GRY, colors.white]),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("RIGHTPADDING", (0,0), (-1,-1), 6),
                ("TOPPADDING",   (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
                ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
                ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.HexColor("#EEEEEE")),
            ]))
            story.append(biz_tbl)
            story.append(Spacer(1, 8))

        # ── podsumowanie finansowe ──
        if self.cb_summary.isChecked():
            story.append(Paragraph("PODSUMOWANIE FINANSOWE", sSecHdr))
            sum_data = [
                ["Przychod calkowity",  f"{total_rev:,.2f} PLN",    ""],
                ["Koszt zakupow",       f"{total_cost:,.2f} PLN",   ""],
                ["Zysk netto",          f"{total_profit:,.2f} PLN", ""],
                ["Liczba transakcji",   str(len(order_list)),       ""],
            ]
            margin_pct = (total_profit / total_rev * 100) if total_rev > 0 else 0
            sum_data[2][2] = f"marza: {margin_pct:.1f}%"
            sum_tbl = Table(sum_data, colWidths=[6*cm, 5*cm, W-11*cm])
            sum_tbl.setStyle(TableStyle([
                ("FONTNAME",   (0,0), (-1,-1), font_name),
                ("FONTNAME",   (0,0), (0,-1),  font_bold),
                ("FONTNAME",   (1,0), (1,0),   font_bold),  # przychód bold
                ("FONTNAME",   (1,2), (1,2),   font_bold),  # zysk bold
                ("FONTSIZE",   (0,0), (-1,-1), 10),
                ("TEXTCOLOR",  (0,0), (0,-1),  C_MED),
                ("TEXTCOLOR",  (1,0), (1,0),   C_GRN),
                ("TEXTCOLOR",  (1,1), (1,1),   C_ORG),
                ("TEXTCOLOR",  (1,2), (1,2),   C_GRN if total_profit >= 0 else C_RED),
                ("TEXTCOLOR",  (1,3), (1,3),   C_DRK),
                ("TEXTCOLOR",  (2,0), (-1,-1), C_MED),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_GRY, colors.white]),
                ("LEFTPADDING",  (0,0), (-1,-1), 8),
                ("RIGHTPADDING", (0,0), (-1,-1), 8),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0), (-1,-1), 5),
                ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
                ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.HexColor("#EEEEEE")),
            ]))
            story.append(sum_tbl)
            story.append(Spacer(1, 8))

        # ── analiza limitu US ──
        if self._include_us():
            story.append(Paragraph("ANALIZA LIMITU DZIALALNOSCI NIEREJESTROWANEJ", sSecHdr))
            pct = total_rev / linfo["limit"] * 100 if linfo["limit"] > 0 else 0
            over = total_rev > linfo["limit"]
            lim_data = [
                ["Rok podatkowy",            str(linfo["year"])],
                ["Minimalne wynagrodzenie",   f"{linfo['wage']:,.2f} PLN"],
                [linfo["label"],              f"{linfo['limit']:,.2f} PLN"],
                ["Przychod w okresie",        f"{total_rev:,.2f} PLN"],
                ["Pozostalo do limitu",       f"{max(linfo['limit']-total_rev,0):,.2f} PLN"],
                ["Uzycie limitu",             f"{pct:.1f}%"],
            ]
            lim_tbl = Table(lim_data, colWidths=[7*cm, W-7*cm])
            lim_tbl.setStyle(TableStyle([
                ("FONTNAME",   (0,0), (-1,-1), font_name),
                ("FONTNAME",   (0,0), (0,-1),  font_bold),
                ("FONTSIZE",   (0,0), (-1,-1), 9),
                ("TEXTCOLOR",  (0,0), (0,-1),  C_MED),
                ("TEXTCOLOR",  (1,3), (1,3),   C_RED if over else C_GRN),
                ("TEXTCOLOR",  (1,5), (1,5),   C_RED if pct > 90 else (C_ORG if pct > 70 else C_GRN)),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_GRY, colors.white]),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("TOPPADDING",   (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
                ("BOX",        (0,0), (-1,-1), 0.5, colors.HexColor("#DDDDDD")),
                ("INNERGRID",  (0,0), (-1,-1), 0.3, colors.HexColor("#EEEEEE")),
            ]))
            story.append(lim_tbl)
            story.append(Spacer(1, 6))

            # baner statusu
            if over:
                story.append(Paragraph(
                    "UWAGA: PRZEKROCZONO LIMIT DZIALALNOSCI NIEREJESTROWANEJ!  "
                    "Konieczna rejestracja dzialalnosci gospodarczej.", sWarn))
            else:
                remaining = linfo["limit"] - total_rev
                story.append(Paragraph(
                    f"Status: W granicach limitu  |  Pozostalo: {remaining:,.2f} PLN  ({100-pct:.1f}%)", sOK))
            story.append(Spacer(1, 8))

        # ── tabela transakcji ──
        if self.cb_sales.isChecked() and order_list:
            story.append(Paragraph("SZCZEGOLOWA EWIDENCJA SPRZEDAZY", sSecHdr))
            col_w = [2.2*cm, 3*cm, 0, 2.8*cm, 2.8*cm, 2.8*cm]
            col_w[2] = W - sum(col_w[:2]) - sum(col_w[3:])

            hdr_row = ["Data", "Platforma", "Produkty (SKU x ilosc)",
                       "Przychod PLN", "Koszt PLN", "Zysk netto PLN"]
            tbl_data = [hdr_row]
            for o in order_list:
                tbl_data.append([
                    o["date"], o["platform"],
                    ", ".join(o["products"]),
                    f"{o['pln']:,.2f}", f"{o['cost']:,.2f}",
                    f"{o['pln']-o['cost']:,.2f}"
                ])
            # suma
            tbl_data.append(["", "SUMA", "",
                              f"{total_rev:,.2f}", f"{total_cost:,.2f}",
                              f"{total_profit:,.2f}"])

            n = len(tbl_data)
            ts = TableStyle([
                # nagłówek
                ("FONTNAME",    (0,0), (-1,0),  font_bold),
                ("FONTSIZE",    (0,0), (-1,0),  8),
                ("BACKGROUND",  (0,0), (-1,0),  C_RED),
                ("TEXTCOLOR",   (0,0), (-1,0),  colors.white),
                ("ALIGN",       (0,0), (-1,0),  "CENTER"),
                ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
                # dane
                ("FONTNAME",    (0,1), (-1,-1), font_name),
                ("FONTSIZE",    (0,1), (-1,-1), 8),
                ("TEXTCOLOR",   (0,1), (-1,-1), C_DRK),
                ("ALIGN",       (3,1), (5,-1),  "RIGHT"),
                ("ALIGN",       (0,1), (1,-1),  "CENTER"),
                ("ROWBACKGROUNDS", (0,1), (-1,n-2), [colors.white, C_GRY]),
                # wiersz sumy
                ("FONTNAME",    (0,n-1), (-1,n-1), font_bold),
                ("FONTSIZE",    (0,n-1), (-1,n-1), 9),
                ("BACKGROUND",  (0,n-1), (-1,n-1), colors.HexColor("#FFE0B2")),
                ("TEXTCOLOR",   (3,n-1), (3,n-1), C_GRN),
                ("TEXTCOLOR",   (4,n-1), (4,n-1), C_ORG),
                ("TEXTCOLOR",   (5,n-1), (5,n-1), C_GRN if total_profit >= 0 else C_RED),
                # siatka
                ("GRID",        (0,0), (-1,-1), 0.3, colors.HexColor("#DDDDDD")),
                ("LINEBELOW",   (0,0), (-1,0),  1.0, C_RED),
                ("LINEABOVE",   (0,n-1),(-1,n-1),1.0, C_RED),
                # padding
                ("LEFTPADDING",  (0,0), (-1,-1), 4),
                ("RIGHTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING",   (0,0), (-1,-1), 4),
                ("BOTTOMPADDING",(0,0), (-1,-1), 4),
            ])
            # kolorowanie zysku
            for ri, o in enumerate(order_list, 1):
                profit = o["pln"] - o["cost"]
                if profit < 0:
                    ts.add("TEXTCOLOR", (5,ri), (5,ri), C_RED)
                else:
                    ts.add("TEXTCOLOR", (5,ri), (5,ri), C_GRN)

            sales_tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
            sales_tbl.setStyle(ts)
            story.append(sales_tbl)
            story.append(Spacer(1, 10))

        # ── tabela zakupów ──
        if self.cb_purchases.isChecked() and purch:
            story.append(Paragraph("EWIDENCJA ZAKUPOW", sSecHdr))
            pcol_w = [1.5*cm, 2.5*cm, 0, 1.8*cm, 2.8*cm, 2.5*cm]
            pcol_w[2] = W - sum(pcol_w[:2]) - sum(pcol_w[3:])
            p_hdr = ["ID", "SKU", "Nazwa", "Ilosc", "Koszt PLN", "Data"]
            p_data = [p_hdr]
            p_total = 0.0
            for p in purch:
                p_data.append([str(p["id"]), p["sku"], p["title"],
                                str(p["qty"]), f"{p['total_pln']:,.2f}", p["date"]])
                p_total += p["total_pln"]
            p_data.append(["", "", "SUMA", "", f"{p_total:,.2f}", ""])
            pn = len(p_data)
            pts = TableStyle([
                ("FONTNAME",  (0,0), (-1,0),  font_bold),
                ("FONTSIZE",  (0,0), (-1,-1), 8),
                ("BACKGROUND",(0,0), (-1,0),  colors.HexColor("#1565C0")),
                ("TEXTCOLOR", (0,0), (-1,0),  colors.white),
                ("ALIGN",     (0,0), (-1,0),  "CENTER"),
                ("VALIGN",    (0,0), (-1,-1), "MIDDLE"),
                ("FONTNAME",  (0,1), (-1,-1), font_name),
                ("ROWBACKGROUNDS",(0,1),(-1,pn-2),[colors.white, C_GRY]),
                ("FONTNAME",  (0,pn-1),(-1,pn-1), font_bold),
                ("BACKGROUND",(0,pn-1),(-1,pn-1), C_BLBG),
                ("GRID",      (0,0), (-1,-1), 0.3, colors.HexColor("#DDDDDD")),
                ("LEFTPADDING",  (0,0),(-1,-1), 4),
                ("RIGHTPADDING", (0,0),(-1,-1), 4),
                ("TOPPADDING",   (0,0),(-1,-1), 4),
                ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                ("ALIGN",     (4,1), (4,-1), "RIGHT"),
            ])
            p_tbl = Table(p_data, colWidths=pcol_w, repeatRows=1)
            p_tbl.setStyle(pts)
            story.append(p_tbl)

        # stopka
        story.append(Spacer(1, 12))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"Raport wygenerowany przez {APP_NAME} v{APP_VERSION}  |  "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            sSmall))

        doc.build(story)
        return True


class BackupDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db; self.setWindowTitle("Archiwizacja danych"); self.resize(560,380)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("🗄  Archiwizacja bazy danych", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        dir_row = QHBoxLayout()
        dir_row.addWidget(QLabel("Katalog:"))
        self.dir_edit = QLineEdit(os.path.join(os.getcwd(),"backup")); dir_row.addWidget(self.dir_edit)
        br = btn("📁 Przeglądaj","secondary"); br.clicked.connect(self._browse); dir_row.addWidget(br); v.addLayout(dir_row)
        v.addWidget(QLabel("Istniejące kopie:", styleSheet=f"color:{T()['text2']};"))
        self.lst = QListWidget(); v.addWidget(self.lst)
        btns = QHBoxLayout()
        cb = btn("💾 Utwórz kopię","success"); cb.clicked.connect(self._create)
        rb = btn("🔄 Przywróć","secondary"); rb.clicked.connect(self._restore)
        db2 = btn("🗑 Usuń","danger"); db2.clicked.connect(self._delete)
        cl = btn("Zamknij","secondary"); cl.clicked.connect(self.accept)
        for b2 in [cb,rb,db2]: btns.addWidget(b2)
        btns.addStretch(); btns.addWidget(cl); v.addLayout(btns)
        self._reload()

    def _browse(self):
        d = QFileDialog.getExistingDirectory(self,"Katalog",self.dir_edit.text())
        if d: self.dir_edit.setText(d)

    def _reload(self):
        self.lst.clear(); bdir = self.dir_edit.text()
        if os.path.isdir(bdir):
            for f in sorted([f for f in os.listdir(bdir) if f.endswith(".db")], reverse=True):
                fp = os.path.join(bdir,f); sz = os.path.getsize(fp)//1024
                mt = datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M")
                self.lst.addItem(f"  {f}  ({sz} KB)  –  {mt}")

    def _create(self):
        bdir = self.dir_edit.text(); os.makedirs(bdir,exist_ok=True)
        dest = os.path.join(bdir,f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
        try:
            self.db.backup(dest); self._reload()
            QMessageBox.information(self,"Sukces",f"Kopia zapasowa:\n{dest}")
        except Exception as e: QMessageBox.critical(self,"Błąd",str(e))

    def _restore(self):
        item = self.lst.currentItem()
        if not item: QMessageBox.warning(self,"Brak","Wybierz kopię z listy."); return
        fname = item.text().strip().split("  ")[0]
        src = os.path.join(self.dir_edit.text(),fname)
        if QMessageBox.question(self,"Przywróć",f"Przywrócić z:\n{src}\n\nAktualna baza zostanie nadpisana!",
                                QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            try: shutil.copy2(src,self.db.path); QMessageBox.information(self,"OK","Przywrócono. Uruchom program ponownie.")
            except Exception as e: QMessageBox.critical(self,"Błąd",str(e))

    def _delete(self):
        item = self.lst.currentItem()
        if not item: return
        fname = item.text().strip().split("  ")[0]
        fp = os.path.join(self.dir_edit.text(),fname)
        if QMessageBox.question(self,"Usuń",f"Usunąć: {fname}?",QMessageBox.Yes|QMessageBox.No)==QMessageBox.Yes:
            try: os.remove(fp); self._reload()
            except Exception as e: QMessageBox.critical(self,"Błąd",str(e))


class BusinessInfoDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config; self.setWindowTitle("Dane osobiste"); self.setFixedSize(460,380)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("👤  Dane sprzedawcy", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        info = self.config.get_business_info()
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        def f(ph,key):
            e = QLineEdit(info.get(key,"")); e.setPlaceholderText(ph); return e
        self.name   = f("Imię i Nazwisko","name")
        self.addr   = f("Ulica, nr domu","address")
        self.postal = f("00-000","postal_code")
        self.city   = f("Miejscowość","city")
        self.pesel  = f("11-cyfrowy PESEL","pesel")
        self.nip    = f("NIP (opcjonalnie)","nip")
        for lbl, field in [("Imię i nazwisko:",self.name),("Adres:",self.addr),
                            ("Kod pocztowy:",self.postal),("Miejscowość:",self.city),
                            ("PESEL:",self.pesel),("NIP:",self.nip)]:
            form.addRow(lbl,field)
        v.addLayout(form); v.addStretch()
        btns = QHBoxLayout()
        ok = btn("Zapisz","success"); ok.clicked.connect(self._save)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)

    def _save(self):
        pesel = self.pesel.text().strip()
        if pesel and (len(pesel)!=11 or not pesel.isdigit()):
            QMessageBox.warning(self,"Błąd","PESEL musi mieć 11 cyfr."); return
        self.config.update_business_info({"name":self.name.text().strip(),"address":self.addr.text().strip(),
            "postal_code":self.postal.text().strip(),"city":self.city.text().strip(),
            "pesel":pesel,"nip":self.nip.text().strip()})
        QMessageBox.information(self,"OK","Dane zapisane."); self.accept()


class InvoiceConfigDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config; self.setWindowTitle("Konfiguracja rachunków"); self.setFixedSize(460,300)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("🧾  Konfiguracja rachunków", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        cfg = self.config.get_invoice_config()
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        self.prefix = QLineEdit(self.config.get("invoice_prefix","R"))
        self.footer = QLineEdit(cfg.get("footer_text","Dziękuję za zakup!"))
        self.seller_info = QLineEdit(cfg.get("seller_info",""))
        self.save_pdf = QCheckBox("Zapisuj PDF do katalogu 'rachunki'"); self.save_pdf.setChecked(self.config.should_save_pdf())
        form.addRow("Prefiks numeru:",self.prefix); form.addRow("Stopka rachunku:",self.footer)
        form.addRow("Dodatkowe info:",self.seller_info); form.addRow("",self.save_pdf)
        v.addLayout(form); v.addStretch()
        btns = QHBoxLayout()
        ok = btn("Zapisz","success"); ok.clicked.connect(self._save)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)

    def _save(self):
        self.config.set("invoice_prefix",self.prefix.text().strip() or "R")
        self.config.set("save_pdf",self.save_pdf.isChecked())
        self.config.update_invoice_config({"footer_text":self.footer.text().strip(),"seller_info":self.seller_info.text().strip()})
        QMessageBox.information(self,"OK","Konfiguracja rachunków zapisana."); self.accept()


class LimitsConfigDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config; self.setWindowTitle("Limity US"); self.setFixedSize(500,380)
        v = QVBoxLayout(self); v.setSpacing(10)
        v.addWidget(QLabel("⚖️  Limity działalności nierejestrowanej", styleSheet=f"font-size:15px;font-weight:700;color:{T()['text']};"))
        v.addWidget(Separator(self))
        lim = self.config.get_limits()
        form = QFormLayout(); form.setSpacing(8); form.setLabelAlignment(Qt.AlignRight)
        self.wage = QDoubleSpinBox(); self.wage.setRange(0,20000); self.wage.setDecimals(2)
        self.wage.setSuffix(" PLN"); self.wage.setValue(lim.get("minimal_wage",4666))
        self.qmult = QDoubleSpinBox(); self.qmult.setRange(0,10); self.qmult.setDecimals(2)
        self.qmult.setValue(lim.get("quarterly_multiplier",2.25))
        self.use_q = QCheckBox("Używaj limitów kwartalnych (od 2026)"); self.use_q.setChecked(lim.get("use_quarterly",True))
        form.addRow("Minimalne wynagrodzenie:",self.wage)
        form.addRow("Mnożnik kwartalny:",self.qmult)
        form.addRow("",self.use_q); v.addLayout(form)
        info = QLabel("Od 2026 roku obowiązują limity kwartalne.\n"
                      "Limit mies. = 75% min. wynagrodzenia\n"
                      "Limit kw. = mnożnik × min. wynagrodzenie (domyślnie 2.25)")
        info.setStyleSheet(f"color:{T()['text3']};font-size:11px;"); v.addWidget(info)
        self.calc = QLabel(); self.calc.setStyleSheet(f"color:{T()['accent']};font-weight:700;"); v.addWidget(self.calc)
        self.wage.valueChanged.connect(self._recalc); self.qmult.valueChanged.connect(self._recalc); self._recalc()
        v.addStretch()
        btns = QHBoxLayout()
        ok = btn("Zapisz","success"); ok.clicked.connect(self._save)
        ca = btn("Anuluj","secondary"); ca.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(ca); v.addLayout(btns)

    def _recalc(self):
        w = self.wage.value()
        self.calc.setText(f"Limit miesięczny: {w*0.75:.2f} PLN  |  Limit kwartalny: {w*self.qmult.value():.2f} PLN")

    def _save(self):
        lim = self.config.get_limits()
        lim["minimal_wage"] = self.wage.value(); lim["quarterly_multiplier"] = self.qmult.value()
        lim["use_quarterly"] = self.use_q.isChecked()
        self.config.update_limits(lim); QMessageBox.information(self,"OK","Limity zapisane."); self.accept()


class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("O programie"); self.setFixedSize(480,360)
        v = QVBoxLayout(self); v.setSpacing(12)
        logo = QLabel(APP_NAME); logo.setStyleSheet(f"font-size:20px;font-weight:800;color:{T()['accent']};")
        logo.setAlignment(Qt.AlignCenter); v.addWidget(logo)
        ver = QLabel(f"Wersja {APP_VERSION}  •  {BUILD_DATE}")
        ver.setStyleSheet(f"color:{T()['text3']};font-size:12px;"); ver.setAlignment(Qt.AlignCenter); v.addWidget(ver)
        v.addWidget(Separator(self))
        features = QLabel(
            "<ul>"
            f"<li>Przełączanie motywu dziennego i nocnego</li>"
            f"<li>Limit {PLATFORM_LIMIT} sprzedaży/rok/platforma z wizualizacją</li>"
            f"<li>Zysk netto widoczny wszędzie (FIFO)</li>"
            "<li>Dashboard KPI z wykresem miesięcznym</li>"
            "<li>Raporty CSV / XLSX / PDF</li>"
            "<li>Ewidencja uproszczona US, limity kwartalne od 2026</li>"
            "<li>Archiwizacja z przywracaniem kopii</li>"
            "</ul>")
        features.setTextFormat(Qt.RichText)
        features.setStyleSheet(f"color:{T()['text2']};font-size:12px;"); v.addWidget(features)
        v.addStretch()
        author = QLabel(f"Autor: {APP_AUTHOR}  •  Licencja: {APP_LICENSE}")
        author.setStyleSheet(f"color:{T()['text3']};font-size:11px;"); author.setAlignment(Qt.AlignCenter); v.addWidget(author)
        ok = btn("Zamknij","secondary"); ok.clicked.connect(self.accept); v.addWidget(ok)


# ─────────────────────────────────────────────────────────
#  GŁÓWNE OKNO
# ─────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config   = Config()
        self.db_path  = self.config.get_db_path()
        self.db       = DB(self.db_path)
        self.setWindowTitle(f"{APP_NAME}  v{APP_VERSION}  –  {os.path.basename(self.db_path)}")
        self.resize(1320,820)
        self._build_ui()
        self._build_menu()
        self._build_toolbar()

    def _build_ui(self):
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)
        self.dashboard    = DashboardWidget(self.db,self.config)
        self.products_tab = ProductsWidget(self.db,self.config)
        self.tabs.addTab(self.dashboard,    "📊  Dashboard")
        self.tabs.addTab(self.products_tab, "📦  Magazyn")
        self.tabs.currentChanged.connect(self._tab_changed)
        self.status_bar = QStatusBar(); self.setStatusBar(self.status_bar); self._upd_status()

    def _tab_changed(self, idx):
        if idx==0: self.dashboard.refresh()
        elif idx==1: self.products_tab.refresh()

    def _upd_status(self):
        self.status_bar.showMessage(
            f"Baza: {self.db_path}  |  {APP_NAME} v{APP_VERSION}  |  "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")

    def _act(self, menu, text, slot, shortcut=None):
        a = QAction(text,self)
        if shortcut: a.setShortcut(shortcut)
        a.triggered.connect(slot); menu.addAction(a); return a

    def _build_menu(self):
        mb = self.menuBar()
        mf = mb.addMenu("&Plik")
        self._act(mf,"📂 Otwórz bazę danych…",self._open_db,"Ctrl+O")
        self._act(mf,"🆕 Nowa baza danych…",self._new_db)
        mf.addSeparator()
        self._act(mf,"🗄 Archiwizacja…",self._backup,"Ctrl+B")
        mf.addSeparator()
        self._act(mf,"📤 Eksport CSV…",self._quick_export)
        mf.addSeparator()
        self._act(mf,"❌ Zakończ",self.close,"Ctrl+Q")

        mm = mb.addMenu("&Magazyn")
        self._act(mm,"➕ Dodaj produkt",self._add_product,"Ctrl+N")
        self._act(mm,"📦 Rejestruj zakup…",self._add_purchase,"Ctrl+Z")
        mm.addSeparator()
        self._act(mm,"📋 Inwentaryzacja…",self._inventory)
        self._act(mm,"📜 Historia zakupów…",self._show_purchases,"Ctrl+Shift+Z")

        ms = mb.addMenu("&Sprzedaż")
        self._act(ms,"💰 Nowa sprzedaż…",self._add_sale,"Ctrl+S")
        ms.addSeparator()
        self._act(ms,"📜 Historia sprzedaży…",self._show_sales,"Ctrl+Shift+S")
        self._act(ms,"🧾 Historia rachunków…",self._show_invoices,"Ctrl+Shift+R")

        mr = mb.addMenu("&Raporty")
        self._act(mr,"📅 Raport miesięczny…", lambda: self._report("monthly"),  "Ctrl+1")
        self._act(mr,"📆 Raport kwartalny…",  lambda: self._report("quarterly"), "Ctrl+2")
        self._act(mr,"📈 Raport roczny…",     lambda: self._report("yearly"),    "Ctrl+3")
        self._act(mr,"🗓 Raport za okres…",   lambda: self._report("custom"),    "Ctrl+4")

        mc = mb.addMenu("&Konfiguracja")
        self._act(mc,"👤 Dane osobiste…",         self._biz_info)
        self._act(mc,"🧾 Ustawienia rachunków…",  self._inv_cfg)
        self._act(mc,"⚖️ Limity US…",             self._lim_cfg)

        mh = mb.addMenu("&Pomoc")
        self._act(mh,"⟳ Odśwież",self._refresh,"F5")
        mh.addSeparator()
        self._act(mh,"ℹ️ O programie…",self._about)

    def _build_toolbar(self):
        tb = self.addToolBar("Główne"); tb.setMovable(False); tb.setIconSize(QSize(20,20))
        def ta(text,slot,tip=""): a=QAction(text,self); a.setToolTip(tip); a.triggered.connect(slot); tb.addAction(a)
        ta("➕ Produkt",    self._add_product,   "Nowy produkt (Ctrl+N)")
        tb.addSeparator()
        ta("📦 Zakup",      self._add_purchase,  "Rejestruj zakup (Ctrl+Z)")
        ta("💰 Sprzedaż",   self._add_sale,      "Nowa sprzedaż (Ctrl+S)")
        tb.addSeparator()
        ta("📅 Mies.",  lambda: self._report("monthly"),  "Raport miesięczny")
        ta("📆 Kw.",    lambda: self._report("quarterly"), "Raport kwartalny")
        tb.addSeparator()
        ta("🗄 Backup",     self._backup,         "Archiwizacja (Ctrl+B)")
        ta("⟳ Odśwież",    self._refresh,        "Odśwież (F5)")

    # ── akcje ──
    def _refresh(self): self.dashboard.refresh(); self.products_tab.refresh(); self._upd_status()
    def _add_product(self):
        if ProductDialog(self.db,parent=self).exec(): self.products_tab.refresh(); self.dashboard.refresh()
    def _add_purchase(self):
        d = PurchaseDialog(self.db,parent=self)
        if d.exec():
            self.db.add_purchase_order(d.result_cost,d.result_date,d.result_items)
            self.products_tab.refresh(); self.dashboard.refresh()
    def _add_sale(self):
        if SaleDialog(self.db,self.config,parent=self).exec():
            self.products_tab.refresh(); self.dashboard.refresh()
    def _show_purchases(self):
        rows = self.db.list_purchases()
        data = [(r["id"],r["sku"],r["title"],r["qty"],f"{r['total_pln']:.2f}",r["date"]) for r in rows]
        HistoryDialog("Historia zakupów",["ID","SKU","Nazwa","Ilość","Koszt PLN","Data"],data,self.db.delete_purchase,self).exec()
        self._refresh()
    def _show_sales(self):
        rows = self.db.list_sales()
        data = [(r["id"],r["platform"],f"{r['total_pln']:.2f}",f"{r['total_eur']:.2f}",
                 f"{r['purchase_cost']:.2f}",f"{r['profit']:.2f}",r["date"],r["items"]) for r in rows]
        HistoryDialog("Historia sprzedaży",["ID","Platforma","PLN","EUR","Koszt","Zysk netto","Data","Pozycje"],
                      data,self.db.delete_sale,self).exec()
        self._refresh()
    def _show_invoices(self): InvoicesDialog(self.db,self.config,self).exec()
    def _inventory(self): InventoryDialog(self.db,self).exec(); self._refresh()
    def _report(self, rt): ReportDialog(self.db,self.config,self,rt).exec()
    def _biz_info(self): BusinessInfoDialog(self.config,self).exec()
    def _inv_cfg(self): InvoiceConfigDialog(self.config,self).exec()
    def _lim_cfg(self): LimitsConfigDialog(self.config,self).exec()
    def _about(self): AboutDialog(self).exec()
    def _backup(self): BackupDialog(self.db,self).exec()

    def _open_db(self):
        path,_ = QFileDialog.getOpenFileName(self,"Otwórz bazę","","SQLite Database (*.db)")
        if path: self._switch(path)
    def _new_db(self):
        path,_ = QFileDialog.getSaveFileName(self,"Nowa baza","","SQLite Database (*.db)")
        if path: self._switch(path)
    def _switch(self, path):
        try:
            if hasattr(self.db,"conn"): self.db.conn.close()
            self.config.set_db_path(path); self.db_path = path; self.db = DB(path)
            self.dashboard.db = self.db; self.products_tab.db = self.db
            self.setWindowTitle(f"{APP_NAME}  v{APP_VERSION}  –  {os.path.basename(path)}")
            self._refresh(); QMessageBox.information(self,"OK",f"Załadowano bazę:\n{path}")
        except Exception as e: QMessageBox.critical(self,"Błąd",str(e))
    def _quick_export(self):
        now = datetime.now(); df=f"{now.year}-01-01"; dt=now.strftime("%Y-%m-%d")
        path,_ = QFileDialog.getSaveFileName(self,"Eksport CSV",f"eksport_{now.year}.csv","CSV Files (*.csv)")
        if path:
            try: self.db.export_csv(path,df,dt); QMessageBox.information(self,"OK",f"Zapisano:\n{path}")
            except Exception as e: QMessageBox.critical(self,"Błąd",str(e))
    def closeEvent(self, event):
        try:
            if hasattr(self.db,"conn"): self.db.conn.close()
        except: pass
        event.accept()


# ─────────────────────────────────────────────────────────
#  START
# ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    apply_theme(app, THEME_DAY)   # domyślnie motyw dzienny
    if not os.path.exists("data.db"):
        DB("data.db").conn.close()
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
