import sys
import os
from datetime import datetime, timedelta
from PySide6.QtWidgets import *
from PySide6.QtCore import QDate, Qt, QTimer, QSize
from PySide6.QtGui import QFont, QAction, QKeySequence, QIcon

from db import DB
from currency import get_eur_rate
from config import Config

PLATFORMS = ["Vinted", "OLX", "Allegro Lokalnie", "FB Marketplace", "Inne"]

# ================== WERSJA ==================
try:
    from version import __version__, display_version
    APP_VERSION = __version__
except ImportError:
    APP_VERSION = "2.1.0"
    def display_version():
        print(f"System Magazynowo-SprzedaÅ¼owy v{APP_VERSION}")

try:
    import config
    if hasattr(config, "PLATFORMS"):
        PLATFORMS = config.PLATFORMS
except ImportError:
    pass

# ================== STYL ==================
RED_WHITE_QSS = """
QWidget {
    background-color: white;
    font-family: Segoe UI, Arial;
    font-size: 12px;
    color: black
}
QPushButton {
    background-color: #c62828;
    color: white;
    padding: 6px 10px;
    border-radius: 6px;
}
QPushButton:hover {
    background-color: #b71c1c;
}
QHeaderView::section {
    background-color: #f2f2f2;
    padding: 4px;
    color: black;
    border: 1px solid #ddd;
}
QHeaderView::section:checked {
    background-color: #e0e0e0;
}
QTableWidget {
    border: 1px solid #ddd;
    color: black;
}
QTableWidget::item {
    padding: 4px;
}
QCheckBox {
    spacing: 5px;
}
QGroupBox {
    border: 1px solid #ddd;
    border-radius: 5px;
    margin-top: 10px;
    padding-top: 10px;
    font-weight: bold;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 5px 0 5px;
}
/* Styl dla menu ribbon - POPRAWIONE */
QMenuBar {
    background-color: #2c3e50;
    color: white;
    font-weight: bold;
    spacing: 5px;
}
QMenuBar::item {
    background-color: transparent;
    color: white;
    padding: 8px 15px;
    margin: 0 1px;
    border-radius: 3px;
}
QMenuBar::item:selected {
    background-color: #34495e;
}
QMenuBar::item:pressed {
    background-color: #2c3e50;
}
/* Menu rozwijane - POPRAWIONE DO POZIOMEGO */
QMenu {
    background-color: white;
    border: 1px solid #ddd;
    border-radius: 3px;
    margin: 0;
    padding: 5px;
}
QMenu::item {
    padding: 8px 25px 8px 15px;
    margin: 2px 0;
    border-radius: 3px;
}
QMenu::item:selected {
    background-color: #c62828;
    color: white;
}
QMenu::item:disabled {
    color: #999;
}
QMenu::separator {
    height: 1px;
    background: #ddd;
    margin: 5px 10px;
}
/* Dla menu podrÄ™cznych - POPRAWIONE */
QMenu::right-arrow {
    image: none;
    width: 0;
}
/* Toolbar */
QToolBar {
    background-color: #f8f9fa;
    border-bottom: 1px solid #ddd;
    spacing: 5px;
    padding: 5px;
}
QToolButton {
    padding: 6px 10px;
    border-radius: 4px;
    margin: 0 2px;
    text-align: center;
}
QToolButton:hover {
    background-color: #e9ecef;
}
QToolButton:pressed {
    background-color: #dee2e6;
}
QToolButton[popupMode="1"] {
    padding-right: 20px;
}
/* Styl dla akcji */
QAction {
    spacing: 8px;
}
"""

# ================== POMOCNICZE ==================
def product_combo(db):
    combo = QComboBox()
    for p in db.list_products():
        combo.addItem(
            f"{p['id']} | {p['sku']} | {p['title']}",
            p["id"]
        )
    return combo

# ================== DIALOG DODAWANIA PRODUKTU ==================
class AddProductDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Dodaj nowy produkt")
        self.resize(400, 200)

        v = QVBoxLayout(self)

        # Formularz
        form = QFormLayout()

        self.sku_input = QLineEdit()
        self.sku_input.setPlaceholderText("np. PROD001")
        form.addRow("SKU:", self.sku_input)

        self.title_input = QLineEdit()
        self.title_input.setPlaceholderText("np. Koszulka baweÅ‚niana")
        form.addRow("Nazwa:", self.title_input)

        self.initial_stock = QSpinBox()
        self.initial_stock.setRange(0, 100000)
        self.initial_stock.setValue(0)
        form.addRow("Stan poczÄ…tkowy:", self.initial_stock)

        v.addLayout(form)

        # Przyciski
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("Zapisz")
        btn_save.clicked.connect(self.save_product)
        btn_save.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_cancel)
        v.addLayout(button_layout)

    def save_product(self):
        sku = self.sku_input.text().strip()
        title = self.title_input.text().strip()
        
        if not sku:
            QMessageBox.warning(self, "BÅ‚Ä…d", "SKU nie moÅ¼e byÄ‡ puste.")
            return
            
        if not title:
            QMessageBox.warning(self, "BÅ‚Ä…d", "Nazwa nie moÅ¼e byÄ‡ pusta.")
            return
        
        try:
            # SprawdÅº czy SKU juÅ¼ istnieje
            existing = self.db.check_sku_exists(sku)
            if existing:
                QMessageBox.warning(self, "BÅ‚Ä…d", f"SKU '{sku}' juÅ¼ istnieje w bazie.")
                return
            
            # Dodaj produkt
            self.db.add_product(sku, title)
            
            # JeÅ›li stan poczÄ…tkowy > 0, dodaj zakup
            initial_qty = self.initial_stock.value()
            if initial_qty > 0:
                pid = self.db.get_product_id_by_sku(sku)
                if pid:
                    self.db.add_purchase_order(
                        0.0,
                        datetime.now().strftime("%Y-%m-%d"),
                        [(pid, initial_qty)]
                    )
            
            QMessageBox.information(self, "Sukces", "Produkt zostaÅ‚ dodany.")
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", f"Nie udaÅ‚o siÄ™ dodaÄ‡ produktu:\n{str(e)}")

# ================== DIALOG KONFIGURACJI DANYCH OSOBOWYCH ==================
class BusinessInfoDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Dane osobiste dla ewidencji uproszczonej")
        self.resize(500, 400)
        
        v = QVBoxLayout(self)
        
        info = self.config.get_business_info()
        
        form = QFormLayout()
        
        self.name_input = QLineEdit()
        self.name_input.setText(info.get("name", ""))
        self.name_input.setPlaceholderText("ImiÄ™ i nazwisko")
        form.addRow("ImiÄ™ i nazwisko:", self.name_input)
        
        self.address_input = QLineEdit()
        self.address_input.setText(info.get("address", ""))
        self.address_input.setPlaceholderText("Ulica, nr domu/mieszkania")
        form.addRow("Adres zamieszkania:", self.address_input)
        
        self.postal_input = QLineEdit()
        self.postal_input.setText(info.get("postal_code", ""))
        self.postal_input.setPlaceholderText("00-000")
        form.addRow("Kod pocztowy:", self.postal_input)
        
        self.city_input = QLineEdit()
        self.city_input.setText(info.get("city", ""))
        self.city_input.setPlaceholderText("MiejscowoÅ›Ä‡")
        form.addRow("MiejscowoÅ›Ä‡:", self.city_input)
        
        self.pesel_input = QLineEdit()
        self.pesel_input.setText(info.get("pesel", ""))
        self.pesel_input.setPlaceholderText("11-cyfrowy numer PESEL")
        form.addRow("PESEL:", self.pesel_input)
        
        self.nip_input = QLineEdit()
        self.nip_input.setText(info.get("nip", ""))
        self.nip_input.setPlaceholderText("NIP (opcjonalnie)")
        form.addRow("NIP:", self.nip_input)
        
        self.regon_input = QLineEdit()
        self.regon_input.setText(info.get("regon", ""))
        self.regon_input.setPlaceholderText("REGON (opcjonalnie)")
        form.addRow("REGON:", self.regon_input)
        
        v.addLayout(form)
        
        # Przyciski
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("Zapisz")
        btn_save.clicked.connect(self.save_info)
        btn_save.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_cancel)
        
        v.addLayout(button_layout)
    
    def save_info(self):
        info = {
            "name": self.name_input.text().strip(),
            "address": self.address_input.text().strip(),
            "postal_code": self.postal_input.text().strip(),
            "city": self.city_input.text().strip(),
            "pesel": self.pesel_input.text().strip(),
            "nip": self.nip_input.text().strip(),
            "regon": self.regon_input.text().strip()
        }
        
        # Walidacja
        if not info["name"]:
            QMessageBox.warning(self, "BÅ‚Ä…d", "ImiÄ™ i nazwisko jest wymagane.")
            return
            
        if not info["address"]:
            QMessageBox.warning(self, "BÅ‚Ä…d", "Adres jest wymagany.")
            return
            
        if not info["postal_code"]:
            QMessageBox.warning(self, "BÅ‚Ä…d", "Kod pocztowy jest wymagany.")
            return
            
        if not info["city"]:
            QMessageBox.warning(self, "BÅ‚Ä…d", "MiejscowoÅ›Ä‡ jest wymagana.")
            return
            
        if not info["pesel"] or len(info["pesel"]) != 11 or not info["pesel"].isdigit():
            QMessageBox.warning(self, "BÅ‚Ä…d", "PESEL musi skÅ‚adaÄ‡ siÄ™ z 11 cyfr.")
            return
        
        self.config.update_business_info(info)
        QMessageBox.information(self, "Sukces", "Dane zostaÅ‚y zapisane.")
        self.accept()

# ================== DIALOG KONFIGURACJI LIMITÃ“W ==================
class LimitsConfigDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Konfiguracja limitÃ³w US")
        self.resize(600, 500)
        
        v = QVBoxLayout(self)
        
        # Informacja o limitach
        info_group = QGroupBox("Informacje o limitach")
        info_layout = QVBoxLayout()
        
        info_text = QLabel(
            "<b>Od 2026 roku w Polsce obowiÄ…zujÄ… limity kwartalne dla dziaÅ‚alnoÅ›ci nierejestrowanej:</b><br><br>"
            "â€¢ <b>Limit miesiÄ™czny:</b> 75% minimalnego wynagrodzenia<br>"
            "â€¢ <b>Limit kwartalny:</b> 225% minimalnego wynagrodzenia<br><br>"
            "System automatycznie oblicza limity na podstawie wprowadzonej kwoty minimalnego wynagrodzenia."
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)
        info_group.setLayout(info_layout)
        v.addWidget(info_group)
        
        # Grupa: OgÃ³lne ustawienia
        gb_general = QGroupBox("OgÃ³lne ustawienia limitÃ³w")
        general_layout = QFormLayout()
        
        self.minimal_wage_input = QDoubleSpinBox()
        self.minimal_wage_input.setRange(0, 10000)
        self.minimal_wage_input.setDecimals(2)
        self.minimal_wage_input.setSuffix(" PLN")
        self.minimal_wage_input.setSingleStep(100)
        
        self.quarterly_multiplier_input = QDoubleSpinBox()
        self.quarterly_multiplier_input.setRange(0, 10)
        self.quarterly_multiplier_input.setDecimals(2)
        self.quarterly_multiplier_input.setSuffix("%")
        self.quarterly_multiplier_input.setSingleStep(0.25)
        
        self.use_quarterly_cb = QCheckBox("UÅ¼ywaj limitÃ³w kwartalnych (od 2026 roku)")
        self.use_quarterly_cb.setChecked(True)
        
        general_layout.addRow("Minimalne wynagrodzenie:", self.minimal_wage_input)
        general_layout.addRow("MnoÅ¼nik limitu kwartalnego:", self.quarterly_multiplier_input)
        general_layout.addRow("", self.use_quarterly_cb)
        
        gb_general.setLayout(general_layout)
        v.addWidget(gb_general)
        
        # Grupa: Limity dla poszczegÃ³lnych lat
        gb_years = QGroupBox("Limity dla poszczegÃ³lnych lat")
        years_layout = QVBoxLayout()
        
        self.years_table = QTableWidget()
        self.years_table.setColumnCount(5)
        self.years_table.setHorizontalHeaderLabels(["Rok", "Minimalne wynagrodzenie", "Limit miesiÄ™czny", "Limit kwartalny", "Edytuj"])
        self.years_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        years_layout.addWidget(self.years_table)
        gb_years.setLayout(years_layout)
        v.addWidget(gb_years)
        
        # Obliczone wartoÅ›ci
        gb_calculated = QGroupBox("Obliczone wartoÅ›ci (aktualne)")
        calculated_layout = QVBoxLayout()
        
        self.calculated_label = QLabel()
        self.calculated_label.setStyleSheet("font-weight: bold;")
        calculated_layout.addWidget(self.calculated_label)
        
        gb_calculated.setLayout(calculated_layout)
        v.addWidget(gb_calculated)
        
        # Przyciski
        button_layout = QHBoxLayout()
        
        btn_save = QPushButton("Zapisz konfiguracjÄ™")
        btn_save.clicked.connect(self.save_config)
        btn_save.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_calculate = QPushButton("Przelicz")
        btn_calculate.clicked.connect(self.calculate_values)
        
        btn_add_year = QPushButton("Dodaj rok")
        btn_add_year.clicked.connect(self.add_year)
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_calculate)
        button_layout.addWidget(btn_add_year)
        button_layout.addWidget(btn_cancel)
        
        v.addLayout(button_layout)
        
        self.load_config()
        self.calculate_values()
    
    def load_config(self):
        """Wczytuje aktualnÄ… konfiguracjÄ™"""
        limits_config = self.config.get_limits_config()
        
        # OgÃ³lne ustawienia
        self.minimal_wage_input.setValue(limits_config.get("minimal_wage", 4242.00))
        self.quarterly_multiplier_input.setValue(limits_config.get("quarterly_limit_multiplier", 2.25))
        self.use_quarterly_cb.setChecked(limits_config.get("use_quarterly_limits", True))
        
        # Tabela z latami
        self.load_years_table()
    
    def load_years_table(self):
        """Wczytuje dane lat do tabeli"""
        limits_config = self.config.get_limits_config()
        year_limits = limits_config.get("year_limits", {})
        
        years = sorted(year_limits.keys(), key=lambda x: int(x))
        self.years_table.setRowCount(len(years))
        
        for i, year in enumerate(years):
            data = year_limits[year]
            
            # Rok
            self.years_table.setItem(i, 0, QTableWidgetItem(year))
            
            # Minimalne wynagrodzenie
            wage_item = QTableWidgetItem(f"{data.get('minimal_wage', 0):.2f} PLN")
            wage_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.years_table.setItem(i, 1, wage_item)
            
            # Limit miesiÄ™czny
            monthly_limit = data.get('minimal_wage', 0) * 0.75
            monthly_item = QTableWidgetItem(f"{monthly_limit:.2f} PLN")
            monthly_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.years_table.setItem(i, 2, monthly_item)
            
            # Limit kwartalny
            quarterly_limit = data.get('minimal_wage', 0) * data.get('quarterly_limit_multiplier', 2.25)
            quarterly_item = QTableWidgetItem(f"{quarterly_limit:.2f} PLN")
            quarterly_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.years_table.setItem(i, 3, quarterly_item)
            
            # Przycisk edycji
            edit_btn = QPushButton("Edytuj")
            edit_btn.clicked.connect(lambda checked, y=year: self.edit_year(y))
            self.years_table.setCellWidget(i, 4, edit_btn)
    
    def calculate_values(self):
        """Przelicza wartoÅ›ci na podstawie wprowadzonych danych"""
        minimal_wage = self.minimal_wage_input.value()
        quarterly_multiplier = self.quarterly_multiplier_input.value()
        
        monthly_limit = minimal_wage * 0.75
        quarterly_limit = minimal_wage * quarterly_multiplier
        
        self.calculated_label.setText(
            f"<b>Obliczone wartoÅ›ci (aktualne ustawienia):</b><br>"
            f"Minimalne wynagrodzenie: {minimal_wage:.2f} PLN<br>"
            f"Limit miesiÄ™czny (75%): {monthly_limit:.2f} PLN<br>"
            f"Limit kwartalny ({quarterly_multiplier*100:.0f}%): {quarterly_limit:.2f} PLN"
        )
    
    def add_year(self):
        """Dodaje nowy rok do konfiguracji"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Dodaj nowy rok")
        dialog.resize(300, 200)
        
        layout = QVBoxLayout(dialog)
        
        form = QFormLayout()
        
        year_spin = QSpinBox()
        year_spin.setRange(2000, 2100)
        year_spin.setValue(datetime.now().year)
        form.addRow("Rok:", year_spin)
        
        wage_input = QDoubleSpinBox()
        wage_input.setRange(0, 10000)
        wage_input.setDecimals(2)
        wage_input.setValue(self.minimal_wage_input.value())
        wage_input.setSuffix(" PLN")
        form.addRow("Minimalne wynagrodzenie:", wage_input)
        
        layout.addLayout(form)
        
        button_layout = QHBoxLayout()
        btn_ok = QPushButton("Dodaj")
        btn_ok.clicked.connect(dialog.accept)
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(dialog.reject)
        
        button_layout.addWidget(btn_ok)
        button_layout.addWidget(btn_cancel)
        layout.addLayout(button_layout)
        
        if dialog.exec():
            year = year_spin.value()
            minimal_wage = wage_input.value()
            
            # Dodaj rok do konfiguracji
            config = {
                "minimal_wage": minimal_wage,
                "monthly_limit": minimal_wage * 0.75,
                "quarterly_limit": minimal_wage * self.quarterly_multiplier_input.value()
            }
            
            self.config.update_year_limit_config(year, config)
            self.load_years_table()
    
    def edit_year(self, year):
        """Edytuje konfiguracjÄ™ dla danego roku"""
        limits_config = self.config.get_limits_config()
        year_limits = limits_config.get("year_limits", {})
        
        if str(year) not in year_limits:
            QMessageBox.warning(self, "BÅ‚Ä…d", f"Nie znaleziono konfiguracji dla roku {year}")
            return
        
        config = year_limits[str(year)]
        
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Edytuj rok {year}")
        dialog.resize(300, 200)
        
        layout = QVBoxLayout(dialog)
        
        form = QFormLayout()
        
        year_label = QLabel(str(year))
        form.addRow("Rok:", year_label)
        
        wage_input = QDoubleSpinBox()
        wage_input.setRange(0, 10000)
        wage_input.setDecimals(2)
        wage_input.setValue(config.get("minimal_wage", self.minimal_wage_input.value()))
        wage_input.setSuffix(" PLN")
        form.addRow("Minimalne wynagrodzenie:", wage_input)
        
        layout.addLayout(form)
        
        button_layout = QHBoxLayout()
        btn_save = QPushButton("Zapisz")
        btn_save.clicked.connect(dialog.accept)
        btn_delete = QPushButton("UsuÅ„")
        btn_delete.clicked.connect(lambda: self.delete_year(year, dialog))
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(dialog.reject)
        
        button_layout.addWidget(btn_save)
        button_layout.addWidget(btn_delete)
        button_layout.addWidget(btn_cancel)
        layout.addLayout(button_layout)
        
        if dialog.exec():
            minimal_wage = wage_input.value()
            
            # Zaktualizuj konfiguracjÄ™
            config = {
                "minimal_wage": minimal_wage,
                "monthly_limit": minimal_wage * 0.75,
                "quarterly_limit": minimal_wage * self.quarterly_multiplier_input.value()
            }
            
            self.config.update_year_limit_config(year, config)
            self.load_years_table()
    
    def delete_year(self, year, dialog):
        """Usuwa konfiguracjÄ™ dla danego roku"""
        if QMessageBox.question(self, "PotwierdÅº usuniÄ™cie", 
                               f"Czy na pewno usunÄ…Ä‡ konfiguracjÄ™ dla roku {year}?") == QMessageBox.Yes:
            limits_config = self.config.get_limits_config()
            if "year_limits" in limits_config and str(year) in limits_config["year_limits"]:
                del limits_config["year_limits"][str(year)]
                self.config.update_limits_config(limits_config)
                self.load_years_table()
                dialog.reject()
    
    def save_config(self):
        """Zapisuje konfiguracjÄ™ limitÃ³w"""
        limits_info = {
            "minimal_wage": self.minimal_wage_input.value(),
            "quarterly_limit_multiplier": self.quarterly_multiplier_input.value(),
            "use_quarterly_limits": self.use_quarterly_cb.isChecked()
        }
        
        self.config.update_limits_config(limits_info)
        QMessageBox.information(self, "Sukces", "Konfiguracja limitÃ³w zostaÅ‚a zapisana.")
        self.accept()

# ================== DIALOG RAPORTU (z drukowaniem) ==================
class ReportDialog(QDialog):
    def __init__(self, db, config, parent=None, report_type="monthly"):
        super().__init__(parent)
        self.db = db
        self.config = config
        self.report_type = report_type  # "monthly", "yearly", "quarterly", "custom"
        
        if report_type == "monthly":
            self.setWindowTitle("Generuj raport miesiÄ™czny")
        elif report_type == "yearly":
            self.setWindowTitle("Generuj raport roczny")
        elif report_type == "quarterly":
            self.setWindowTitle("Generuj raport kwartalny")
        else:
            self.setWindowTitle("Generuj raport okresowy")
            
        self.resize(600, 650)

        v = QVBoxLayout(self)

        # Typ raportu (tylko informacyjnie)
        gb_type = QGroupBox("Typ raportu")
        type_layout = QVBoxLayout()
        
        if report_type == "monthly":
            self.type_label = QLabel("Raport miesiÄ™czny")
        elif report_type == "yearly":
            self.type_label = QLabel("Raport roczny")
        elif report_type == "quarterly":
            self.type_label = QLabel("Raport kwartalny")
        else:
            self.type_label = QLabel("Raport okresowy")
            
        self.type_label.setStyleSheet("font-weight: bold; color: #2E7D32;")
        type_layout.addWidget(self.type_label)
        gb_type.setLayout(type_layout)
        v.addWidget(gb_type)

        # Kontenery dla rÃ³Å¼nych typÃ³w raportÃ³w
        if report_type == "monthly":
            self.monthly_widget = QWidget()
            monthly_layout = QHBoxLayout(self.monthly_widget)
            self.month_combo = QComboBox()
            self.month_combo.addItems([
                "StyczeÅ„", "Luty", "Marzec", "KwiecieÅ„", "Maj", "Czerwiec",
                "Lipiec", "SierpieÅ„", "WrzesieÅ„", "PaÅºdziernik", "Listopad", "GrudzieÅ„"
            ])
            self.month_combo.setCurrentIndex(datetime.now().month - 1)
            
            self.year_spin = QSpinBox()
            self.year_spin.setRange(2000, 2100)
            self.year_spin.setValue(datetime.now().year)
            
            monthly_layout.addWidget(QLabel("MiesiÄ…c:"))
            monthly_layout.addWidget(self.month_combo)
            monthly_layout.addWidget(QLabel("Rok:"))
            monthly_layout.addWidget(self.year_spin)
            monthly_layout.addStretch()
            
            v.addWidget(self.monthly_widget)
            
        elif report_type == "yearly":
            self.yearly_widget = QWidget()
            yearly_layout = QHBoxLayout(self.yearly_widget)
            self.year_only_spin = QSpinBox()
            self.year_only_spin.setRange(2000, 2100)
            self.year_only_spin.setValue(datetime.now().year)
            
            yearly_layout.addWidget(QLabel("Rok:"))
            yearly_layout.addWidget(self.year_only_spin)
            yearly_layout.addStretch()
            
            v.addWidget(self.yearly_widget)
            
        elif report_type == "quarterly":
            self.quarterly_widget = QWidget()
            quarterly_layout = QHBoxLayout(self.quarterly_widget)
            
            self.quarter_combo = QComboBox()
            self.quarter_combo.addItems([
                "I kwartaÅ‚ (styczeÅ„-marzec)",
                "II kwartaÅ‚ (kwiecieÅ„-czerwiec)",
                "III kwartaÅ‚ (lipiec-wrzesieÅ„)",
                "IV kwartaÅ‚ (paÅºdziernik-grudzieÅ„)"
            ])
            
            # Ustal bieÅ¼Ä…cy kwartaÅ‚
            current_month = datetime.now().month
            current_quarter = (current_month - 1) // 3
            self.quarter_combo.setCurrentIndex(current_quarter)
            
            self.quarter_year_spin = QSpinBox()
            self.quarter_year_spin.setRange(2000, 2100)
            self.quarter_year_spin.setValue(datetime.now().year)
            
            quarterly_layout.addWidget(QLabel("KwartaÅ‚:"))
            quarterly_layout.addWidget(self.quarter_combo)
            quarterly_layout.addWidget(QLabel("Rok:"))
            quarterly_layout.addWidget(self.quarter_year_spin)
            quarterly_layout.addStretch()
            
            v.addWidget(self.quarterly_widget)
            
        else:  # custom
            self.custom_widget = QWidget()
            custom_layout = QHBoxLayout(self.custom_widget)
            self.date_from = QDateEdit(QDate.currentDate().addMonths(-1))
            self.date_from.setCalendarPopup(True)
            self.date_to = QDateEdit(QDate.currentDate())
            self.date_to.setCalendarPopup(True)
            
            custom_layout.addWidget(QLabel("Od:"))
            custom_layout.addWidget(self.date_from)
            custom_layout.addWidget(QLabel("Do:"))
            custom_layout.addWidget(self.date_to)
            custom_layout.addStretch()
            
            v.addWidget(self.custom_widget)

        # Format eksportu
        gb_format = QGroupBox("Format eksportu")
        format_layout = QHBoxLayout()
        
        self.rb_csv = QRadioButton("CSV")
        self.rb_csv.setChecked(True)
        self.rb_excel = QRadioButton("Excel (XLSX)")
        self.rb_pdf = QRadioButton("PDF")
        
        format_layout.addWidget(self.rb_csv)
        format_layout.addWidget(self.rb_excel)
        format_layout.addWidget(self.rb_pdf)
        format_layout.addStretch()
        gb_format.setLayout(format_layout)
        v.addWidget(gb_format)

        # Opcje raportu
        gb_options = QGroupBox("Opcje raportu")
        options_layout = QVBoxLayout()
        
        self.cb_consolidate_sales = QCheckBox("ÅÄ…cz pozycje w ramach jednej sprzedaÅ¼y (zalecane)")
        self.cb_consolidate_sales.setChecked(True)
        self.cb_consolidate_sales.setToolTip("ÅÄ…czy wszystkie produkty z jednej sprzedaÅ¼y w jednÄ… pozycjÄ™")
        
        self.cb_show_products = QCheckBox("Pokazuj listÄ™ produktÃ³w w sprzedaÅ¼y")
        self.cb_show_products.setChecked(True)
        
        self.cb_purchases = QCheckBox("UwzglÄ™dnij zakupy")
        self.cb_purchases.setChecked(False)
        self.cb_sales = QCheckBox("UwzglÄ™dnij sprzedaÅ¼")
        self.cb_sales.setChecked(True)
        self.cb_summary = QCheckBox("Podsumowanie finansowe")
        self.cb_summary.setChecked(True)
        
        self.cb_simple_register = QCheckBox("Uproszczony rejestr sprzedaÅ¼y z danymi osobowymi")
        self.cb_simple_register.setChecked(False)
        self.cb_simple_register.setToolTip("Generuje raport z limitami US i danymi osobowymi")
        
        options_layout.addWidget(self.cb_consolidate_sales)
        options_layout.addWidget(self.cb_show_products)
        options_layout.addWidget(self.cb_purchases)
        options_layout.addWidget(self.cb_sales)
        options_layout.addWidget(self.cb_summary)
        options_layout.addWidget(self.cb_simple_register)
        
        gb_options.setLayout(options_layout)
        v.addWidget(gb_options)

        # Przyciski
        button_layout = QHBoxLayout()
        
        btn_generate = QPushButton("Generuj raport")
        btn_generate.clicked.connect(self.generate_report)
        btn_generate.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_print = QPushButton("ðŸ–¨ï¸ Drukuj")
        btn_print.clicked.connect(self.print_report)
        btn_print.setToolTip("Drukuj raport bezpoÅ›rednio na drukarce")
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(btn_generate)
        button_layout.addWidget(btn_print)
        button_layout.addWidget(btn_cancel)
        v.addLayout(button_layout)

    def get_date_range(self):
        try:
            if self.report_type == "monthly":
                month = self.month_combo.currentIndex() + 1
                year = self.year_spin.value()
                date_from = f"{year}-{month:02d}-01"
                
                if month == 12:
                    date_to = f"{year}-12-31"
                else:
                    next_month = datetime(year, month + 1, 1)
                    last_day = next_month - timedelta(days=1)
                    date_to = last_day.strftime("%Y-%m-%d")
                    
            elif self.report_type == "yearly":
                year = self.year_only_spin.value()
                date_from = f"{year}-01-01"
                date_to = f"{year}-12-31"
                
            elif self.report_type == "quarterly":
                quarter = self.quarter_combo.currentIndex() + 1
                year = self.quarter_year_spin.value()
                
                # OkreÅ›l miesiÄ…ce dla kwartaÅ‚u
                if quarter == 1:
                    date_from = f"{year}-01-01"
                    date_to = f"{year}-03-31"
                elif quarter == 2:
                    date_from = f"{year}-04-01"
                    date_to = f"{year}-06-30"
                elif quarter == 3:
                    date_from = f"{year}-07-01"
                    date_to = f"{year}-09-30"
                else:  # quarter == 4
                    date_from = f"{year}-10-01"
                    date_to = f"{year}-12-31"
                
            else:  # custom
                date_from = self.date_from.date().toString("yyyy-MM-dd")
                date_to = self.date_to.date().toString("yyyy-MM-dd")
                
            return date_from, date_to
            
        except Exception as e:
            today = datetime.now()
            date_from = today.replace(day=1).strftime("%Y-%m-%d")
            if today.month == 12:
                date_to = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                date_to = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
            return date_from.strftime("%Y-%m-%d"), date_to.strftime("%Y-%m-%d")
    
    def get_quarter_name(self):
        """Pobiera nazwÄ™ kwartaÅ‚u"""
        if self.report_type == "quarterly":
            quarter = self.quarter_combo.currentIndex() + 1
            year = self.quarter_year_spin.value()
            
            quarter_names = {
                1: "I kwartaÅ‚ (styczeÅ„-marzec)",
                2: "II kwartaÅ‚ (kwiecieÅ„-czerwiec)",
                3: "III kwartaÅ‚ (lipiec-wrzesieÅ„)",
                4: "IV kwartaÅ‚ (paÅºdziernik-grudzieÅ„)"
            }
            
            return f"{quarter_names[quarter]} {year}"
        return None

    def get_consolidated_sales_data(self, date_from, date_to):
        """Pobiera i Å‚Ä…czy dane sprzedaÅ¼y w ramach jednego zamÃ³wienia"""
        try:
            # Pobierz szczegÃ³Å‚owe dane sprzedaÅ¼y
            detailed_sales = self.db.get_detailed_sales(date_from, date_to)
            if not detailed_sales:
                return []
            
            # Grupuj po order_id
            orders_dict = {}
            for sale in detailed_sales:
                order_id = sale['order_id']
                if order_id not in orders_dict:
                    orders_dict[order_id] = {
                        'order_id': order_id,
                        'platform': sale['platform'],
                        'date': sale['date'],
                        'total_pln': sale['order_total_pln'],
                        'total_eur': sale['order_total_eur'],
                        'total_cost': sale['order_total_cost'],
                        'total_profit': sale['order_total_pln'] - sale['order_total_cost'],
                        'products': [],
                        'quantities': [],
                        'revenue_items': [],
                        'cost_items': [],
                        'profit_items': []
                    }
                
                # Dodaj produkt do listy - SKRÃ“CONA NAZWA DLA WYÅšWIETLANIA
                product_info = f"{sale['sku']} - {sale['title']} (x{sale['qty']})"
                orders_dict[order_id]['products'].append(product_info)
                orders_dict[order_id]['quantities'].append(sale['qty'])
                orders_dict[order_id]['revenue_items'].append(sale['item_revenue_pln'])
                orders_dict[order_id]['cost_items'].append(sale['item_cost'])
                orders_dict[order_id]['profit_items'].append(sale['item_profit'])
            
            # PrzeksztaÅ‚Ä‡ do listy
            consolidated_sales = []
            for order_data in orders_dict.values():
                # Oblicz Å‚Ä…cznÄ… iloÅ›Ä‡
                total_qty = sum(order_data['quantities'])
                
                # Przelicz na rzeczywiste wartoÅ›ci (suma kosztÃ³w, nie Å›rednia)
                actual_total_cost = sum(order_data['cost_items'])
                
                # Tworzymy listÄ™ produktÃ³w w formacie poziomym (oddzielone przecinkami)
                # UMIARKOWANE OGRANICZENIE DÅUGOÅšCI - 150 ZNAKÃ“W
                products_horizontal = ", ".join(order_data['products'])
                if len(products_horizontal) > 150:
                    products_horizontal = products_horizontal[:147] + "..."
                
                consolidated_sales.append({
                    'order_id': order_data['order_id'],
                    'date': order_data['date'],
                    'platform': order_data['platform'],
                    'products_list': order_data['products'],
                    'products_horizontal': products_horizontal,
                    'total_quantity': total_qty,
                    'revenue_pln': order_data['total_pln'],
                    'cost_pln': actual_total_cost,
                    'profit_pln': order_data['total_pln'] - actual_total_cost,
                    'revenue_per_item': order_data['total_pln'] / total_qty if total_qty > 0 else 0,
                    'products_display': products_horizontal if self.cb_show_products.isChecked() else f"{len(order_data['products'])} produktÃ³w"
                })
            
            return consolidated_sales
            
        except Exception as e:
            print(f"BÅ‚Ä…d w get_consolidated_sales_data: {e}")
            import traceback
            traceback.print_exc()
            return []

    def print_report(self):
        """BezpoÅ›rednie drukowanie raportu"""
        try:
            from PySide6.QtPrintSupport import QPrinter, QPrintDialog
            
            printer = QPrinter()
            
            if printer.printerName() == "":
                QMessageBox.information(self, "Brak drukarki", 
                    "Nie znaleziono domyÅ›lnej drukarki. Skonfiguruj drukarkÄ™ w systemie.")
                return
            
            date_from, date_to = self.get_date_range()
            personal_data = self.config.get_business_info() if self.cb_simple_register.isChecked() else {}
            
            if self.cb_simple_register.isChecked():
                required_fields = ['name', 'address', 'postal_code', 'city', 'pesel']
                missing_fields = [field for field in required_fields if not personal_data.get(field)]
                
                if missing_fields:
                    QMessageBox.warning(self, "Brak danych", 
                                      f"UzupeÅ‚nij dane osobowe w konfiguracji.\nBrakujÄ…ce pola: {', '.join(missing_fields)}")
                    return
            
            # Pobierz dane sprzedaÅ¼y
            if self.cb_consolidate_sales.isChecked():
                sales_data = self.get_consolidated_sales_data(date_from, date_to)
                register_data = None
            else:
                register_data = self.db.get_simple_sales_register_with_cumulative(date_from, date_to, personal_data or {})
                sales_data = None
            
            if not sales_data and (not register_data or not register_data.get("transakcje")):
                QMessageBox.warning(self, "Brak danych", "Brak danych do wydrukowania w wybranym okresie.")
                return
            
            html_content = self.generate_print_html(date_from, date_to, sales_data, register_data, personal_data)
            
            print_dialog = QPrintDialog(printer, self)
            print_dialog.setWindowTitle("Drukuj raport")
            
            if print_dialog.exec() == QPrintDialog.Accepted:
                from PySide6.QtGui import QTextDocument
                document = QTextDocument()
                document.setHtml(html_content)
                document.print_(printer)
                
                QMessageBox.information(self, "Drukowanie", "Raport zostaÅ‚ wysÅ‚any do druku.")
                
        except ImportError:
            QMessageBox.warning(self, "BÅ‚Ä…d drukowania", 
                              "ModuÅ‚ drukowania nie jest dostÄ™pny.\nUpewnij siÄ™, Å¼e PySide6 jest poprawnie zainstalowany.")
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d drukowania", 
                               f"WystÄ…piÅ‚ bÅ‚Ä…d podczas drukowania:\n{str(e)}")

    def generate_print_html(self, date_from, date_to, sales_data=None, register_data=None, personal_data=None):
        """Generuje HTML do drukowania z uwzglÄ™dnieniem limitÃ³w kwartalnych"""
        try:
            # TytuÅ‚ raportu
            if self.report_type == "monthly":
                month_name = self.month_combo.currentText()
                year = self.year_spin.value()
                title = f"Raport miesiÄ™czny - {month_name} {year}"
            elif self.report_type == "yearly":
                year = self.year_only_spin.value()
                title = f"Raport roczny - {year}"
            elif self.report_type == "quarterly":
                quarter_name = self.get_quarter_name()
                title = f"Raport kwartalny - {quarter_name}"
            else:
                title = f"Raport okresowy - {date_from} do {date_to}"
            
            # Dane sprzedawcy - tylko jeÅ›li zaznaczono uproszczony rejestr
            seller_info = ""
            if self.cb_simple_register.isChecked() and personal_data and personal_data.get('name'):
                seller_info = f"""
                <div style="margin-bottom: 12px;">
                    <h3 style="font-size: 11px; margin-bottom: 4px;">Dane sprzedawcy:</h3>
                    <p style="font-size: 9px; margin: 1px 0;"><b>ImiÄ™ i nazwisko:</b> {personal_data.get('name', '')}</p>
                    <p style="font-size: 9px; margin: 1px 0;"><b>Adres:</b> {personal_data.get('address', '')}</p>
                    <p style="font-size: 9px; margin: 1px 0;"><b>Kod pocztowy i miejscowoÅ›Ä‡:</b> {personal_data.get('postal_code', '')} {personal_data.get('city', '')}</p>
                    <p style="font-size: 9px; margin: 1px 0;"><b>PESEL:</b> {personal_data.get('pesel', '')}</p>
                    {"<p style='font-size: 9px; margin: 1px 0;'><b>NIP:</b> " + personal_data.get('nip', '') + "</p>" if personal_data.get('nip') else ""}
                    {"<p style='font-size: 9px; margin: 1px 0;'><b>REGON:</b> " + personal_data.get('regon', '') + "</p>" if personal_data.get('regon') else ""}
                </div>
                """
            
            # Oblicz podsumowanie
            total_revenue = 0
            total_cost = 0
            total_profit = 0
            total_transactions = 0
            
            if sales_data:
                total_transactions = len(sales_data)
                for sale in sales_data:
                    total_revenue += sale['revenue_pln']
                    total_cost += sale['cost_pln']
                    total_profit += sale['profit_pln']
            elif register_data and register_data.get("transakcje"):
                summary = register_data["podsumowanie_ogolne"]
                total_revenue = summary['przychod_calkowity']
                total_cost = summary['koszt_calkowity']
                total_profit = summary['zysk_calkowity']
                total_transactions = summary['liczba_transakcji']
            
            # Podsumowanie ogÃ³lne
            summary_html = f"""
            <div style="margin-bottom: 12px;">
                <h3 style="font-size: 11px; margin-bottom: 4px;">Podsumowanie ogÃ³lne:</h3>
                <table border="1" cellpadding="3" style="border-collapse: collapse; width: 100%; font-size: 9px;">
                    <tr>
                        <td><b>PrzychÃ³d caÅ‚kowity:</b></td>
                        <td align="right">{total_revenue:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>Koszt caÅ‚kowity:</b></td>
                        <td align="right">{total_cost:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>Zysk caÅ‚kowity:</b></td>
                        <td align="right">{total_profit:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>Liczba transakcji:</b></td>
                        <td align="right">{total_transactions}</td>
                    </tr>
                </table>
            </div>
            """
            
            # Analiza limitÃ³w US - POKAZUJ ZAWSZE, nie tylko dla uproszczonego rejestru!
            year = int(date_from[:4])
            minimal_wage = self.config.get_minimal_wage(year)
            
            # OkreÅ›l czy to raport kwartalny
            is_quarterly_report = self.report_type == "quarterly"
            
            if is_quarterly_report and self.config.use_quarterly_limits():
                # UÅ¼ywamy limitÃ³w kwartalnych
                limit_type = "kwartalny"
                limit_multiplier = self.config.get_limits_config().get("quarterly_limit_multiplier", 2.25)
                limit = minimal_wage * limit_multiplier
                limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia"
            else:
                # UÅ¼ywamy limitÃ³w miesiÄ™cznych
                limit_type = "miesiÄ™czny"
                limit = minimal_wage * 0.75
                limit_text = "75% minimalnego wynagrodzenia"
            
            analysis_html = f"""
            <div style="margin-bottom: 12px;">
                <h3 style="font-size: 11px; margin-bottom: 4px;">Analiza progu dla US ({year} r.):</h3>
                <table border="1" cellpadding="3" style="border-collapse: collapse; width: 100%; font-size: 9px;">
                    <tr>
                        <td><b>Minimalne wynagrodzenie:</b></td>
                        <td align="right">{minimal_wage:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>{limit_text} (limit {limit_type}):</b></td>
                        <td align="right">{limit:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>PrzychÃ³d narastajÄ…co:</b></td>
                        <td align="right">{total_revenue:.2f} PLN</td>
                    </tr>
                    <tr>
                        <td><b>Stan:</b></td>
                        <td align="right" style="color: {'red' if total_revenue > limit else 'green'}; font-weight: bold;">
                            {'PRZEKROCZONO LIMIT!' if total_revenue > limit else 'W LIMICIE'}
                        </td>
                    </tr>
                </table>
            </div>
            """
            
            # Dodatkowa informacja dla raportÃ³w kwartalnych
            if is_quarterly_report and self.config.use_quarterly_limits():
                analysis_html += f"""
                <div style="margin-bottom: 12px; padding: 6px; background-color: #f0f0f0; border-left: 3px solid #2E7D32; font-size: 8px;">
                    <p><b>UWAGA:</b> Od 2026 roku obowiÄ…zujÄ… limity kwartalne dla dziaÅ‚alnoÅ›ci nierejestrowanej.</p>
                    <p>Limit kwartalny wynosi {limit_multiplier*100:.0f}% minimalnego wynagrodzenia.</p>
                    <p>PrzychÃ³d w tym kwartale: <b>{total_revenue:.2f} PLN</b></p>
                </div>
                """
            
            # SzczegÃ³Å‚owa ewidencja transakcji - ZOPTYMALIZOWANA DLA A4
            transactions_html = ""
            if sales_data:
                trans_rows = ""
                for sale in sales_data:
                    # UÅ¼ywamy formatu poziomego zamiast pionowego
                    products_display = sale['products_horizontal']
                    
                    trans_rows += f"""
                    <tr>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{sale['date']}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{sale['platform']}</td>
                        <td style="font-size: 6px; padding: 2px; border: 1px solid #ddd; max-width: 160px; word-wrap: break-word; text-align: left;">{products_display}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{sale['total_quantity']}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{sale['revenue_pln']:.2f}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{sale['cost_pln']:.2f}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{sale['profit_pln']:.2f}</td>
                    </tr>
                    """
                
                transactions_html = f"""
                <div style="margin-bottom: 15px; page-break-before: always;">
                    <h4 style="font-size: 9px; margin-bottom: 3px; font-weight: bold; color: #333;">SzczegÃ³Å‚owa ewidencja transakcji (poÅ‚Ä…czone):</h4>
                    <table border="0" cellpadding="0" cellspacing="0" style="border-collapse: collapse; width: 100%; font-size: 7px; table-layout: fixed;">
                        <thead>
                            <tr style="background-color: #f2f2f2;">
                                <th style="width: 50px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Data</th>
                                <th style="width: 60px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Platforma</th>
                                <th style="width: 160px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Produkty</th>
                                <th style="width: 35px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">IloÅ›Ä‡</th>
                                <th style="width: 55px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">WartoÅ›Ä‡</th>
                                <th style="width: 55px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Koszt</th>
                                <th style="width: 55px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Zysk</th>
                            </tr>
                        </thead>
                        <tbody>
                            {trans_rows}
                        </tbody>
                    </table>
                    <p style="font-size: 6px; color: #666; margin-top: 2px; font-style: italic;">
                        * Produkty wyÅ›wietlane poziomo, nazwy skrÃ³cone do 150 znakÃ³w dla czytelnoÅ›ci
                    </p>
                </div>
                """
            elif register_data and register_data.get("transakcje"):
                # Stary format dla niepoÅ‚Ä…czonych transakcji
                trans_rows = ""
                for transaction in register_data["transakcje"]:
                    product_display = f"{transaction['nazwa_produktu']} ({transaction['kod_produktu']})"
                    if len(product_display) > 80:
                        product_display = product_display[:77] + "..."
                    
                    trans_rows += f"""
                    <tr>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{transaction['data_sprzedazy']}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{transaction['platforma']}</td>
                        <td style="font-size: 6px; padding: 2px; border: 1px solid #ddd; max-width: 120px; word-wrap: break-word; text-align: left;">{product_display}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: center;">{transaction['ilosc']}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{transaction['wartosc_sprzedazy_pln']:.2f}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{transaction['koszt_zakupu']:.2f}</td>
                        <td style="font-size: 7px; padding: 2px; border: 1px solid #ddd; text-align: right;">{transaction['zysk_brutto']:.2f}</td>
                    </tr>
                    """
                
                transactions_html = f"""
                <div style="margin-bottom: 15px; page-break-before: always;">
                    <h4 style="font-size: 9px; margin-bottom: 3px; font-weight: bold; color: #333;">SzczegÃ³Å‚owa ewidencja transakcji (niepoÅ‚Ä…czone):</h4>
                    <table border="0" cellpadding="0" cellspacing="0" style="border-collapse: collapse; width: 100%; font-size: 7px; table-layout: fixed;">
                        <thead>
                            <tr style="background-color: #f2f2f2;">
                                <th style="width: 50px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Data</th>
                                <th style="width: 70px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Platforma</th>
                                <th style="width: 120px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Produkt</th>
                                <th style="width: 40px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">IloÅ›Ä‡</th>
                                <th style="width: 60px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">WartoÅ›Ä‡</th>
                                <th style="width: 60px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Koszt</th>
                                <th style="width: 60px; font-size: 7px; padding: 3px; border: 1px solid #ddd; font-weight: bold; text-align: center;">Zysk</th>
                            </tr>
                        </thead>
                        <tbody>
                            {trans_rows}
                        </tbody>
                    </table>
                </div>
                """
            
            # Stopka
            footer = f"""
            <div style="margin-top: 15px; border-top: 1px solid #ccc; padding-top: 6px; font-size: 7px; color: #666;">
                <p>Raport wygenerowany: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p>Typ raportu: {self.report_type}</p>
                <p>ÅÄ…czenie sprzedaÅ¼y: {'Tak' if self.cb_consolidate_sales.isChecked() else 'Nie'}</p>
                <p>Uproszczony rejestr: {'Tak' if self.cb_simple_register.isChecked() else 'Nie'}</p>
                <p>System Magazynowo-SprzedaÅ¼owy v{APP_VERSION}</p>
            </div>
            """
            
            # CaÅ‚y dokument HTML - ZOPTYMALIZOWANY DLA DRUKU
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <title>{title}</title>
                <style>
                    @media print {{
                        @page {{
                            size: A4;
                            margin: 8mm 10mm 8mm 10mm;
                        }}
                        body {{
                            margin: 0;
                            padding: 0;
                            font-family: Arial, sans-serif;
                            font-size: 8px;
                            color: black;
                        }}
                        h1 {{
                            color: #2E7D32;
                            border-bottom: 1px solid #2E7D32;
                            padding-bottom: 4px;
                            font-size: 12px;
                            margin-bottom: 8px;
                        }}
                        h2 {{
                            color: #333;
                            margin-top: 10px;
                            font-size: 10px;
                            margin-bottom: 5px;
                        }}
                        h3 {{
                            color: #333;
                            margin-top: 8px;
                            font-size: 9px;
                            margin-bottom: 4px;
                        }}
                        h4 {{
                            color: #333;
                            margin-top: 6px;
                            font-size: 8px;
                            margin-bottom: 3px;
                        }}
                        p {{
                            margin: 2px 0;
                            font-size: 8px;
                        }}
                        table {{
                            margin-top: 5px;
                            width: 100%;
                            table-layout: fixed;
                            border-collapse: collapse;
                        }}
                        th {{
                            background-color: #f2f2f2;
                            padding: 3px;
                            text-align: left;
                            font-weight: bold;
                            font-size: 7px;
                            border: 1px solid #ddd;
                        }}
                        td {{
                            padding: 2px;
                            vertical-align: top;
                            border: 1px solid #ddd;
                            font-size: 7px;
                        }}
                        .footer {{
                            margin-top: 12px;
                            border-top: 1px solid #ccc;
                            padding-top: 5px;
                            font-size: 7px;
                            color: #666;
                        }}
                    }}
                    @media screen {{
                        body {{
                            font-family: Arial, sans-serif;
                            margin: 8px;
                            font-size: 8px;
                        }}
                        h1 {{
                            color: #2E7D32;
                            border-bottom: 1px solid #2E7D32;
                            padding-bottom: 5px;
                            font-size: 12px;
                        }}
                        h2 {{ font-size: 10px; }}
                        h3 {{ font-size: 9px; }}
                        h4 {{ font-size: 8px; }}
                        p {{ font-size: 8px; }}
                        table {{ 
                            width: 100%; 
                            table-layout: fixed;
                            border-collapse: collapse;
                        }}
                        th {{ 
                            background-color: #f2f2f2; 
                            padding: 3px; 
                            border: 1px solid #ddd; 
                            font-size: 7px;
                        }}
                        td {{ 
                            padding: 2px; 
                            border: 1px solid #ddd; 
                            font-size: 7px;
                        }}
                    }}
                </style>
            </head>
            <body>
                <h1>{title}</h1>
                <p><b>Okres:</b> {date_from} - {date_to}</p>
                <p><b>Wygenerowano:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                <p><b>ÅÄ…czenie sprzedaÅ¼y:</b> {'Tak' if self.cb_consolidate_sales.isChecked() else 'Nie'}</p>
                <p><b>Uproszczony rejestr:</b> {'Tak' if self.cb_simple_register.isChecked() else 'Nie'}</p>
                
                {seller_info}
                {summary_html}
                {analysis_html}
                {transactions_html}
                {footer}
            </body>
            </html>
            """
            
            return html
            
        except Exception as e:
            print(f"BÅ‚Ä…d generowania HTML: {e}")
            return f"<html><body><h1>BÅ‚Ä…d generowania raportu</h1><p>{str(e)}</p></body></html>"

    def generate_report(self):
        try:
            date_from, date_to = self.get_date_range()
            
            if not self.cb_purchases.isChecked() and not self.cb_sales.isChecked():
                QMessageBox.warning(self, "Brak danych", "Wybierz przynajmniej jeden typ danych (zakupy lub sprzedaÅ¼).")
                return
            
            personal_data = {}
            if self.cb_simple_register.isChecked():
                personal_data = self.config.get_business_info()
                required_fields = ['name', 'address', 'postal_code', 'city', 'pesel']
                missing_fields = [field for field in required_fields if not personal_data.get(field)]
                
                if missing_fields:
                    QMessageBox.warning(self, "Brak danych", 
                                      f"UzupeÅ‚nij dane osobowe w konfiguracji.\nBrakujÄ…ce pola: {', '.join(missing_fields)}")
                    return
            
            # Ustal format pliku
            if self.rb_csv.isChecked():
                file_filter = "CSV Files (*.csv)"
                default_ext = ".csv"
            elif self.rb_excel.isChecked():
                file_filter = "Excel Files (*.xlsx)"
                default_ext = ".xlsx"
            else:  # PDF
                file_filter = "PDF Files (*.pdf)"
                default_ext = ".pdf"
            
            # Sugerowana nazwa pliku
            if self.report_type == "monthly":
                month_name = self.month_combo.currentText().lower()
                year = self.year_spin.value()
                suggested_name = f"raport_{month_name}_{year}{default_ext}"
            elif self.report_type == "yearly":
                year = self.year_only_spin.value()
                suggested_name = f"raport_{year}{default_ext}"
            elif self.report_type == "quarterly":
                quarter = self.quarter_combo.currentIndex() + 1
                year = self.quarter_year_spin.value()
                suggested_name = f"raport_kwartal{quarter}_{year}{default_ext}"
            else:
                from_str = date_from.replace("-", "")
                to_str = date_to.replace("-", "")
                suggested_name = f"raport_{from_str}_do_{to_str}{default_ext}"
            
            # Dodaj info o typie raportu do nazwy
            if self.cb_simple_register.isChecked():
                suggested_name = suggested_name.replace(default_ext, f"_us{default_ext}")
            
            path, _ = QFileDialog.getSaveFileName(
                self, "Zapisz raport", 
                os.path.join(os.getcwd(), suggested_name),
                file_filter
            )
            
            if path:
                try:
                    if self.rb_csv.isChecked():
                        success = self.export_consolidated_report_csv(
                            path, date_from, date_to,
                            personal_data=personal_data if self.cb_simple_register.isChecked() else None,
                            report_type=self.report_type,
                            config=self.config
                        )
                    elif self.rb_excel.isChecked():
                        success = self.export_consolidated_report_excel(
                            path, date_from, date_to,
                            personal_data=personal_data if self.cb_simple_register.isChecked() else None,
                            report_type=self.report_type,
                            config=self.config
                        )
                    else:  # PDF
                        success = self.export_consolidated_report_pdf(
                            path, date_from, date_to,
                            personal_data=personal_data if self.cb_simple_register.isChecked() else None
                        )
                    
                    if success:
                        QMessageBox.information(self, "Sukces", f"Raport zostaÅ‚ wygenerowany:\n{path}")
                        self.accept()
                    else:
                        QMessageBox.warning(self, "BÅ‚Ä…d", "Nie udaÅ‚o siÄ™ wygenerowaÄ‡ raportu.")
                        
                except Exception as e:
                    error_msg = str(e)
                    if "openpyxl" in error_msg.lower():
                        error_msg += "\n\nUpewnij siÄ™, Å¼e openpyxl jest zainstalowany:\npip install openpyxl"
                    QMessageBox.critical(self, "BÅ‚Ä…d", f"WystÄ…piÅ‚ bÅ‚Ä…d podczas generowania raportu:\n{error_msg}")
                    
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", f"WystÄ…piÅ‚ bÅ‚Ä…d:\n{str(e)}")
            import traceback
            print(traceback.format_exc())

    def export_consolidated_report_csv(self, path, date_from, date_to, personal_data=None, report_type=None, config=None):
        """Eksportuje raport CSV z poÅ‚Ä…czonymi sprzedaÅ¼ami"""
        try:
            import csv
            
            sales_data = self.get_consolidated_sales_data(date_from, date_to)
            
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                
                title = "RAPORT SPRZEDAÅ»Y (POÅÄ„CZONY)"
                if report_type == "quarterly":
                    title = "RAPORT KWARTALNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
                elif report_type == "monthly":
                    title = "RAPORT MIESIÄ˜CZNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
                elif report_type == "yearly":
                    title = "RAPORT ROCZNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
                
                w.writerow([title])
                w.writerow([f"Okres: {date_from} - {date_to}"])
                w.writerow([f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                w.writerow([f"ÅÄ…czenie sprzedaÅ¼y: TAK"])
                w.writerow([])
                
                # Dane sprzedawcy - tylko jeÅ›li zaznaczono uproszczony rejestr
                if personal_data:
                    w.writerow(["DANE SPRZEDAWCY"])
                    w.writerow([f"ImiÄ™ i nazwisko: {personal_data.get('name', '')}"])
                    w.writerow([f"Adres: {personal_data.get('address', '')}"])
                    w.writerow([f"Kod pocztowy: {personal_data.get('postal_code', '')}"])
                    w.writerow([f"MiejscowoÅ›Ä‡: {personal_data.get('city', '')}"])
                    w.writerow([f"PESEL: {personal_data.get('pesel', '')}"])
                    if personal_data.get('nip'):
                        w.writerow([f"NIP: {personal_data.get('nip', '')}"])
                    if personal_data.get('regon'):
                        w.writerow([f"REGON: {personal_data.get('regon', '')}"])
                    w.writerow([])
                
                # Podsumowanie
                total_revenue = sum(sale['revenue_pln'] for sale in sales_data)
                total_cost = sum(sale['cost_pln'] for sale in sales_data)
                total_profit = total_revenue - total_cost
                
                w.writerow(["PODSUMOWANIE OGÃ“LNE"])
                w.writerow(["PrzychÃ³d caÅ‚kowity:", f"{total_revenue:.2f} PLN"])
                w.writerow(["Koszt caÅ‚kowity:", f"{total_cost:.2f} PLN"])
                w.writerow(["Zysk caÅ‚kowity:", f"{total_profit:.2f} PLN"])
                w.writerow(["Liczba transakcji (zamÃ³wieÅ„):", len(sales_data)])
                w.writerow([])
                
                # Analiza limitÃ³w US - POKAZUJ ZAWSZE, nie tylko dla uproszczonego rejestru!
                year = int(date_from[:4])
                minimal_wage = config.get_minimal_wage(year) if config else 4242.00
                
                is_quarterly = report_type == "quarterly"
                use_quarterly = config.use_quarterly_limits() if config else True
                
                if is_quarterly and use_quarterly:
                    limit_multiplier = config.get_limits_config().get("quarterly_limit_multiplier", 2.25) if config else 2.25
                    limit = minimal_wage * limit_multiplier
                    limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia (limit kwartalny)"
                else:
                    limit = minimal_wage * 0.75
                    limit_text = "75% minimalnego wynagrodzenia (limit miesiÄ™czny)"
                
                w.writerow([f"ANALIZA PROGU LIMITU ({year} r.)"])
                w.writerow([f"Minimalne wynagrodzenie: {minimal_wage} PLN"])
                w.writerow([f"{limit_text}: {limit:.2f} PLN"])
                w.writerow([f"PrzychÃ³d narastajÄ…co w roku {year}: {total_revenue:.2f} PLN"])
                
                if total_revenue > limit:
                    w.writerow(["UWAGA: Przekroczono limit dziaÅ‚alnoÅ›ci nierejestrowanej!"])
                    w.writerow(["Konieczna rejestracja dziaÅ‚alnoÅ›ci gospodarczej"])
                else:
                    w.writerow(["OK: PrzychÃ³d mieÅ›ci siÄ™ w limicie dziaÅ‚alnoÅ›ci nierejestrowanej"])
                
                w.writerow([])
                
                # SzczegÃ³Å‚owa ewidencja
                w.writerow(["SZCZEGÃ“ÅOWA EWIDENCJA TRANSAKCJI (POÅÄ„CZONE)"])
                w.writerow(["Data", "Platforma", "Produkty", "IloÅ›Ä‡ Å‚Ä…cznie", "WartoÅ›Ä‡ sprzedaÅ¼y", "Koszt zakupu", "Zysk"])
                
                for sale in sales_data:
                    products_display = sale['products_horizontal']
                    
                    w.writerow([
                        sale['date'],
                        sale['platform'],
                        products_display,
                        sale['total_quantity'],
                        f"{sale['revenue_pln']:.2f}",
                        f"{sale['cost_pln']:.2f}",
                        f"{sale['profit_pln']:.2f}"
                    ])
                
                return True
                
        except Exception as e:
            print(f"BÅ‚Ä…d w export_consolidated_report_csv: {e}")
            return False

    def export_consolidated_report_excel(self, path, date_from, date_to, personal_data=None, report_type=None, config=None):
        """Eksportuje raport Excel z poÅ‚Ä…czonymi sprzedaÅ¼ami"""
        if not HAS_EXCEL:
            raise ImportError("Biblioteka openpyxl nie jest zainstalowana. UÅ¼yj: pip install openpyxl")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            sales_data = self.get_consolidated_sales_data(date_from, date_to)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Raport sprzedaÅ¼y"
            
            # Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            money_format = '#,##0.00'
            
            row = 1
            
            # TytuÅ‚
            title = "RAPORT SPRZEDAÅ»Y (POÅÄ„CZONY)"
            if report_type == "quarterly":
                title = "RAPORT KWARTALNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
            elif report_type == "monthly":
                title = "RAPORT MIESIÄ˜CZNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
            elif report_type == "yearly":
                title = "RAPORT ROCZNY SPRZEDAÅ»Y (POÅÄ„CZONY)"
            
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
            row += 2
            
            ws.cell(row=row, column=1, value=f"Okres: {date_from} - {date_to}").font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=1, value=f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            row += 1
            ws.cell(row=row, column=1, value=f"ÅÄ…czenie sprzedaÅ¼y: TAK")
            row += 1
            ws.cell(row=row, column=1, value=f"Uproszczony rejestr: {'TAK' if personal_data else 'NIE'}")
            row += 2
            
            # Dane sprzedawcy - tylko jeÅ›li zaznaczono uproszczony rejestr
            if personal_data:
                ws.cell(row=row, column=1, value="DANE SPRZEDAWCY").font = Font(bold=True)
                row += 1
                
                ws.cell(row=row, column=1, value=f"ImiÄ™ i nazwisko: {personal_data.get('name', '')}")
                row += 1
                ws.cell(row=row, column=1, value=f"Adres: {personal_data.get('address', '')}")
                row += 1
                ws.cell(row=row, column=1, value=f"Kod pocztowy: {personal_data.get('postal_code', '')}")
                row += 1
                ws.cell(row=row, column=1, value=f"MiejscowoÅ›Ä‡: {personal_data.get('city', '')}")
                row += 1
                ws.cell(row=row, column=1, value=f"PESEL: {personal_data.get('pesel', '')}")
                row += 2
            
            # Podsumowanie
            total_revenue = sum(sale['revenue_pln'] for sale in sales_data)
            total_cost = sum(sale['cost_pln'] for sale in sales_data)
            total_profit = total_revenue - total_cost
            
            ws.cell(row=row, column=1, value="PODSUMOWANIE OGÃ“LNE").font = Font(bold=True)
            row += 1
            
            summary_data = [
                ["PrzychÃ³d caÅ‚kowity:", total_revenue],
                ["Koszt caÅ‚kowity:", total_cost],
                ["Zysk caÅ‚kowity:", total_profit],
                ["Liczba transakcji (zamÃ³wieÅ„):", len(sales_data)]
            ]
            
            for label, value in summary_data:
                ws.cell(row=row, column=1, value=label)
                if isinstance(value, (int, float)):
                    ws.cell(row=row, column=2, value=value)
                    ws.cell(row=row, column=2).number_format = money_format
                else:
                    ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 1
            
            # Analiza limitÃ³w US - POKAZUJ ZAWSZE
            year = int(date_from[:4])
            minimal_wage = config.get_minimal_wage(year) if config else 4242.00
            
            is_quarterly = report_type == "quarterly"
            use_quarterly = config.use_quarterly_limits() if config else True
            
            if is_quarterly and use_quarterly:
                limit_multiplier = config.get_limits_config().get("quarterly_limit_multiplier", 2.25) if config else 2.25
                limit = minimal_wage * limit_multiplier
                limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia (limit kwartalny)"
            else:
                limit = minimal_wage * 0.75
                limit_text = "75% minimalnego wynagrodzenia (limit miesiÄ™czny)"
            
            ws.cell(row=row, column=1, value=f"ANALIZA PROGU LIMITU ({year} r.)").font = Font(bold=True, color="FF0000")
            row += 1
            
            limit_data = [
                ["Minimalne wynagrodzenie:", minimal_wage],
                [limit_text, limit],
                ["PrzychÃ³d narastajÄ…co:", total_revenue],
                ["Stan:", "PRZEKROCZONO LIMIT!" if total_revenue > limit else "W LIMICIE"]
            ]
            
            for label, value in limit_data:
                ws.cell(row=row, column=1, value=label)
                if isinstance(value, (int, float)):
                    ws.cell(row=row, column=2, value=value)
                    ws.cell(row=row, column=2).number_format = money_format
                else:
                    ws.cell(row=row, column=2, value=value)
                row += 1
            
            row += 2
            
            # NagÅ‚Ã³wki tabeli
            headers = ["Data", "Platforma", "Produkty", "IloÅ›Ä‡ Å‚Ä…cznie", "WartoÅ›Ä‡ sprzedaÅ¼y", "Koszt zakupu", "Zysk"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row += 1
            
            # Dane - FORMAT POZIOMY Z UMIARKOWANYMI NAZWAMI
            for sale in sales_data:
                ws.cell(row=row, column=1, value=sale['date'])
                ws.cell(row=row, column=2, value=sale['platform'])
                
                # Produkty w formacie poziomym (oddzielone przecinkami)
                products_display = sale['products_horizontal']
                ws.cell(row=row, column=3, value=products_display)
                ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True, vertical='top')
                
                ws.cell(row=row, column=4, value=sale['total_quantity'])
                ws.cell(row=row, column=5, value=sale['revenue_pln']).number_format = money_format
                ws.cell(row=row, column=6, value=sale['cost_pln']).number_format = money_format
                ws.cell(row=row, column=7, value=sale['profit_pln']).number_format = money_format
                
                row += 1
            
            # Ustaw szerokoÅ›ci kolumn - ROZSÄ„DNA SZEROKOÅšÄ† DLA PRODUKTÃ“W
            column_widths = [12, 15, 40, 12, 15, 15, 15]  # Kolumna Produkty ma 40 znakÃ³w
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            wb.save(path)
            return True
            
        except Exception as e:
            print(f"BÅ‚Ä…d w export_consolidated_report_excel: {e}")
            return False

    def export_consolidated_report_pdf(self, path, date_from, date_to, personal_data=None):
        """Eksportuje raport PDF z poÅ‚Ä…czonymi sprzedaÅ¼ami"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # PrÃ³buj uÅ¼yÄ‡ czcionki z polskimi znakami
            font_paths = [
                '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                'C:/Windows/Fonts/arial.ttf',
                'arial.ttf'
            ]
            
            font_name = 'Helvetica'
            for font_path in font_paths:
                try:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('PolishFont', font_path))
                        font_name = 'PolishFont'
                        break
                except:
                    continue
            
            sales_data = self.get_consolidated_sales_data(date_from, date_to)
            
            doc = SimpleDocTemplate(
                path,
                pagesize=A4,
                rightMargin=1.2*cm,
                leftMargin=1.2*cm,
                topMargin=1.2*cm,
                bottomMargin=1.2*cm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Styl tytuÅ‚u
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=11,
                spaceAfter=12,
                alignment=1
            )
            
            # Styl normalny
            normal_style = ParagraphStyle(
                'Normal',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=7,
                spaceAfter=3
            )
            
            # TytuÅ‚ raportu
            if self.report_type == "quarterly":
                quarter_name = self.get_quarter_name()
                story.append(Paragraph(f"RAPORT KWARTALNY - {quarter_name}", title_style))
            elif self.report_type == "monthly":
                month_name = self.month_combo.currentText()
                year = self.year_spin.value()
                story.append(Paragraph(f"RAPORT MIESIÄ˜CZNY - {month_name} {year}", title_style))
            elif self.report_type == "yearly":
                year = self.year_only_spin.value()
                story.append(Paragraph(f"RAPORT ROCZNY - {year}", title_style))
            else:
                story.append(Paragraph("RAPORT SZCZEGÃ“ÅOWY", title_style))
                
            story.append(Spacer(1, 6))
            
            # Okres raportu
            story.append(Paragraph(f"Okres: {date_from} - {date_to}", normal_style))
            story.append(Paragraph(f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
            story.append(Paragraph(f"ÅÄ…czenie sprzedaÅ¼y: TAK", normal_style))
            story.append(Paragraph(f"Uproszczony rejestr: {'TAK' if personal_data else 'NIE'}", normal_style))
            story.append(Spacer(1, 10))
            
            # Dane sprzedawcy - tylko jeÅ›li zaznaczono uproszczony rejestr
            if personal_data:
                story.append(Paragraph("DANE SPRZEDAWCY", ParagraphStyle('Subtitle', parent=styles['Heading2'], fontName=font_name, fontSize=8)))
                
                seller_info = [
                    f"ImiÄ™ i nazwisko: {personal_data.get('name', '')}",
                    f"Adres: {personal_data.get('address', '')}",
                    f"{personal_data.get('postal_code', '')} {personal_data.get('city', '')}",
                    f"PESEL: {personal_data.get('pesel', '')}"
                ]
                
                if personal_data.get('nip'):
                    seller_info.append(f"NIP: {personal_data.get('nip', '')}")
                if personal_data.get('regon'):
                    seller_info.append(f"REGON: {personal_data.get('regon', '')}")
                
                for info in seller_info:
                    story.append(Paragraph(info, normal_style))
                
                story.append(Spacer(1, 10))
            
            # Podsumowanie
            total_revenue = sum(sale['revenue_pln'] for sale in sales_data)
            total_cost = sum(sale['cost_pln'] for sale in sales_data)
            total_profit = total_revenue - total_cost
            
            story.append(Paragraph("PODSUMOWANIE OGÃ“LNE", ParagraphStyle('Subtitle', parent=styles['Heading2'], fontName=font_name, fontSize=8)))
            
            summary_data = [
                ["PrzychÃ³d caÅ‚kowity:", f"{total_revenue:.2f} PLN"],
                ["Koszt caÅ‚kowity:", f"{total_cost:.2f} PLN"],
                ["Zysk caÅ‚kowity:", f"{total_profit:.2f} PLN"],
                ["Liczba transakcji:", str(len(sales_data))]
            ]
            
            summary_table = Table(summary_data, colWidths=[5*cm, 3.5*cm])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 10))
            
            # Analiza limitÃ³w US - POKAZUJ ZAWSZE
            year = int(date_from[:4])
            minimal_wage = self.config.get_minimal_wage(year)
            
            is_quarterly = self.report_type == "quarterly"
            use_quarterly = self.config.use_quarterly_limits()
            
            if is_quarterly and use_quarterly:
                limit_multiplier = self.config.get_limits_config().get("quarterly_limit_multiplier", 2.25)
                limit = minimal_wage * limit_multiplier
                limit_text = f"{limit_multiplier*100:.0f}% minimalnego wynagrodzenia (limit kwartalny)"
            else:
                limit = minimal_wage * 0.75
                limit_text = "75% minimalnego wynagrodzenia (limit miesiÄ™czny)"
            
            story.append(Paragraph("ANALIZA PROGU DLA US", ParagraphStyle('Subtitle', parent=styles['Heading2'], fontName=font_name, fontSize=8)))
            
            limit_data = [
                ["Minimalne wynagrodzenie:", f"{minimal_wage:.2f} PLN"],
                [limit_text, f"{limit:.2f} PLN"],
                ["PrzychÃ³d narastajÄ…co:", f"{total_revenue:.2f} PLN"],
                ["Stan:", "PRZEKROCZONO LIMIT!" if total_revenue > limit else "W LIMICIE"]
            ]
            
            limit_table = Table(limit_data, colWidths=[5*cm, 3.5*cm])
            limit_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('TEXTCOLOR', (1, 3), (1, 3), colors.red if total_revenue > limit else colors.green),
            ]))
            
            story.append(limit_table)
            story.append(Spacer(1, 10))
            
            # Dodatkowa informacja dla raportÃ³w kwartalnych
            if is_quarterly and use_quarterly:
                info_text = Paragraph(
                    f"<b>UWAGA:</b> Od 2026 roku obowiÄ…zujÄ… limity kwartalne dla dziaÅ‚alnoÅ›ci nierejestrowanej.<br/>"
                    f"Limit kwartalny wynosi {limit_multiplier*100:.0f}% minimalnego wynagrodzenia.<br/>"
                    f"PrzychÃ³d w tym kwartale: <b>{total_revenue:.2f} PLN</b>",
                    ParagraphStyle('Info', parent=styles['Normal'], fontName=font_name, fontSize=6,
                                 backColor=colors.lightgrey, borderPadding=6,
                                 leftIndent=8, rightIndent=8)
                )
                story.append(info_text)
                story.append(Spacer(1, 10))
            
            # SzczegÃ³Å‚owa ewidencja - NOWA STRONA Z ZOPTYMALIZOWANÄ„ TABELÄ„
            if sales_data:
                story.append(PageBreak())
                story.append(Paragraph("SZCZEGÃ“ÅOWA EWIDENCJA TRANSAKCJI (POÅÄ„CZONE)", 
                                     ParagraphStyle('Subtitle', parent=styles['Heading2'], fontName=font_name, fontSize=8)))
                
                headers = ["Data", "Platforma", "Produkty", "IloÅ›Ä‡", "WartoÅ›Ä‡", "Koszt", "Zysk"]
                
                # ZOPTYMALIZOWANE SZEROKOÅšCI KOLUMN DLA A4 - ZACHOWUJÄ„CE MIEJSCE
                col_widths = [1.2*cm, 1.8*cm, 7.0*cm, 1.2*cm, 2.0*cm, 2.0*cm, 2.0*cm]
                
                # NagÅ‚Ã³wki tabeli
                header_data = []
                for header in headers:
                    header_data.append(Paragraph(header, ParagraphStyle('Header', parent=styles['Normal'], 
                                                                       fontName=font_name, fontSize=6, 
                                                                       alignment=1, textColor=colors.white)))
                
                table_data = [header_data]
                
                # Dane transakcji Z UMIARKOWANYMI NAZWAMI
                for sale in sales_data:
                    # UÅ¼ywamy formatu poziomego (oddzielone przecinkami)
                    products_text = sale['products_horizontal']
                    
                    row = [
                        Paragraph(sale['date'], ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=1)),
                        Paragraph(sale['platform'], ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=1)),
                        Paragraph(products_text, ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=4, alignment=0)),
                        Paragraph(str(sale['total_quantity']), ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=1)),
                        Paragraph(f"{sale['revenue_pln']:.2f} PLN", ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=2)),
                        Paragraph(f"{sale['cost_pln']:.2f} PLN", ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=2)),
                        Paragraph(f"{sale['profit_pln']:.2f} PLN", ParagraphStyle('Cell', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=2))
                    ]
                    table_data.append(row)
                
                # Tworzymy tabelÄ™ z zoptymalizowanymi szerokoÅ›ciami
                transactions_table = Table(table_data, colWidths=col_widths, repeatRows=1)
                transactions_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 5),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('ALIGN', (3, 0), (3, -1), 'CENTER'),
                    ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ]))
                
                story.append(transactions_table)
                story.append(Spacer(1, 6))
                
                # Informacja o ograniczeniach
                story.append(Paragraph("* Nazwy produktÃ³w ograniczone do 150 znakÃ³w dla zachowania czytelnoÅ›ci", 
                                     ParagraphStyle('Info', parent=styles['Normal'], fontName=font_name, fontSize=4, 
                                                  alignment=0, textColor=colors.grey)))
            else:
                story.append(Paragraph("Brak danych sprzedaÅ¼y w wybranym okresie", normal_style))
            
            # Stopka
            story.append(Spacer(1, 12))
            footer = Paragraph(f"Raport wygenerowany przez System Magazynowo-SprzedaÅ¼owy v{APP_VERSION}", 
                             ParagraphStyle('Footer', parent=styles['Normal'], fontName=font_name, fontSize=5, alignment=1))
            story.append(footer)
            
            doc.build(story)
            return True
            
        except Exception as e:
            print(f"BÅ‚Ä…d w export_consolidated_report_pdf: {e}")
            import traceback
            traceback.print_exc()
            return False
# ================== SORTOWANIE ==================
class SortableTableWidget(QTableWidget):
    def __init__(self, rows=0, columns=0, parent=None):
        super().__init__(rows, columns, parent)
        self.sort_order = {}
        self.current_sorted_column = -1
        self.current_sort_order = Qt.AscendingOrder
        
        self.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        
    def on_header_clicked(self, column):
        if column in self.sort_order:
            self.sort_order[column] = not self.sort_order[column]
        else:
            self.sort_order[column] = True
        
        sort_ascending = self.sort_order[column]
        
        self.current_sorted_column = column
        self.current_sort_order = Qt.AscendingOrder if sort_ascending else Qt.DescendingOrder
        
        self.sort_by_column(column, self.current_sort_order)
        self.mark_sorted_column(column)
    
    def sort_by_column(self, column, order):
        self.sortItems(column, order)
    
    def mark_sorted_column(self, column):
        header = self.horizontalHeader()
        for i in range(header.count()):
            if i == column:
                sort_text = " â†‘" if self.sort_order.get(column, True) else " â†“"
                original_text = self.get_column_name(i)
                header.model().setHeaderData(i, Qt.Horizontal, original_text + sort_text)
                
                header.setStyleSheet("""
                    QHeaderView::section {
                        background-color: #e0e0e0;
                        font-weight: bold;
                    }
                """)
            else:
                original_text = self.get_column_name(i)
                header.model().setHeaderData(i, Qt.Horizontal, original_text)
        
        if column == -1:
            header.setStyleSheet("""
                QHeaderView::section {
                    background-color: #f2f2f2;
                    font-weight: normal;
                }
            """)
    
    def get_column_name(self, column):
        header = self.horizontalHeader()
        original_text = header.model().headerData(column, Qt.Horizontal)
        
        if original_text:
            for arrow in [" â†‘", " â†“"]:
                if original_text.endswith(arrow):
                    return original_text[:-2]
            return original_text
        
        names = ["ID", "SKU", "Nazwa", "Stan", ""]
        if column < len(names):
            return names[column]
        return f"Kolumna {column+1}"
    
    def load_data(self, data):
        self.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                
                if j in [0, 3]:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    try:
                        item.setData(Qt.EditRole, float(value))
                    except ValueError:
                        pass
                
                self.setItem(i, j, item)
        
        self.sort_order = {}
        self.current_sorted_column = -1
        self.mark_sorted_column(-1)

# ================== HISTORIA Z ZAZNACZANIEM ==================
class HistoryDialog(QDialog):
    def __init__(self, title, headers, rows, delete_cb, parent=None, allow_multiple=True):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(1000, 550)
        self.rows = rows
        self.delete_cb = delete_cb
        self.headers = headers
        self.allow_multiple = allow_multiple

        v = QVBoxLayout(self)

        button_panel = QHBoxLayout()
        
        self.select_all_checkbox = QCheckBox("Zaznacz wszystkie")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        button_panel.addWidget(self.select_all_checkbox)
        
        button_panel.addStretch()
        
        b_select_none = QPushButton("Odznacz wszystkie")
        b_select_none.clicked.connect(self.select_none)
        button_panel.addWidget(b_select_none)
        
        v.addLayout(button_panel)

        self.table = SortableTableWidget()
        self.table.setColumnCount(len(headers) + 1)
        table_headers = ["âœ“"] + list(headers)
        self.table.setHorizontalHeaderLabels(table_headers)
        
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        for i in range(1, len(table_headers)):
            self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)
        
        v.addWidget(self.table)

        delete_panel = QHBoxLayout()
        
        b_del_selected = QPushButton("ðŸ—‘ï¸ UsuÅ„ zaznaczone")
        b_del_selected.setStyleSheet("background-color: #d32f2f; font-weight: bold;")
        b_del_selected.clicked.connect(self.delete_selected)
        delete_panel.addWidget(b_del_selected)
        
        b_del_single = QPushButton("UsuÅ„ pojedynczy")
        b_del_single.clicked.connect(self.delete_single)
        delete_panel.addWidget(b_del_single)
        
        delete_panel.addStretch()
        
        b_close = QPushButton("Zamknij")
        b_close.clicked.connect(self.accept)
        delete_panel.addWidget(b_close)
        
        v.addLayout(delete_panel)

        self.load()

    def load(self):
        self.table.setRowCount(len(self.rows))
        for i, r in enumerate(self.rows):
            checkbox = QCheckBox()
            checkbox.setChecked(False)
            self.table.setCellWidget(i, 0, checkbox)
            
            for j, val in enumerate(r):
                item = QTableWidgetItem(str(val))
                
                if j in [0, 3, 4]:
                    if isinstance(val, (int, float)):
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        try:
                            item.setData(Qt.EditRole, float(val))
                        except (ValueError, TypeError):
                            pass
                
                self.table.setItem(i, j + 1, item)

    def toggle_select_all(self, state):
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(state == Qt.Checked)

    def select_none(self):
        self.select_all_checkbox.setChecked(False)
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(False)

    def get_selected_ids(self):
        selected_ids = []
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                id_item = self.table.item(row, 1)
                if id_item:
                    try:
                        selected_ids.append(int(id_item.text()))
                    except ValueError:
                        pass
        return selected_ids

    def delete_selected(self):
        selected_ids = self.get_selected_ids()
        if not selected_ids:
            QMessageBox.warning(self, "Brak zaznaczenia", "Nie zaznaczono Å¼adnych wpisÃ³w do usuniÄ™cia.")
            return
        
        count = len(selected_ids)
        if QMessageBox.question(
            self, "PotwierdÅº usuniÄ™cie",
            f"Czy na pewno usunÄ…Ä‡ {count} zaznaczonych wpisÃ³w?\n"
            f"Spowoduje to takÅ¼e usuniÄ™cie odpowiednich iloÅ›ci z magazynu.",
            QMessageBox.Yes | QMessageBox.No
        ) == QMessageBox.Yes:
            for order_id in selected_ids:
                self.delete_cb(order_id)
            self.accept()

    def delete_single(self):
        selected_ids = self.get_selected_ids()
        if len(selected_ids) == 0:
            QMessageBox.warning(self, "Brak zaznaczenia", "Nie zaznaczono Å¼adnego wpisu do usuniÄ™cia.")
            return
        elif len(selected_ids) > 1:
            QMessageBox.warning(self, "Za duÅ¼o zaznaczonych", "Zaznaczono wiÄ™cej niÅ¼ jeden wpis. UÅ¼yj 'UsuÅ„ zaznaczone' dla wielu pozycji.")
            return
        
        order_id = selected_ids[0]
        if QMessageBox.question(self, "PotwierdÅº", "UsunÄ…Ä‡ wybrany wpis?") == QMessageBox.Yes:
            self.delete_cb(order_id)
            self.accept()

# ================== INWENTARYZACJA ==================
class InventoryDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Inwentaryzacja magazynu")
        self.resize(800, 450)

        v = QVBoxLayout(self)

        self.table = SortableTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["ID", "SKU", "Nazwa", "Stan systemowy", "Stan rzeczywisty"]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        v.addWidget(self.table)

        b = QPushButton("Zapisz korekty")
        b.clicked.connect(self.apply)
        v.addWidget(b)

        self.load()

    def load(self):
        rows = self.db.list_products()
        data = []
        for r in rows:
            data.append([
                r["id"],
                r["sku"],
                r["title"],
                r["stock"],
                r["stock"]
            ])
        self.table.load_data(data)

    def apply(self):
        for r in range(self.table.rowCount()):
            pid = int(self.table.item(r, 0).text())
            system = int(self.table.item(r, 3).text())
            real = int(self.table.item(r, 4).text())
            delta = real - system
            if delta != 0:
                self.db.update_stock(pid, delta)
        QMessageBox.information(self, "OK", "Inwentaryzacja zapisana")
        self.accept()

# ================== ZAKUP ==================
class PurchaseDialog(QDialog):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.setWindowTitle("Dodaj zakup")
        v = QVBoxLayout(self)

        self.cost = QDoubleSpinBox()
        self.cost.setMaximum(1e9)
        self.date = QDateEdit(QDate.currentDate())
        self.date.setCalendarPopup(True)

        form = QFormLayout()
        form.addRow("Koszt PLN", self.cost)
        form.addRow("Data", self.date)
        v.addLayout(form)

        self.items = SortableTableWidget(0, 2)
        self.items.setHorizontalHeaderLabels(["Produkt", "IloÅ›Ä‡"])
        v.addWidget(self.items)

        b_add = QPushButton("Dodaj pozycjÄ™")
        b_add.clicked.connect(self.add_item)
        v.addWidget(b_add)

        b_ok = QPushButton("Zapisz")
        b_ok.clicked.connect(self.accept)
        v.addWidget(b_ok)

    def add_item(self):
        r = self.items.rowCount()
        self.items.insertRow(r)
        
        combo = product_combo(self.db)
        self.items.setCellWidget(r, 0, combo)

        qty = QSpinBox()
        qty.setMinimum(1)
        qty.setMaximum(100000)
        self.items.setCellWidget(r, 1, qty)

    def get_items(self):
        return [
            (
                self.items.cellWidget(r, 0).currentData(),
                self.items.cellWidget(r, 1).value()
            )
            for r in range(self.items.rowCount())
        ]

# ================== DIALOG HISTORII RACHUNKÃ“W ==================
class InvoicesHistoryDialog(QDialog):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db = db
        self.config = config
        self.setWindowTitle("Historia wygenerowanych rachunkÃ³w")
        self.resize(1200, 600)
        
        v = QVBoxLayout(self)
        
        # Filtr daty
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Od:"))
        self.date_from = QDateEdit(QDate.currentDate().addMonths(-1))
        self.date_from.setCalendarPopup(True)
        filter_layout.addWidget(self.date_from)
        
        filter_layout.addWidget(QLabel("Do:"))
        self.date_to = QDateEdit(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        filter_layout.addWidget(self.date_to)
        
        btn_filter = QPushButton("Filtruj")
        btn_filter.clicked.connect(self.load_invoices)
        filter_layout.addWidget(btn_filter)
        
        btn_refresh = QPushButton("OdÅ›wieÅ¼")
        btn_refresh.clicked.connect(self.load_invoices)
        filter_layout.addWidget(btn_refresh)
        
        # Przycisk resetu numeracji
        btn_reset_counter = QPushButton("âŸ³ Resetuj licznik")
        btn_reset_counter.setToolTip("Resetuj licznik numeracji rachunkÃ³w")
        btn_reset_counter.clicked.connect(self.reset_invoice_counter)
        filter_layout.addWidget(btn_reset_counter)
        
        filter_layout.addStretch()
        v.addLayout(filter_layout)
        
        # Tabela z rachunkami
        self.table = SortableTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "âœ“", "Numer rachunku", "Data wystawienia", "Klient", "Kwota", 
            "Platforma", "Data sprzedaÅ¼y", "ÅšcieÅ¼ka pliku", "Akcje"
        ])
        
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        
        v.addWidget(self.table)
        
        # Przyciski akcji
        button_layout = QHBoxLayout()
        
        self.select_all_checkbox = QCheckBox("Zaznacz wszystkie")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all)
        button_layout.addWidget(self.select_all_checkbox)
        
        btn_open_selected = QPushButton("ðŸ— OtwÃ³rz zaznaczone")
        btn_open_selected.clicked.connect(self.open_selected_invoices)
        button_layout.addWidget(btn_open_selected)
        
        btn_delete_selected = QPushButton("ðŸ—‘ UsuÅ„ zaznaczone")
        btn_delete_selected.setStyleSheet("background-color: #d32f2f; font-weight: bold;")
        btn_delete_selected.clicked.connect(self.delete_selected_invoices)
        button_layout.addWidget(btn_delete_selected)
        
        button_layout.addStretch()
        
        btn_close = QPushButton("Zamknij")
        btn_close.clicked.connect(self.accept)
        button_layout.addWidget(btn_close)
        
        v.addLayout(button_layout)
        
        self.load_invoices()
    
    def load_invoices(self):
        date_from = self.date_from.date().toString("yyyy-MM-dd")
        date_to = self.date_to.date().toString("yyyy-MM-dd")
        
        invoices = self.db.list_invoices(date_from, date_to)
        self.table.setRowCount(len(invoices))
        
        for i, inv in enumerate(invoices):
            # Checkbox
            checkbox = QCheckBox()
            self.table.setCellWidget(i, 0, checkbox)
            
            # Numer rachunku
            self.table.setItem(i, 1, QTableWidgetItem(inv['invoice_number']))
            
            # Data wystawienia
            self.table.setItem(i, 2, QTableWidgetItem(inv['issue_date']))
            
            # Klient
            customer = inv['customer_name']
            if not customer and inv['sale_order_id']:
                customer = f"ZamÃ³wienie #{inv['sale_order_id']}"
            self.table.setItem(i, 3, QTableWidgetItem(customer or "Brak danych"))
            
            # Kwota
            amount_item = QTableWidgetItem(f"{inv['total_amount']:.2f} PLN")
            amount_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.table.setItem(i, 4, amount_item)
            
            # Platforma
            self.table.setItem(i, 5, QTableWidgetItem(inv['platform'] or "Brak"))
            
            # Data sprzedaÅ¼y
            self.table.setItem(i, 6, QTableWidgetItem(inv['sale_date'] or inv['issue_date']))
            
            # ÅšcieÅ¼ka pliku
            path_item = QTableWidgetItem(inv['file_path'])
            path_item.setToolTip(inv['file_path'])
            self.table.setItem(i, 7, path_item)
            
            # Przyciski akcji
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(4, 4, 4, 4)
            action_layout.setSpacing(4)
            
            btn_open = QPushButton("ðŸ“„ OtwÃ³rz")
            btn_open.setFixedWidth(80)
            btn_open.setToolTip("OtwÃ³rz plik PDF")
            btn_open.clicked.connect(lambda checked, path=inv['file_path']: self.open_invoice(path))
            action_layout.addWidget(btn_open)
            
            btn_delete = QPushButton("ðŸ—‘ UsuÅ„")
            btn_delete.setFixedWidth(80)
            btn_delete.setStyleSheet("background-color: #ff4444; color: white;")
            btn_delete.setToolTip("UsuÅ„ rachunek i plik PDF")
            btn_delete.clicked.connect(lambda checked, inv_id=inv['id'], inv_num=inv['invoice_number']: 
                                      self.delete_single_invoice(inv_id, inv_num))
            action_layout.addWidget(btn_delete)
            
            action_layout.addStretch()
            self.table.setCellWidget(i, 8, action_widget)
    
    def toggle_select_all(self, state):
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(state == Qt.Checked)
    
    def get_selected_invoice_ids(self):
        selected_ids = []
        for row in range(self.table.rowCount()):
            checkbox = self.table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                invoice_number_item = self.table.item(row, 1)
                if invoice_number_item:
                    invoice_number = invoice_number_item.text()
                    invoice = self.db.get_invoice_by_number(invoice_number)
                    if invoice:
                        selected_ids.append(invoice['id'])
        return selected_ids
    
    def open_selected_invoices(self):
        selected_ids = self.get_selected_invoice_ids()
        if not selected_ids:
            QMessageBox.warning(self, "Brak zaznaczenia", "Nie zaznaczono Å¼adnych rachunkÃ³w.")
            return
        
        for inv_id in selected_ids:
            invoices = self.db.list_invoices()
            for inv in invoices:
                if inv['id'] == inv_id and os.path.exists(inv['file_path']):
                    self.open_invoice(inv['file_path'])
    
    def open_invoice(self, file_path):
        if os.path.exists(file_path):
            try:
                if os.name == 'nt':
                    os.startfile(file_path)
                elif os.name == 'posix':
                    import subprocess
                    subprocess.call(['open', file_path] if sys.platform == 'darwin' 
                                   else ['xdg-open', file_path])
            except Exception as e:
                QMessageBox.warning(self, "BÅ‚Ä…d", f"Nie moÅ¼na otworzyÄ‡ pliku:\n{file_path}\n\nBÅ‚Ä…d: {str(e)}")
        else:
            QMessageBox.warning(self, "Brak pliku", f"Plik nie istnieje:\n{file_path}")
    
    def delete_selected_invoices(self):
        selected_ids = self.get_selected_invoice_ids()
        if not selected_ids:
            QMessageBox.warning(self, "Brak zaznaczenia", "Nie zaznaczono Å¼adnych rachunkÃ³w do usuniÄ™cia.")
            return
        
        count = len(selected_ids)
        reply = QMessageBox.question(
            self, "PotwierdÅº usuniÄ™cie",
            f"Czy na pewno usunÄ…Ä‡ {count} zaznaczonych rachunkÃ³w?\n"
            f"UWAGA: Spowoduje to rÃ³wnieÅ¼ usuniÄ™cie plikÃ³w PDF z dysku!",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success_count = 0
            for inv_id in selected_ids:
                if self.db.delete_invoice(inv_id):
                    success_count += 1
            
            QMessageBox.information(self, "UsuniÄ™to", 
                                  f"UsuniÄ™to {success_count} z {count} rachunkÃ³w.")
            self.load_invoices()
    
    def delete_single_invoice(self, invoice_id, invoice_number):
        reply = QMessageBox.question(
            self, "PotwierdÅº usuniÄ™cie",
            f"Czy na pewno usunÄ…Ä‡ rachunek:\n{invoice_number}?\n\n"
            f"UWAGA: Spowoduje to rÃ³wnieÅ¼ usuniÄ™cie pliku PDF z dysku!",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if self.db.delete_invoice(invoice_id):
                QMessageBox.information(self, "UsuniÄ™to", f"Rachunek {invoice_number} zostaÅ‚ usuniÄ™ty.")
                self.load_invoices()
    
    def reset_invoice_counter(self):
        """Resetuje licznik numeracji rachunkÃ³w"""
        reply = QMessageBox.question(
            self, "Resetuj licznik",
            "Czy na pewno chcesz zresetowaÄ‡ licznik numeracji rachunkÃ³w?\n\n"
            "UWAGA: To ustawi nastÄ™pny numer rachunku na 1.\n"
            "MoÅ¼e spowodowaÄ‡ duplikaty numerÃ³w jeÅ›li stare rachunki nadal istniejÄ…!",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Pobierz aktualnÄ… konfiguracjÄ™
            current_config = self.config.load()
            
            # Resetuj licznik
            current_config["invoice"]["next_number"] = 1
            
            # Zapisz z powrotem
            with open("config.json", 'w', encoding='utf-8') as f:
                import json
                json.dump(current_config, f, ensure_ascii=False, indent=2)
            
            # PrzeÅ‚aduj konfiguracjÄ™
            self.config = Config()
            
            QMessageBox.information(self, "Licznik zresetowany", 
                                  "Licznik numeracji rachunkÃ³w zostaÅ‚ zresetowany do 1.")

# ================== SPRZEDAÅ» Z RACHUNKIEM ==================
class SaleDialog(QDialog):
    def __init__(self, db, config):
        super().__init__()
        self.db = db
        self.config = config
        self.setWindowTitle("Dodaj sprzedaÅ¼")
        v = QVBoxLayout(self)
        
        # Platforma z moÅ¼liwoÅ›ciÄ… wprowadzenia wÅ‚asnej nazwy dla opcji "Inne"
        self.platform = QComboBox()
        self.platform.addItems(PLATFORMS)
        self.platform.currentTextChanged.connect(self.on_platform_changed)
        
        # Pole do wprowadzenia wÅ‚asnej nazwy platformy (widoczne tylko dla "Inne")
        self.custom_platform_label = QLabel("WprowadÅº nazwÄ™ platformy:")
        self.custom_platform_label.setVisible(False)
        self.custom_platform_input = QLineEdit()
        self.custom_platform_input.setPlaceholderText("np. Allegro, eBay, wÅ‚asny sklep...")
        self.custom_platform_input.setVisible(False)
        
        self.pln = QDoubleSpinBox()
        self.pln.setMaximum(1e9)
        self.pln.valueChanged.connect(self.update_fifo_cost)
        
        self.date = QDateEdit(QDate.currentDate())
        self.date.setCalendarPopup(True)

        # Koszt automatyczny z FIFO
        self.auto_cost_label = QLabel("Koszt zakupu (FIFO): 0.00 PLN")
        self.auto_cost_label.setStyleSheet("font-weight: bold; color: #c62828;")
        
        # Checkbox dla rachunku uproszczonego
        self.create_invoice_checkbox = QCheckBox("Wygeneruj rachunek uproszczony")
        self.create_invoice_checkbox.setChecked(False)
        self.create_invoice_checkbox.toggled.connect(self.toggle_invoice_fields)
        
        # Dane klienta (opcjonalne) - UKRYTE DOMYÅšLNIE
        self.client_group = QGroupBox("Dane klienta (opcjonalnie dla rachunku)")
        self.client_group.setVisible(False)
        
        client_layout = QFormLayout()
        
        self.client_name = QLineEdit()
        self.client_name.setPlaceholderText("ImiÄ™ i nazwisko/Nazwa firmy")
        client_layout.addRow("Nabywca:", self.client_name)
        
        self.client_address = QLineEdit()
        self.client_address.setPlaceholderText("Adres")
        client_layout.addRow("Adres:", self.client_address)
        
        self.client_group.setLayout(client_layout)
        
        form = QFormLayout()
        form.addRow("Platforma", self.platform)
        form.addRow(self.custom_platform_label)
        form.addRow(self.custom_platform_input)
        form.addRow("Cena PLN", self.pln)
        form.addRow("Data", self.date)
        form.addRow("", self.auto_cost_label)
        form.addRow("", self.create_invoice_checkbox)
        v.addLayout(form)
        v.addWidget(self.client_group)

        # Tabela z pozycjami
        self.items = SortableTableWidget(0, 2)
        self.items.setHorizontalHeaderLabels(["Produkt", "IloÅ›Ä‡"])
        v.addWidget(self.items)

        b_add = QPushButton("Dodaj pozycjÄ™")
        b_add.clicked.connect(self.add_item)
        v.addWidget(b_add)

        # Przyciski
        button_layout = QHBoxLayout()
        
        b_ok = QPushButton("Zapisz i wygeneruj rachunek")
        b_ok.clicked.connect(self.save_sale_with_invoice)
        b_ok.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        b_ok_simple = QPushButton("Zapisz bez rachunku")
        b_ok_simple.clicked.connect(self.save_sale_without_invoice)
        
        b_cancel = QPushButton("Anuluj")
        b_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(b_ok)
        button_layout.addWidget(b_ok_simple)
        button_layout.addWidget(b_cancel)
        
        v.addLayout(button_layout)
        
        self.fifo_cost = 0.0

    def on_platform_changed(self, platform_name):
        """Pokazuje/ukrywa pole do wprowadzenia wÅ‚asnej nazwy platformy"""
        if platform_name == "Inne":
            self.custom_platform_label.setVisible(True)
            self.custom_platform_input.setVisible(True)
            self.custom_platform_input.setFocus()
        else:
            self.custom_platform_label.setVisible(False)
            self.custom_platform_input.setVisible(False)

    def get_platform_name(self):
        """Zwraca wÅ‚aÅ›ciwÄ… nazwÄ™ platformy (uÅ¼ytkownika lub domyÅ›lnÄ…)"""
        if self.platform.currentText() == "Inne":
            custom_name = self.custom_platform_input.text().strip()
            return custom_name if custom_name else "Inne"
        return self.platform.currentText()

    def toggle_invoice_fields(self, checked):
        """Pokazuje/ukrywa pola danych klienta w zaleÅ¼noÅ›ci od zaznaczenia checkboxa"""
        self.client_group.setVisible(checked)
        
        if checked:
            self.client_group.setTitle("Dane klienta (opcjonalnie dla rachunku)")
            self.client_name.setEnabled(True)
            self.client_address.setEnabled(True)
        else:
            self.client_name.clear()
            self.client_address.clear()
            self.client_name.setEnabled(False)
            self.client_address.setEnabled(False)

    def add_item(self):
        r = self.items.rowCount()
        self.items.insertRow(r)
        
        combo = product_combo(self.db)
        combo.currentIndexChanged.connect(self.update_fifo_cost)
        self.items.setCellWidget(r, 0, combo)

        qty = QSpinBox()
        qty.setMinimum(1)
        qty.setMaximum(100000)
        qty.valueChanged.connect(self.update_fifo_cost)
        self.items.setCellWidget(r, 1, qty)
        
        QTimer.singleShot(100, self.update_fifo_cost)

    def get_items(self):
        items = []
        for r in range(self.items.rowCount()):
            combo = self.items.cellWidget(r, 0)
            qty_widget = self.items.cellWidget(r, 1)
            if combo and qty_widget:
                pid = combo.currentData()
                qty = qty_widget.value()
                items.append((pid, qty))
        return items

    def update_fifo_cost(self):
        try:
            total_cost = 0.0
            items = self.get_items()
            
            for pid, qty in items:
                fifo_batches = self.db.get_fifo_batches(pid, qty)
                if not fifo_batches:
                    QMessageBox.warning(self, "Brak stanu", 
                        f"Brak wystarczajÄ…cego stanu dla produktu ID: {pid}\n"
                        f"Wymagane: {qty}, dostÄ™pne: 0")
                    total_cost = 0.0
                    break
                
                for batch in fifo_batches:
                    batch_qty = min(qty, batch["available_qty"])
                    total_cost += batch["unit_cost"] * batch_qty
                    qty -= batch_qty
                    if qty <= 0:
                        break
            
            self.fifo_cost = total_cost
            self.auto_cost_label.setText(f"Koszt zakupu (FIFO): {total_cost:.2f} PLN")
            return total_cost
        except Exception as e:
            print(f"BÅ‚Ä…d w obliczaniu kosztu FIFO: {e}")
            self.fifo_cost = 0.0
            self.auto_cost_label.setText("Koszt zakupu (FIFO): 0.00 PLN")
            return 0.0

    def generate_invoice(self, sale_id, items, total_pln, fifo_cost, profit):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            # PrÃ³buj uÅ¼yÄ‡ czcionki z polskimi znakami
            font_paths = [
                '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
                'C:/Windows/Fonts/arial.ttf',
                'arial.ttf'
            ]
            
            font_name = 'Helvetica'
            for font_path in font_paths:
                try:
                    if os.path.exists(font_path):
                        pdfmetrics.registerFont(TTFont('PolishFont', font_path))
                        font_name = 'PolishFont'
                        break
                except:
                    continue
            
            business_info = self.config.get_business_info()
            invoice_config = self.config.get_invoice_config()
            
            invoice_number = self.config.get_next_invoice_number()
            
            # SprawdÅº czy zapisywaÄ‡ PDF
            if not self.config.should_save_pdf():
                # Tymczasowy plik
                import tempfile
                filename = f"rachunek_{invoice_number.replace('/', '_')}.pdf"
                path = os.path.join(tempfile.gettempdir(), filename)
            else:
                filename = f"rachunek_{invoice_number.replace('/', '_')}.pdf"
                path = os.path.join(os.getcwd(), "rachunki", filename)
                os.makedirs(os.path.dirname(path), exist_ok=True)
            
            doc = SimpleDocTemplate(
                path,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            # Styl tytuÅ‚u
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=16,
                spaceAfter=30,
                alignment=1
            )
            
            # Styl normalny
            normal_style = ParagraphStyle(
                'Normal',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                spaceAfter=6
            )
            
            # Styl nagÅ‚Ã³wka tabeli
            header_style = ParagraphStyle(
                'Header',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=10,
                textColor=colors.whitesmoke,
                alignment=1
            )
            
            # Styl stopki
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=9,
                alignment=1,
                textColor=colors.grey
            )
            
            # TytuÅ‚
            story.append(Paragraph("RACHUMEK UPROSZCZONY", title_style))
            story.append(Spacer(1, 20))
            
            # Dane sprzedawcy
            seller_data = [
                ["SPRZEDAWCA:", ""],
                [f"ImiÄ™ i nazwisko: {business_info.get('name', '')}", ""],
                [f"Adres: {business_info.get('address', '')}", ""],
                [f"{business_info.get('postal_code', '')} {business_info.get('city', '')}", ""],
                [f"PESEL: {business_info.get('pesel', '')}", ""],
            ]
            
            # Dodaj opcjonalne dane firmy
            company_name = business_info.get('company_name', '')
            company_address = business_info.get('company_address', '')
            phone = business_info.get('phone', '')
            email = business_info.get('email', '')
            
            if company_name:
                seller_data.insert(1, [f"Nazwa firmy: {company_name}", ""])
            if company_address:
                seller_data.insert(2, [f"Adres firmy: {company_address}", ""])
            if phone:
                seller_data.append([f"Telefon: {phone}", ""])
            if email:
                seller_data.append([f"Email: {email}", ""])
            
            if business_info.get('nip'):
                seller_data.append([f"NIP: {business_info.get('nip', '')}", ""])
            
            if business_info.get('regon'):
                seller_data.append([f"REGON: {business_info.get('regon', '')}", ""])
            
            seller_table = Table(seller_data, colWidths=[12*cm, 6*cm])
            seller_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
            ]))
            
            story.append(seller_table)
            story.append(Spacer(1, 20))
            
            # Dane nabywcy
            client_name = self.client_name.text().strip()
            client_address = self.client_address.text().strip()
            
            if client_name or client_address:
                buyer_data = [
                    ["NABYWCA:", ""],
                    [f"Nazwa: {client_name}", ""],
                    [f"Adres: {client_address}", ""],
                ]
                
                buyer_table = Table(buyer_data, colWidths=[12*cm, 6*cm])
                buyer_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
                ]))
                
                story.append(buyer_table)
                story.append(Spacer(1, 20))
            
            # Informacje o fakturze
            invoice_data = [
                ["Numer rachunku:", invoice_number],
                ["Data sprzedaÅ¼y:", self.date.date().toString("dd.MM.yyyy")],
                ["Data wystawienia:", QDate.currentDate().toString("dd.MM.yyyy")],
                ["Platforma:", self.get_platform_name()],
            ]
            
            invoice_table = Table(invoice_data, colWidths=[6*cm, 12*cm])
            invoice_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
            ]))
            story.append(invoice_table)
            story.append(Spacer(1, 30))
            
            # Pozycje faktury
            items_data = [["Lp.", "Nazwa towaru/usÅ‚ugi", "IloÅ›Ä‡", "Cena jdn. (PLN)", "WartoÅ›Ä‡ (PLN)"]]
            
            for i, (pid, qty) in enumerate(self.get_items(), 1):
                product_info = self.db.get_product_info(pid)
                if product_info:
                    item_price = total_pln / sum(q for _, q in self.get_items()) if self.get_items() else 0
                    item_total = item_price * qty
                    items_data.append([
                        str(i),
                        f"{product_info['sku']} - {product_info['title']}",
                        str(qty),
                        f"{item_price:.2f}",
                        f"{item_total:.2f}"
                    ])
            
            items_table = Table(items_data, colWidths=[1*cm, 9*cm, 2*cm, 3*cm, 3*cm])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            
            story.append(items_table)
            story.append(Spacer(1, 20))
            
            # Podsumowanie
            summary_data = [
                ["RAZEM DO ZAPÅATY:", f"{total_pln:.2f} PLN"],
                ["W tym:", ""],
                ["- koszt wÅ‚asny:", f"{fifo_cost:.2f} PLN"],
                ["- zysk brutto:", f"{profit:.2f} PLN"],
            ]
            
            summary_table = Table(summary_data, colWidths=[12*cm, 6*cm])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (0, 0), 12),
                ('FONTSIZE', (1, 0), (1, 0), 12),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('TEXTCOLOR', (1, 0), (1, 0), colors.red),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # Stopka
            footer_text = self.config.get_invoice_footer()
            story.append(Paragraph(footer_text, footer_style))
            
            story.append(Paragraph("Rachunek uproszczony jest dokumentem sprzedaÅ¼y dla potrzeb ewidencji przychodÃ³w", 
                                 ParagraphStyle('Info', parent=styles['Normal'], fontName=font_name, fontSize=8, alignment=1)))
            
            doc.build(story)
            
            # Zapisz fakturÄ™ do bazy danych
            success = self.db.add_invoice(
                invoice_number=invoice_number,
                sale_order_id=sale_id,
                file_path=path,
                customer_name=client_name,
                customer_address=client_address,
                issue_date=QDate.currentDate().toString("yyyy-MM-dd"),
                total_amount=total_pln
            )
            
            if not success:
                print("Uwaga: Nie udaÅ‚o siÄ™ zapisaÄ‡ faktury do bazy danych")
            
            # OtwÃ³rz automatycznie jeÅ›li ustawione
            if self.config.should_auto_open_invoice():
                import subprocess
                try:
                    if os.name == 'nt':
                        os.startfile(path)
                    elif os.name == 'posix':
                        subprocess.call(['open', path] if sys.platform == 'darwin' else ['xdg-open', path])
                except:
                    pass
            
            return path, invoice_number
            
        except Exception as e:
            print(f"BÅ‚Ä…d generowania rachunku: {e}")
            import traceback
            traceback.print_exc()
            return None, None
    
    def save_sale_with_invoice(self):
        if not self.create_invoice_checkbox.isChecked():
            QMessageBox.warning(self, "Informacja", 
                              "Aby wygenerowaÄ‡ rachunek, zaznacz opcjÄ™ 'Wygeneruj rachunek uproszczony'")
            return self.save_sale_without_invoice()
        
        business_info = self.config.get_business_info()
        required_fields = ['name', 'address', 'postal_code', 'city', 'pesel']
        for field in required_fields:
            if not business_info.get(field):
                QMessageBox.warning(self, "Brak danych", 
                                  f"UzupeÅ‚nij dane sprzedawcy w konfiguracji.\nBrakujÄ…ce pole: {field}")
                return
        
        sale_id = self._save_sale()
        if sale_id:
            invoice_path, invoice_number = self.generate_invoice(
                sale_id, 
                self.get_items(), 
                self.pln.value(), 
                self.fifo_cost,
                self.pln.value() - self.fifo_cost
            )
            
            if invoice_path:
                QMessageBox.information(self, "Sukces", 
                                      f"SprzedaÅ¼ zostaÅ‚a dodana.\nRachunek {invoice_number} wygenerowany:\n{invoice_path}")
                self.accept()
    
    def save_sale_without_invoice(self):
        sale_id = self._save_sale()
        if sale_id:
            QMessageBox.information(self, "Sukces", "SprzedaÅ¼ zostaÅ‚a dodana (bez rachunku).")
            self.accept()
    
    def _save_sale(self):
        items = self.get_items()
        if not items:
            QMessageBox.warning(self, "Brak pozycji", "Dodaj przynajmniej jednÄ… pozycjÄ™ do sprzedaÅ¼y.")
            return None
            
        for pid, qty in items:
            if not self.db.check_stock(pid, qty):
                product_info = self.db.get_product_info(pid)
                if product_info:
                    QMessageBox.warning(self, "Brak stanu", 
                        f"Brak wystarczajÄ…cego stanu dla produktu:\n"
                        f"SKU: {product_info['sku']}\n"
                        f"Nazwa: {product_info['title']}\n"
                        f"Wymagane: {qty}, dostÄ™pne: {product_info['stock']}")
                return None
        
        fifo_cost = self.update_fifo_cost()
        
        date = self.date.date().toString("yyyy-MM-dd")
        try:
            eur_rate = get_eur_rate(date)
            eur = round(self.pln.value() / eur_rate, 2) if eur_rate else 0
        except:
            eur = 0
        
        try:
            # UÅ¼ywamy get_platform_name() zamiast currentText()
            sale_id = self.db.add_sale_order_with_reset(
                self.get_platform_name(),
                self.pln.value(),
                eur,
                date,
                items,
                fifo_cost
            )
            return sale_id
            
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", f"Nie udaÅ‚o siÄ™ dodaÄ‡ sprzedaÅ¼y:\n{str(e)}")
            return None

# ================== NOWE DIALOGI - DODANE FUNKCJE ==================

class OpenDatabaseDialog(QDialog):
    """Dialog do otwierania istniejÄ…cej bazy danych"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("OtwÃ³rz bazÄ™ danych")
        self.resize(500, 300)
        
        v = QVBoxLayout(self)
        
        # Informacje
        info_label = QLabel("Wybierz istniejÄ…cÄ… bazÄ™ danych (.db) lub utwÃ³rz nowÄ…:")
        v.addWidget(info_label)
        
        # Lista dostÄ™pnych baz
        self.db_list = QListWidget()
        self.refresh_db_list()
        v.addWidget(self.db_list)
        
        # ÅšcieÅ¼ka rÄ™czna
        path_layout = QHBoxLayout()
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("ÅšcieÅ¼ka do pliku .db...")
        btn_browse = QPushButton("PrzeglÄ…daj...")
        btn_browse.clicked.connect(self.browse_db_file)
        
        path_layout.addWidget(self.path_edit)
        path_layout.addWidget(btn_browse)
        v.addLayout(path_layout)
        
        # Przyciski akcji
        btn_layout = QHBoxLayout()
        
        btn_open = QPushButton("OtwÃ³rz")
        btn_open.clicked.connect(self.open_selected)
        btn_open.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_new = QPushButton("UtwÃ³rz nowÄ…")
        btn_new.clicked.connect(self.create_new)
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_open)
        btn_layout.addWidget(btn_new)
        btn_layout.addWidget(btn_cancel)
        
        v.addLayout(btn_layout)
    
    def refresh_db_list(self):
        """OdÅ›wieÅ¼ listÄ™ dostÄ™pnych baz danych w bieÅ¼Ä…cym folderze"""
        self.db_list.clear()
        current_dir = os.getcwd()
        
        for file in os.listdir(current_dir):
            if file.endswith('.db'):
                size = os.path.getsize(file)
                size_str = f"({size/1024:.1f} KB)" if size < 1024*1024 else f"({size/(1024*1024):.1f} MB)"
                item_text = f"{file} {size_str}"
                item = QListWidgetItem(item_text)
                item.setData(Qt.UserRole, file)
                self.db_list.addItem(item)
    
    def browse_db_file(self):
        """PrzeglÄ…daj pliki bazy danych"""
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz bazÄ™ danych", 
            os.getcwd(), 
            "Bazy danych (*.db);;Wszystkie pliki (*.*)"
        )
        if path:
            self.path_edit.setText(path)
    
    def open_selected(self):
        """OtwÃ³rz wybranÄ… bazÄ™ danych"""
        if self.db_list.currentItem():
            db_file = self.db_list.currentItem().data(Qt.UserRole)
            path = os.path.join(os.getcwd(), db_file)
        elif self.path_edit.text().strip():
            path = self.path_edit.text().strip()
        else:
            QMessageBox.warning(self, "Brak wyboru", "Wybierz bazÄ™ danych z listy lub podaj Å›cieÅ¼kÄ™.")
            return
        
        if not os.path.exists(path):
            QMessageBox.warning(self, "Brak pliku", f"Plik nie istnieje:\n{path}")
            return
        
        if not path.endswith('.db'):
            QMessageBox.warning(self, "NieprawidÅ‚owy format", "Wybierz plik z rozszerzeniem .db")
            return
        
        self.selected_path = path
        self.accept()
    
    def create_new(self):
        """UtwÃ³rz nowÄ… bazÄ™ danych"""
        path, _ = QFileDialog.getSaveFileName(
            self, "UtwÃ³rz nowÄ… bazÄ™ danych",
            os.getcwd(),
            "Bazy danych (*.db)"
        )
        
        if path:
            if not path.endswith('.db'):
                path += '.db'
            
            # UtwÃ³rz pustÄ… bazÄ™
            import shutil
            try:
                # SprawdÅº czy istnieje domyÅ›lna baza do skopiowania
                if os.path.exists('data.db'):
                    shutil.copy('data.db', path)
                else:
                    # UtwÃ³rz pustÄ… bazÄ™ przez wywoÅ‚anie DB
                    from db import DB
                    temp_db = DB(path)
                    temp_db.conn.close()
                
                QMessageBox.information(self, "Sukces", f"Utworzono nowÄ… bazÄ™ danych:\n{path}")
                self.selected_path = path
                self.accept()
            except Exception as e:
                QMessageBox.critical(self, "BÅ‚Ä…d", f"Nie udaÅ‚o siÄ™ utworzyÄ‡ bazy:\n{str(e)}")

class BackupDialog(QDialog):
    """Dialog do tworzenia kopii zapasowej bazy danych"""
    def __init__(self, current_db_path, parent=None):
        super().__init__(parent)
        self.current_db_path = current_db_path
        self.setWindowTitle("Archiwizuj bazÄ™ danych")
        self.resize(500, 300)
        
        v = QVBoxLayout(self)
        
        # Informacje o aktualnej bazie
        info_label = QLabel(f"Aktualna baza danych: {os.path.basename(current_db_path)}")
        info_label.setWordWrap(True)
        v.addWidget(info_label)
        
        # Opcje archiwizacji
        group = QGroupBox("Opcje archiwizacji")
        group_layout = QVBoxLayout()
        
        self.rb_default = QRadioButton("Archiwizuj do folderu 'backups' (automatyczna nazwa)")
        self.rb_default.setChecked(True)
        group_layout.addWidget(self.rb_default)
        
        self.rb_custom = QRadioButton("Zapisz kopiÄ™ jako... (wybierz lokalizacjÄ™ rÄ™cznie)")
        group_layout.addWidget(self.rb_custom)
        
        self.custom_path_edit = QLineEdit()
        self.custom_path_edit.setPlaceholderText("ÅšcieÅ¼ka do zapisania kopii...")
        self.custom_path_edit.setEnabled(False)
        
        btn_browse = QPushButton("Wybierz...")
        btn_browse.clicked.connect(self.browse_backup_path)
        btn_browse.setEnabled(False)
        
        path_layout = QHBoxLayout()
        path_layout.addWidget(self.custom_path_edit)
        path_layout.addWidget(btn_browse)
        group_layout.addLayout(path_layout)
        
        self.rb_default.toggled.connect(lambda: self.update_path_enabled())
        self.rb_custom.toggled.connect(lambda: self.update_path_enabled())
        
        group.setLayout(group_layout)
        v.addWidget(group)
        
        # Kompresja
        self.cb_compress = QCheckBox("Skompresuj kopiÄ™ (format .zip)")
        self.cb_compress.setChecked(True)
        v.addWidget(self.cb_compress)
        
        # Dodatkowe opcje
        self.cb_backup_config = QCheckBox("UwzglÄ™dnij plik konfiguracyjny (config.json)")
        self.cb_backup_config.setChecked(True)
        v.addWidget(self.cb_backup_config)
        
        self.cb_backup_invoices = QCheckBox("UwzglÄ™dnij folder z rachunkami (rachunki/)")
        self.cb_backup_invoices.setChecked(True)
        v.addWidget(self.cb_backup_invoices)
        
        v.addStretch()
        
        # Przyciski
        btn_layout = QHBoxLayout()
        
        btn_backup = QPushButton("UtwÃ³rz kopiÄ™ zapasowÄ…")
        btn_backup.clicked.connect(self.create_backup)
        btn_backup.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_backup)
        btn_layout.addWidget(btn_cancel)
        
        v.addLayout(btn_layout)
    
    def update_path_enabled(self):
        """WÅ‚Ä…cz/wyÅ‚Ä…cz pola edycji Å›cieÅ¼ki"""
        enabled = self.rb_custom.isChecked()
        self.custom_path_edit.setEnabled(enabled)
        # ZnajdÅº i wÅ‚Ä…cz przycisk przeglÄ…dania
        for i in range(self.layout().count()):
            widget = self.layout().itemAt(i).widget()
            if isinstance(widget, QGroupBox):
                for j in range(widget.layout().count()):
                    item = widget.layout().itemAt(j)
                    if item and item.layout():
                        for k in range(item.layout().count()):
                            sub_widget = item.layout().itemAt(k).widget()
                            if isinstance(sub_widget, QPushButton) and sub_widget.text() == "Wybierz...":
                                sub_widget.setEnabled(enabled)
    
    def browse_backup_path(self):
        """Wybierz Å›cieÅ¼ka do zapisania kopii"""
        default_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        path, _ = QFileDialog.getSaveFileName(
            self, "Zapisz kopiÄ™ zapasowÄ…",
            os.path.join(os.getcwd(), default_name),
            "Bazy danych (*.db);;Wszystkie pliki (*.*)"
        )
        if path:
            self.custom_path_edit.setText(path)
    
    def create_backup(self):
        """UtwÃ³rz kopiÄ™ zapasowÄ…"""
        try:
            import shutil
            import zipfile
            
            if self.rb_default.isChecked():
                # UtwÃ³rz folder backups jeÅ›li nie istnieje
                backup_dir = os.path.join(os.getcwd(), "backups")
                os.makedirs(backup_dir, exist_ok=True)
                
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                backup_name = f"backup_{timestamp}.db"
                backup_path = os.path.join(backup_dir, backup_name)
                
                # Skopiuj bazÄ™ danych
                shutil.copy2(self.current_db_path, backup_path)
                
                if self.cb_compress.isChecked():
                    zip_path = backup_path.replace('.db', '.zip')
                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(backup_path, os.path.basename(backup_path))
                        
                        if self.cb_backup_config.isChecked() and os.path.exists('config.json'):
                            zipf.write('config.json', 'config.json')
                        
                        if self.cb_backup_invoices.isChecked() and os.path.exists('rachunki'):
                            for root, dirs, files in os.walk('rachunki'):
                                for file in files:
                                    file_path = os.path.join(root, file)
                                    arcname = os.path.relpath(file_path, start='.')
                                    zipf.write(file_path, arcname)
                    
                    os.remove(backup_path)
                    backup_path = zip_path
                
                QMessageBox.information(self, "Sukces", 
                    f"Utworzono kopiÄ™ zapasowÄ…:\n{backup_path}")
                
            else:
                backup_path = self.custom_path_edit.text().strip()
                if not backup_path:
                    QMessageBox.warning(self, "Brak Å›cieÅ¼ki", "Podaj Å›cieÅ¼kÄ™ do zapisania kopii.")
                    return
                
                shutil.copy2(self.current_db_path, backup_path)
                
                if self.cb_compress.isChecked():
                    zip_path = backup_path.replace('.db', '.zip')
                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        zipf.write(backup_path, os.path.basename(backup_path))
                    
                    os.remove(backup_path)
                    backup_path = zip_path
                
                QMessageBox.information(self, "Sukces", 
                    f"Utworzono kopiÄ™ zapasowÄ…:\n{backup_path}")
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", 
                f"Nie udaÅ‚o siÄ™ utworzyÄ‡ kopii zapasowej:\n{str(e)}")

class ImportExportDialog(QDialog):
    """Dialog do importu/eksportu danych"""
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db = db
        self.config = config
        self.setWindowTitle("Import/Export danych")
        self.resize(600, 500)
        
        v = QVBoxLayout(self)
        
        # Sekcja importu
        import_group = QGroupBox("Import danych")
        import_layout = QVBoxLayout()
        
        self.import_combo = QComboBox()
        self.import_combo.addItems([
            "Wybierz typ importu...",
            "Towary z pliku CSV/Excel",
            "Cennik dostawcy z CSV/Excel",
            "Transakcje bankowe z CSV",
            "Rachunki z plikÃ³w PDF"
        ])
        import_layout.addWidget(self.import_combo)
        
        self.import_widget = QStackedWidget()
        
        # Widget dla importu towarÃ³w
        import_products_widget = QWidget()
        import_products_layout = QVBoxLayout()
        import_products_layout.addWidget(QLabel("Format pliku: SKU;Nazwa;Stan poczÄ…tkowy;Cena zakupu"))
        import_products_layout.addWidget(QLabel("WskazÃ³wka: UÅ¼yj CSV z separatorem Å›rednika"))
        self.import_products_btn = QPushButton("Wybierz plik do importu...")
        self.import_products_btn.clicked.connect(lambda: self.import_products())
        import_products_layout.addWidget(self.import_products_btn)
        import_products_widget.setLayout(import_products_layout)
        self.import_widget.addWidget(import_products_widget)
        
        # Widget dla importu cennika
        import_prices_widget = QWidget()
        import_prices_layout = QVBoxLayout()
        import_prices_layout.addWidget(QLabel("Format pliku: SKU;Cena zakupu;Data waÅ¼noÅ›ci"))
        self.import_prices_btn = QPushButton("Wybierz plik cennika...")
        import_prices_layout.addWidget(self.import_prices_btn)
        import_prices_widget.setLayout(import_prices_layout)
        self.import_widget.addWidget(import_prices_widget)
        
        # Pusty widget dla pozostaÅ‚ych opcji
        for _ in range(3):
            empty_widget = QWidget()
            empty_layout = QVBoxLayout()
            empty_layout.addWidget(QLabel("Funkcja w przygotowaniu..."))
            empty_widget.setLayout(empty_layout)
            self.import_widget.addWidget(empty_widget)
        
        import_layout.addWidget(self.import_widget)
        import_group.setLayout(import_layout)
        v.addWidget(import_group)
        
        # Sekcja eksportu
        export_group = QGroupBox("Eksport danych")
        export_layout = QVBoxLayout()
        
        self.export_combo = QComboBox()
        self.export_combo.addItems([
            "Wybierz typ eksportu...",
            "Lista produktÃ³w do PDF",
            "Raport magazynowy do Excel",
            "Lista faktur do CSV",
            "Ewidencja sprzedaÅ¼y do PDF"
        ])
        export_layout.addWidget(self.export_combo)
        
        self.export_widget = QStackedWidget()
        
        # Widget dla eksportu produktÃ³w
        export_products_widget = QWidget()
        export_products_layout = QVBoxLayout()
        export_products_layout.addWidget(QLabel("Wygeneruj listÄ™ produktÃ³w w formacie PDF"))
        self.export_products_btn = QPushButton("Eksportuj produkty do PDF...")
        self.export_products_btn.clicked.connect(lambda: self.export_products_pdf())
        export_products_layout.addWidget(self.export_products_btn)
        export_products_widget.setLayout(export_products_layout)
        self.export_widget.addWidget(export_products_widget)
        
        # Widget dla eksportu magazynu
        export_inventory_widget = QWidget()
        export_inventory_layout = QVBoxLayout()
        export_inventory_layout.addWidget(QLabel("Wygeneruj peÅ‚ny raport magazynowy"))
        self.export_inventory_btn = QPushButton("Eksportuj raport do Excel...")
        export_inventory_layout.addWidget(self.export_inventory_btn)
        export_inventory_widget.setLayout(export_inventory_layout)
        self.export_widget.addWidget(export_inventory_widget)
        
        # Pusty widget dla pozostaÅ‚ych opcji
        for _ in range(2):
            empty_widget = QWidget()
            empty_layout = QVBoxLayout()
            empty_layout.addWidget(QLabel("Funkcja w przygotowaniu..."))
            empty_widget.setLayout(empty_layout)
            self.export_widget.addWidget(empty_widget)
        
        export_layout.addWidget(self.export_widget)
        export_group.setLayout(export_layout)
        v.addWidget(export_group)
        
        # Przyciski
        btn_layout = QHBoxLayout()
        btn_close = QPushButton("Zamknij")
        btn_close.clicked.connect(self.accept)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        
        v.addLayout(btn_layout)
        
        # PodÅ‚Ä…czenie sygnaÅ‚Ã³w
        self.import_combo.currentIndexChanged.connect(self.import_widget.setCurrentIndex)
        self.export_combo.currentIndexChanged.connect(self.export_widget.setCurrentIndex)
    
    def import_products(self):
        """Import produktÃ³w z pliku CSV"""
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik do importu",
            os.getcwd(),
            "Pliki CSV (*.csv);;Pliki Excel (*.xlsx *.xls);;Wszystkie pliki (*.*)"
        )
        
        if not path:
            return
        
        try:
            imported = 0
            skipped = 0
            errors = []
            
            if path.endswith('.csv'):
                with open(path, 'r', encoding='utf-8') as f:
                    import csv
                    reader = csv.reader(f, delimiter=';')
                    
                    for row_num, row in enumerate(reader, 1):
                        if len(row) < 2:
                            continue
                        
                        sku = row[0].strip()
                        title = row[1].strip()
                        
                        if not sku or not title:
                            errors.append(f"Wiersz {row_num}: Brak SKU lub nazwy")
                            continue
                        
                        if self.db.check_sku_exists(sku):
                            skipped += 1
                            continue
                        
                        try:
                            self.db.add_product(sku, title)
                            imported += 1
                            
                            # JeÅ›li jest stan poczÄ…tkowy
                            if len(row) >= 3 and row[2].strip():
                                try:
                                    stock = int(row[2].strip())
                                    if stock > 0:
                                        pid = self.db.get_product_id_by_sku(sku)
                                        if pid:
                                            self.db.update_stock(pid, stock)
                                except ValueError:
                                    pass
                            
                        except Exception as e:
                            errors.append(f"Wiersz {row_num}: {str(e)}")
            
            QMessageBox.information(self, "Import zakoÅ„czony",
                f"Zaimportowano: {imported} produktÃ³w\n"
                f"PominiÄ™to (juÅ¼ istniejÄ…): {skipped}\n"
                f"BÅ‚Ä™dy: {len(errors)}")
            
            if errors:
                QMessageBox.warning(self, "BÅ‚Ä™dy importu", 
                    "NiektÃ³re wiersze nie zostaÅ‚y zaimportowane:\n\n" + 
                    "\n".join(errors[:10]) + 
                    ("\n\n... i wiÄ™cej" if len(errors) > 10 else ""))
        
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d importu", 
                f"Nie udaÅ‚o siÄ™ zaimportowaÄ‡ pliku:\n{str(e)}")
    
    def export_products_pdf(self):
        """Eksport listy produktÃ³w do PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            
            path, _ = QFileDialog.getSaveFileName(
                self, "Zapisz listÄ™ produktÃ³w",
                os.path.join(os.getcwd(), f"produkty_{datetime.now().strftime('%Y%m%d')}.pdf"),
                "Pliki PDF (*.pdf)"
            )
            
            if not path:
                return
            
            products = self.db.list_products()
            
            doc = SimpleDocTemplate(path, pagesize=A4,
                                  rightMargin=2*cm, leftMargin=2*cm,
                                  topMargin=2*cm, bottomMargin=2*cm)
            
            story = []
            styles = getSampleStyleSheet()
            
            story.append(Paragraph("LISTA PRODUKTÃ“W W MAGAZYNIE", 
                                 styles['Title']))
            story.append(Spacer(1, 20))
            story.append(Paragraph(f"Data wygenerowania: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 
                                 styles['Normal']))
            story.append(Spacer(1, 30))
            
            # Tabela z produktami
            data = [["SKU", "Nazwa", "Stan"]]
            
            for p in products:
                data.append([p['sku'], p['title'], str(p['stock'])])
            
            table = Table(data, colWidths=[4*cm, 10*cm, 2*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (2, 1), (2, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke])
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            story.append(Paragraph(f"ÅÄ…cznie produktÃ³w: {len(products)}", styles['Normal']))
            
            doc.build(story)
            
            QMessageBox.information(self, "Sukces", 
                f"Lista produktÃ³w zostaÅ‚a wyeksportowana:\n{path}")
        
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d eksportu", 
                f"Nie udaÅ‚o siÄ™ wyeksportowaÄ‡ do PDF:\n{str(e)}")

class PrintDialog(QDialog):
    """Dialog do drukowania"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Drukowanie")
        self.resize(500, 400)
        
        v = QVBoxLayout(self)
        
        # WybÃ³r drukarki
        printer_group = QGroupBox("WybÃ³r drukarki")
        printer_layout = QVBoxLayout()
        
        self.printer_combo = QComboBox()
        self.refresh_printers()
        printer_layout.addWidget(QLabel("Drukarka:"))
        printer_layout.addWidget(self.printer_combo)
        
        btn_refresh = QPushButton("OdÅ›wieÅ¼ listÄ™")
        btn_refresh.clicked.connect(self.refresh_printers)
        printer_layout.addWidget(btn_refresh)
        
        btn_properties = QPushButton("WÅ‚aÅ›ciwoÅ›ci drukarki...")
        btn_properties.clicked.connect(self.show_printer_properties)
        printer_layout.addWidget(btn_properties)
        
        printer_group.setLayout(printer_layout)
        v.addWidget(printer_group)
        
        # Ustawienia druku
        settings_group = QGroupBox("Ustawienia druku")
        settings_layout = QVBoxLayout()
        
        # Zakres stron
        pages_layout = QHBoxLayout()
        pages_layout.addWidget(QLabel("Zakres stron:"))
        
        self.rb_all = QRadioButton("Wszystkie")
        self.rb_all.setChecked(True)
        pages_layout.addWidget(self.rb_all)
        
        self.rb_range = QRadioButton("Strony:")
        pages_layout.addWidget(self.rb_range)
        
        self.page_from = QSpinBox()
        self.page_from.setRange(1, 999)
        self.page_from.setValue(1)
        self.page_from.setEnabled(False)
        pages_layout.addWidget(self.page_from)
        
        pages_layout.addWidget(QLabel("do"))
        
        self.page_to = QSpinBox()
        self.page_to.setRange(1, 999)
        self.page_to.setValue(1)
        self.page_to.setEnabled(False)
        pages_layout.addWidget(self.page_to)
        
        pages_layout.addStretch()
        
        self.rb_all.toggled.connect(self.update_page_range_enabled)
        self.rb_range.toggled.connect(self.update_page_range_enabled)
        
        settings_layout.addLayout(pages_layout)
        
        # Liczba kopii
        copies_layout = QHBoxLayout()
        copies_layout.addWidget(QLabel("Liczba kopii:"))
        
        self.copies_spin = QSpinBox()
        self.copies_spin.setRange(1, 99)
        self.copies_spin.setValue(1)
        copies_layout.addWidget(self.copies_spin)
        
        copies_layout.addStretch()
        settings_layout.addLayout(copies_layout)
        
        # Orientacja
        orientation_layout = QHBoxLayout()
        orientation_layout.addWidget(QLabel("Orientacja:"))
        
        self.rb_portrait = QRadioButton("Pionowa")
        self.rb_portrait.setChecked(True)
        orientation_layout.addWidget(self.rb_portrait)
        
        self.rb_landscape = QRadioButton("Pozioma")
        orientation_layout.addWidget(self.rb_landscape)
        
        orientation_layout.addStretch()
        settings_layout.addLayout(orientation_layout)
        
        # Zaawansowane
        self.cb_color = QCheckBox("Druk kolorowy")
        self.cb_color.setChecked(True)
        settings_layout.addWidget(self.cb_color)
        
        self.cb_duplex = QCheckBox("Druk dwustronny")
        settings_layout.addWidget(self.cb_duplex)
        
        settings_group.setLayout(settings_layout)
        v.addWidget(settings_group)
        
        # PodglÄ…d
        preview_group = QGroupBox("PodglÄ…d wydruku")
        preview_layout = QVBoxLayout()
        
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setMaximumHeight(100)
        self.preview_text.setPlainText("PrzykÅ‚adowy podglÄ…d wydruku...\nTutaj bÄ™dzie wyÅ›wietlana zawartoÅ›Ä‡ dokumentu przed drukiem.")
        preview_layout.addWidget(self.preview_text)
        
        preview_group.setLayout(preview_layout)
        v.addWidget(preview_group)
        
        # Przyciski
        btn_layout = QHBoxLayout()
        
        btn_print = QPushButton("Drukuj")
        btn_print.clicked.connect(self.print_document)
        btn_print.setStyleSheet("background-color: #2E7D32; font-weight: bold;")
        
        btn_preview = QPushButton("PodglÄ…d wydruku...")
        btn_preview.clicked.connect(self.print_preview)
        
        btn_cancel = QPushButton("Anuluj")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_print)
        btn_layout.addWidget(btn_preview)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        
        v.addLayout(btn_layout)
    
    def refresh_printers(self):
        """OdÅ›wieÅ¼ listÄ™ dostÄ™pnych drukarek"""
        from PySide6.QtPrintSupport import QPrinterInfo
        self.printer_combo.clear()
        
        printers = QPrinterInfo.availablePrinters()
        for printer in printers:
            self.printer_combo.addItem(printer.printerName(), printer)
    
    def update_page_range_enabled(self):
        """WÅ‚Ä…cz/wyÅ‚Ä…cz pola zakresu stron"""
        enabled = self.rb_range.isChecked()
        self.page_from.setEnabled(enabled)
        self.page_to.setEnabled(enabled)
    
    def show_printer_properties(self):
        """PokaÅ¼ wÅ‚aÅ›ciwoÅ›ci wybranej drukarki"""
        if self.printer_combo.currentIndex() < 0:
            QMessageBox.warning(self, "Brak drukarki", "Nie wybrano drukarki.")
            return
        
        printer_info = self.printer_combo.currentData()
        printer_name = printer_info.printerName()
        
        QMessageBox.information(self, "WÅ‚aÅ›ciwoÅ›ci drukarki",
            f"Drukarka: {printer_name}\n\n"
            f"Stan: {'Gotowa' if printer_info.isDefault() else 'DostÄ™pna'}\n"
            f"ObsÅ‚uguje kolor: {'Tak' if printer_info.supportsCustomPageSizes() else 'Nie'}\n"
            f"ObsÅ‚uguje duplex: {'Tak' if printer_info.supportsMultipleCopies() else 'Nie'}")
    
    def print_document(self):
        """Wydrukuj dokument"""
        if self.printer_combo.currentIndex() < 0:
            QMessageBox.warning(self, "Brak drukarki", "Wybierz drukarkÄ™ z listy.")
            return
        
        printer_info = self.printer_combo.currentData()
        
        # Tutaj moÅ¼na dodaÄ‡ logikÄ™ drukowania konkretnego dokumentu
        QMessageBox.information(self, "Drukowanie",
            f"RozpoczÄ™to drukowanie na drukarce:\n{printer_info.printerName()}\n\n"
            f"Kopie: {self.copies_spin.value()}\n"
            f"Orientacja: {'Pionowa' if self.rb_portrait.isChecked() else 'Pozioma'}\n"
            f"Zakres stron: {'Wszystkie' if self.rb_all.isChecked() else f'{self.page_from.value()}-{self.page_to.value()}'}")
        
        self.accept()
    
    def print_preview(self):
        """PokaÅ¼ podglÄ…d wydruku"""
        QMessageBox.information(self, "PodglÄ…d wydruku",
            "Ta funkcja jest w przygotowaniu.\n"
            "W przyszÅ‚oÅ›ci bÄ™dzie tutaj wyÅ›wietlany podglÄ…d dokumentu przed drukiem.")

# ================== MAIN (Z MENU RIBBON - POPRAWIONYM) ==================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = Config()
        
        # Wczytaj Å›cieÅ¼kÄ™ bazy z konfiguracji
        self.db_path = self.config.get_database_path()
        self.db = DB(self.db_path)
        
        # Ustaw tytuÅ‚ okna z nazwÄ… bazy
        db_info = self.config.get_database_info()
        self.setWindowTitle(f"System Magazynowo-SprzedaÅ¼owy v{APP_VERSION} - {db_info['filename']}")
        self.resize(1200, 700)

        self.setup_ui()
        self.setup_menu_bar()
        self.setup_toolbar()
        
        self.refresh()

    def setup_ui(self):
        """Konfiguruje gÅ‚Ã³wny interfejs"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        # Tabela produktÃ³w
        self.table = SortableTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["ID", "SKU", "Nazwa", "Stan", ""])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        
        main_layout.addWidget(self.table)
        
        # Pasek statusu z informacjÄ… o bazie danych
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.update_status_bar()

    def update_status_bar(self):
        """Aktualizuje pasek statusu z informacjÄ… o bazie danych"""
        db_info = self.config.get_database_info()
        status_text = f"Baza danych: {db_info['filename']}"
        
        if db_info['last_opened']:
            status_text += f" | Ostatnio otwarta: {db_info['last_opened']}"
        
        self.status_bar.showMessage(status_text)

    def setup_menu_bar(self):
        """Konfiguruje menu ribbon - POPRAWIONE DO POZIOMEGO"""
        menu_bar = self.menuBar()
        
        # ========== MENU PLIK ==========
        file_menu = menu_bar.addMenu("&Plik")
        
        # Nowe funkcje
        open_last_action = QAction("&OtwÃ³rz ostatniÄ… bazÄ™", self)
        open_last_action.triggered.connect(self.open_last_database)
        open_last_action.setToolTip("OtwÃ³rz ostatnio uÅ¼ywanÄ… bazÄ™ danych")
        file_menu.addAction(open_last_action)
        
        open_db_action = QAction("&OtwÃ³rz bazÄ™ danych...", self)
        open_db_action.triggered.connect(self.open_database)
        file_menu.addAction(open_db_action)
        
        backup_action = QAction("&Archiwizuj bazÄ™ danych...", self)
        backup_action.triggered.connect(self.backup_database)
        file_menu.addAction(backup_action)
        
        file_menu.addSeparator()
        
        # Import/Export
        import_export_action = QAction("&Import/Export...", self)
        import_export_action.triggered.connect(self.import_export)
        file_menu.addAction(import_export_action)
        
        file_menu.addSeparator()
        
        # Drukowanie
        print_action = QAction("&Drukowanie...", self)
        print_action.setShortcut("Ctrl+P")
        print_action.triggered.connect(self.print_dialog)
        file_menu.addAction(print_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("&ZakoÅ„cz", self)
        exit_action.setShortcut(QKeySequence.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # ========== MENU MAGAZYN ==========
        warehouse_menu = menu_bar.addMenu("&Magazyn")
        
        add_product_action = QAction("&Dodaj produkt...", self)
        add_product_action.setShortcut("Ctrl+N")
        add_product_action.triggered.connect(self.add_product)
        warehouse_menu.addAction(add_product_action)
        
        delete_product_action = QAction("&UsuÅ„ produkt...", self)
        delete_product_action.setShortcut("Ctrl+Shift+P")
        delete_product_action.triggered.connect(self.delete_product)
        warehouse_menu.addAction(delete_product_action)
        
        warehouse_menu.addSeparator()
        
        add_purchase_action = QAction("&Dodaj zakup...", self)
        add_purchase_action.setShortcut("Ctrl+Z")
        add_purchase_action.triggered.connect(self.add_purchase)
        warehouse_menu.addAction(add_purchase_action)
        
        view_purchases_action = QAction("&Historia zakupÃ³w...", self)
        view_purchases_action.setShortcut("Ctrl+Shift+Z")
        view_purchases_action.triggered.connect(self.show_purchases)
        warehouse_menu.addAction(view_purchases_action)
        
        warehouse_menu.addSeparator()
        
        inventory_action = QAction("&Inwentaryzacja magazynu", self)
        inventory_action.triggered.connect(self.inventory)
        warehouse_menu.addAction(inventory_action)
        
        # ========== MENU SPRZEDAÅ» ==========
        sales_menu = menu_bar.addMenu("&SprzedaÅ¼")
        
        add_sale_action = QAction("&Dodaj sprzedaÅ¼...", self)
        add_sale_action.setShortcut("Ctrl+S")
        add_sale_action.triggered.connect(self.add_sale)
        sales_menu.addAction(add_sale_action)
        
        view_sales_action = QAction("&Historia sprzedaÅ¼y...", self)
        view_sales_action.setShortcut("Ctrl+Shift+S")
        view_sales_action.triggered.connect(self.show_sales)
        sales_menu.addAction(view_sales_action)
        
        sales_menu.addSeparator()
        
        view_invoices_action = QAction("&Historia rachunkÃ³w...", self)
        view_invoices_action.triggered.connect(self.show_invoices)
        sales_menu.addAction(view_invoices_action)
        
        # ========== MENU RAPORTY ==========
        reports_menu = menu_bar.addMenu("&Raporty")
        
        monthly_report_action = QAction("&Raport miesiÄ™czny...", self)
        monthly_report_action.triggered.connect(lambda: self.show_report_dialog("monthly"))
        reports_menu.addAction(monthly_report_action)
        
        quarterly_report_action = QAction("&Raport kwartalny...", self)
        quarterly_report_action.triggered.connect(lambda: self.show_report_dialog("quarterly"))
        reports_menu.addAction(quarterly_report_action)
        
        yearly_report_action = QAction("&Raport roczny...", self)
        yearly_report_action.triggered.connect(lambda: self.show_report_dialog("yearly"))
        reports_menu.addAction(yearly_report_action)
        
        custom_report_action = QAction("&Raport okresowy...", self)
        custom_report_action.triggered.connect(lambda: self.show_report_dialog("custom"))
        reports_menu.addAction(custom_report_action)
        
        # ========== MENU KONFIGURACJA ==========
        config_menu = menu_bar.addMenu("&Konfiguracja")
        
        business_info_action = QAction("&Dane osobiste...", self)
        business_info_action.triggered.connect(self.business_info)
        config_menu.addAction(business_info_action)
        
        invoice_config_action = QAction("&Konfiguracja rachunkÃ³w...", self)
        invoice_config_action.triggered.connect(self.invoice_config)
        config_menu.addAction(invoice_config_action)
        
        # NOWE: Submenu Konfiguracja dodatkowe
        config_submenu = QMenu("&Dodatkowe", self)
        
        limits_config_action = QAction("&Limity US...", self)
        limits_config_action.triggered.connect(self.limits_config)
        config_submenu.addAction(limits_config_action)
        
        config_menu.addMenu(config_submenu)
        
        # ========== MENU POMOC ==========
        help_menu = menu_bar.addMenu("&Pomoc")
        
        refresh_action = QAction("&OdÅ›wieÅ¼", self)
        refresh_action.setShortcut("F5")
        refresh_action.triggered.connect(self.refresh)
        help_menu.addAction(refresh_action)
        
        help_menu.addSeparator()
        
        about_action = QAction("&O programie...", self)
        about_action.triggered.connect(self.about)
        help_menu.addAction(about_action)

    def setup_toolbar(self):
        """Konfiguruje pasek narzÄ™dzi - UPROSZCZONY"""
        toolbar = QToolBar("GÅ‚Ã³wne narzÄ™dzia")
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(24, 24))
        self.addToolBar(toolbar)
        
        # Przycisk dodaj produkt
        add_product_btn = QAction("âž• Dodaj produkt", self)
        add_product_btn.triggered.connect(self.add_product)
        toolbar.addAction(add_product_btn)
        
        toolbar.addSeparator()
        
        # Przycisk dodaj zakup
        add_purchase_btn = QAction("ðŸ“¦ Dodaj zakup", self)
        add_purchase_btn.triggered.connect(self.add_purchase)
        toolbar.addAction(add_purchase_btn)
        
        toolbar.addSeparator()
        
        # Przycisk dodaj sprzedaÅ¼
        add_sale_btn = QAction("ðŸ’° Dodaj sprzedaÅ¼", self)
        add_sale_btn.triggered.connect(self.add_sale)
        toolbar.addAction(add_sale_btn)
        
        toolbar.addSeparator()
        
        # Przycisk odÅ›wieÅ¼
        refresh_btn = QAction("âŸ³ OdÅ›wieÅ¼", self)
        refresh_btn.triggered.connect(self.refresh)
        toolbar.addAction(refresh_btn)
        
        toolbar.addSeparator()
        
        # Przycisk raportu miesiÄ™cznego
        report_btn = QAction("ðŸ“Š Raport miesiÄ™czny", self)
        report_btn.triggered.connect(lambda: self.show_report_dialog("monthly"))
        toolbar.addAction(report_btn)

    def open_last_database(self):
        """Otwiera ostatnio uÅ¼ywanÄ… bazÄ™ danych"""
        db_info = self.config.get_database_info()
        
        if db_info["exists"]:
            try:
                # Zamknij aktualnÄ… bazÄ™
                if hasattr(self.db, 'conn'):
                    self.db.conn.close()
                
                # OtwÃ³rz ostatniÄ… bazÄ™
                self.db_path = db_info["path"]
                self.db = DB(self.db_path)
                
                # OdÅ›wieÅ¼ interfejs
                self.refresh()
                
                # Zaktualizuj tytuÅ‚
                self.setWindowTitle(f"System Magazynowo-SprzedaÅ¼owy v{APP_VERSION} - {db_info['filename']}")
                self.update_status_bar()
                
                QMessageBox.information(self, "Sukces", 
                    f"ZaÅ‚adowano ostatniÄ… bazÄ™ danych:\n{self.db_path}")
                    
            except Exception as e:
                QMessageBox.critical(self, "BÅ‚Ä…d", 
                    f"Nie udaÅ‚o siÄ™ zaÅ‚adowaÄ‡ bazy danych:\n{str(e)}")
        else:
            QMessageBox.warning(self, "Brak pliku", 
                f"Ostatnia baza danych nie istnieje:\n{db_info['path']}")

    def open_database(self):
        """OtwÃ³rz istniejÄ…cÄ… bazÄ™ danych"""
        dialog = OpenDatabaseDialog(self)
        if dialog.exec():
            try:
                new_db_path = dialog.selected_path
                
                # Zamknij aktualnÄ… bazÄ™
                if hasattr(self.db, 'conn'):
                    self.db.conn.close()
                
                # Zapisz nowÄ… Å›cieÅ¼kÄ™ w konfiguracji
                self.config.set_database_path(new_db_path)
                self.db_path = new_db_path
                
                # OtwÃ³rz nowÄ… bazÄ™
                self.db = DB(self.db_path)
                
                # OdÅ›wieÅ¼ interfejs
                self.refresh()
                
                # Zaktualizuj tytuÅ‚ okna
                db_info = self.config.get_database_info()
                self.setWindowTitle(f"System Magazynowo-SprzedaÅ¼owy v{APP_VERSION} - {db_info['filename']}")
                
                # Zaktualizuj pasek statusu
                self.update_status_bar()
                
                QMessageBox.information(self, "Sukces", 
                    f"ZaÅ‚adowano bazÄ™ danych:\n{self.db_path}")
                
            except Exception as e:
                QMessageBox.critical(self, "BÅ‚Ä…d", 
                    f"Nie udaÅ‚o siÄ™ zaÅ‚adowaÄ‡ bazy danych:\n{str(e)}")
                
                # PrzywrÃ³Ä‡ poprzedniÄ… bazÄ™
                self.db_path = self.config.get_database_path()
                self.db = DB(self.db_path)
                self.refresh()

    def backup_database(self):
        """UtwÃ³rz kopiÄ™ zapasowÄ… bazy danych"""
        dialog = BackupDialog(self.db_path, self)
        dialog.exec()

    def import_export(self):
        """OtwÃ³rz dialog importu/eksportu"""
        dialog = ImportExportDialog(self.db, self.config, self)
        dialog.exec()

    def print_dialog(self):
        """OtwÃ³rz dialog drukowania"""
        dialog = PrintDialog(self)
        dialog.exec()

    def show_report_dialog(self, report_type):
        """Pokazuje dialog raportu w zaleÅ¼noÅ›ci od typu"""
        dialog = ReportDialog(self.db, self.config, self, report_type)
        dialog.exec()

    def refresh(self):
        rows = self.db.list_products()
        data = []
        for r in rows:
            data.append([
                r["id"],
                r["sku"],
                r["title"],
                r["stock"],
                "" 
            ])
        
        current_column = self.table.current_sorted_column
        current_order = self.table.current_sort_order
        
        self.table.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, value in enumerate(row[:4]):
                item = QTableWidgetItem(str(value))
                
                if j in [0, 3]:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    try:
                        item.setData(Qt.EditRole, float(value))
                    except ValueError:
                        pass
                
                self.table.setItem(i, j, item)
            
            delete_btn = QPushButton("ðŸ—‘ï¸")
            delete_btn.setFixedWidth(30)
            delete_btn.setToolTip("UsuÅ„ produkt")
            delete_btn.clicked.connect(lambda checked, pid=row[0]: self.delete_single_product(pid))
            self.table.setCellWidget(i, 4, delete_btn)
        
        if current_column >= 0:
            self.table.sort_by_column(current_column, current_order)
            self.table.mark_sorted_column(current_column)
        
        self.status_bar.showMessage(f"ZaÅ‚adowano {len(data)} produktÃ³w | Baza: {os.path.basename(self.db_path)}")

    def add_product(self):
        dialog = AddProductDialog(self.db, self)
        if dialog.exec():
            self.refresh()

    def delete_product(self):
        products = self.db.list_products()
        if not products:
            QMessageBox.warning(self, "Brak produktÃ³w", "Brak produktÃ³w do usuniÄ™cia.")
            return
            
        items = [f"{p['id']} | {p['sku']} | {p['title']}" for p in products]
        product_str, ok = QInputDialog.getItem(self, "UsuÅ„ produkt", "Wybierz produkt:", items, 0, False)
        
        if ok and product_str:
            pid = int(product_str.split("|")[0].strip())
            self.confirm_delete_product(pid)

    def delete_single_product(self, pid):
        self.confirm_delete_product(pid)

    def confirm_delete_product(self, pid):
        product_info = self.db.get_product_info(pid)
        if not product_info:
            QMessageBox.warning(self, "BÅ‚Ä…d", "Produkt nie istnieje.")
            return
            
        if QMessageBox.question(
            self, "PotwierdÅº usuniÄ™cie",
            f"Czy na pewno usunÄ…Ä‡ produkt?\n\n"
            f"SKU: {product_info['sku']}\n"
            f"Nazwa: {product_info['title']}\n"
            f"Stan: {product_info['stock']}",
            QMessageBox.Yes | QMessageBox.No
        ) == QMessageBox.Yes:
            success = self.db.delete_product(pid)
            if success:
                self.refresh()
            else:
                QMessageBox.warning(self, "BÅ‚Ä…d", 
                    "Nie moÅ¼na usunÄ…Ä‡ produktu z dodatnim stanem magazynowym!\n"
                    "Najpierw sprzedaj lub skoryguj stan produktu.")

    def add_purchase(self):
        d = PurchaseDialog(self.db)
        if d.exec():
            self.db.add_purchase_order(
                d.cost.value(),
                d.date.date().toString("yyyy-MM-dd"),
                d.get_items()
            )
            self.refresh()

    def add_sale(self):
        d = SaleDialog(self.db, self.config)
        if d.exec():
            self.refresh()

    def show_purchases(self):
        try:
            purchases = self.db.list_purchases()
            HistoryDialog(
                "Historia zakupÃ³w",
                ["ID", "SKU", "Nazwa", "IloÅ›Ä‡", "PLN", "Data"],
                purchases,
                self.db.delete_purchase,
                self
            ).exec()
            self.refresh()
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", f"Nie moÅ¼na zaÅ‚adowaÄ‡ historii zakupÃ³w:\n{str(e)}")

    def show_sales(self):
        try:
            sales = self.db.list_sales()
            HistoryDialog(
                "Historia sprzedaÅ¼y",
                ["ID", "Platforma", "PLN", "EUR", "Koszt zakupu", "Zysk", "Data", "Pozycje"],
                sales,
                self.db.delete_sale,
                self
            ).exec()
            self.refresh()
        except Exception as e:
            QMessageBox.critical(self, "BÅ‚Ä…d", f"Nie moÅ¼na zaÅ‚adowaÄ‡ historii sprzedaÅ¼y:\n{str(e)}")

    def show_invoices(self):
        dialog = InvoicesHistoryDialog(self.db, self.config, self)
        dialog.exec()

    def invoice_config(self):
        """Konfiguracja ustawieÅ„ rachunkÃ³w"""
        dialog = InvoiceConfigDialog(self.config, self)
        dialog.exec()

    def limits_config(self):
        """Konfiguracja limitÃ³w US"""
        dialog = LimitsConfigDialog(self.config, self)
        dialog.exec()

    def inventory(self):
        InventoryDialog(self.db, self).exec()
        self.refresh()

    def business_info(self):
        dialog = BusinessInfoDialog(self.config, self)
        dialog.exec()

    def about(self):
        """WyÅ›wietla informacje o wersji programu"""
        try:
            from version import get_version
            version_info = get_version()
            
            about_text = f"""
            <h2>{version_info['app_name']}</h2>
            <p><b>Wersja:</b> {version_info['version']}</p>
            <p><b>Data budowy:</b> {version_info['build_date']}</p>
            <p><b>Autor:</b> {version_info['author']}</p>
            <p><b>Licencja:</b> {version_info['license']}</p>
            
            <h3>Funkcje w tej wersji:</h3>
            <ul>
                <li>Historia rachunkÃ³w z moÅ¼liwoÅ›ciÄ… usuwania</li>
                <li>Poprawione polskie znaki w PDF</li>
                <li>Inteligentne liczenie transakcji</li>
                <li>PeÅ‚na ewidencja dla US</li>
                <li>System FIFO dla kosztÃ³w zakupu</li>
                <li>Generowanie raportÃ³w Excel/CSV/PDF</li>
                <li>Konfiguracja wyglÄ…du rachunkÃ³w</li>
                <li>Resetowanie numeracji faktur</li>
                <li>Otwieranie rÃ³Å¼nych baz danych</li>
                <li>Archiwizacja i kopia zapasowa</li>
                <li>Import/Export danych</li>
                <li>Drukowanie dokumentÃ³w (rÃ³wnieÅ¼ bezpoÅ›rednie z raportÃ³w)</li>
                <li><b>NOWE:</b> Limity kwartalne od 2026 roku</li>
                <li><b>NOWE:</b> Raporty kwartalne z automatycznymi obliczeniami</li>
                <li><b>NOWE:</b> ZapamiÄ™tywanie ostatnio uÅ¼ywanej bazy danych</li>
                <li><b>NOWE:</b> Dodano platformÄ™ FB Marketplace</li>
                <li><b>NOWE:</b> MoÅ¼liwoÅ›Ä‡ wprowadzenia wÅ‚asnej nazwy platformy dla opcji "Inne"</li>
            </ul>
            
            <h3>Licencja GNU GPL v3.0:</h3>
            <p>Wolne oprogramowanie - moÅ¼esz uÅ¼ywaÄ‡, modyfikowaÄ‡ i rozpowszechniaÄ‡<br>
            zgodnie z warunkami licencji GNU General Public License wersja 3.</p>
            
            <p><i>Ostatnia aktualizacja: v{version_info['version']}</i></p>
            """
            
            msg = QMessageBox(self)
            msg.setWindowTitle("O programie")
            msg.setTextFormat(Qt.RichText)
            msg.setText(about_text)
            msg.setIcon(QMessageBox.Information)
            msg.exec()
            
        except ImportError:
            about_text = f"""
            <h2>System Magazynowo-SprzedaÅ¼owy</h2>
            <p><b>Wersja:</b> {APP_VERSION}</p>
            
            <h3>Funkcje:</h3>
            <ul>
                <li>ZarzÄ…dzanie magazynem i produktami</li>
                <li>Ewidencja zakupÃ³w i sprzedaÅ¼y</li>
                <li>Generowanie rachunkÃ³w uproszczonych</li>
                <li>Raporty dla dziaÅ‚alnoÅ›ci nierejestrowanej</li>
                <li>Otwieranie rÃ³Å¼nych baz danych</li>
                <li>Archiwizacja i kopia zapasowa</li>
                <li>Import/Export danych</li>
                <li>Drukowanie dokumentÃ³w (rÃ³wnieÅ¼ bezpoÅ›rednie z raportÃ³w)</li>
                <li><b>NOWE:</b> Limity kwartalne od 2026 roku</li>
                <li><b>NOWE:</b> Raporty kwartalne z automatycznymi obliczeniami</li>
                <li><b>NOWE:</b> ZapamiÄ™tywanie ostatnio uÅ¼ywanej bazy danych</li>
                <li><b>NOWE:</b> Dodano platformÄ™ FB Marketplace</li>
                <li><b>NOWE:</b> MoÅ¼liwoÅ›Ä‡ wprowadzenia wÅ‚asnej nazwy platformy dla opcji "Inne"</li>
            </ul>
            
            <p><i>Ostatnia aktualizacja: v{APP_VERSION}</i></p>
            """
            
            msg = QMessageBox(self)
            msg.setWindowTitle("O programie")
            msg.setTextFormat(Qt.RichText)
            msg.setText(about_text)
            msg.setIcon(QMessageBox.Information)
            msg.exec()
        except Exception as e:
            QMessageBox.information(self, "O programie", 
                                  f"System Magazynowo-SprzedaÅ¼owy\nWersja: {APP_VERSION}")

    def closeEvent(self, event):
        """Zapisz stan przed zamkniÄ™ciem programu"""
        try:
            if hasattr(self.db, 'conn'):
                self.db.conn.close()
        except:
            pass
        
        event.accept()

# ================== START ==================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(RED_WHITE_QSS)
    
    # SprawdÅº czy domyÅ›lna baza istnieje
    if not os.path.exists("data.db"):
        # UtwÃ³rz domyÅ›lnÄ… bazÄ™
        from db import DB
        temp_db = DB("data.db")
        temp_db.conn.close()
    
    w = MainWindow()
    w.show()
    sys.exit(app.exec())
